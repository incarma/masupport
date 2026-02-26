# django_ma/accounts/tasks.py
from __future__ import annotations

import logging
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from celery import shared_task
from django.conf import settings
from django.core.cache import cache
from django.db import transaction

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from .constants import (
    CACHE_ERROR_PREFIX,
    CACHE_PROGRESS_PREFIX,
    CACHE_RESULT_PATH_PREFIX,
    CACHE_STATUS_PREFIX,
    CACHE_TIMEOUT_SECONDS,
    cache_key,
)
from .models import CustomUser
from .services.users_excel_import import (
    REQUIRED_COLS,
    build_defaults_from_row,
    pick_worksheet_by_required_cols,
)

logger = logging.getLogger(__name__)

# NOTE:
# part(소속부서) 값 치환/정규화는 users_excel_import.py의 SSOT(build_defaults_from_row/normalize_part)에서 처리한다.


# =============================================================================
# 0) 정책/상수
# =============================================================================

# ✅ 관리자 보호(권장): 기존 이 등급은 엑셀로 grade 강등/권한 필드 덮어쓰기 방지
PROTECTED_GRADES = {"superuser", "head", "leader"}

# ✅ 보호 필드: 평상시 빈값으로 덮어쓰지 않음(단, 재직→퇴사 전환 시 초기화 허용)
PROTECTED_FIELDS = {"position", "team_a", "team_b", "team_c"}

# 결과 리포트 엑셀 시트명
RESULT_SHEET_NAME = "UploadResult"

# 진행률 표시를 위한 최소/최대 보정
PERCENT_MIN = 0
PERCENT_MAX = 100


# =============================================================================
# 1) Cache helpers (keys 단일화)
# =============================================================================

@dataclass(frozen=True)
class UploadCacheKeys:
    percent: str
    status: str
    error: str
    result_path: str


def _keys(task_id: str) -> UploadCacheKeys:
    return UploadCacheKeys(
        percent=cache_key(CACHE_PROGRESS_PREFIX, task_id),
        status=cache_key(CACHE_STATUS_PREFIX, task_id),
        error=cache_key(CACHE_ERROR_PREFIX, task_id),
        result_path=cache_key(CACHE_RESULT_PATH_PREFIX, task_id),
    )


def _cache_init(task_id: str) -> UploadCacheKeys:
    k = _keys(task_id)
    cache.set(k.status, "RUNNING", timeout=CACHE_TIMEOUT_SECONDS)
    cache.set(k.percent, 0, timeout=CACHE_TIMEOUT_SECONDS)
    cache.delete(k.error)
    cache.delete(k.result_path)
    return k


def _cache_set_percent(k: UploadCacheKeys, percent: int) -> None:
    p = max(PERCENT_MIN, min(PERCENT_MAX, int(percent)))
    cache.set(k.percent, p, timeout=CACHE_TIMEOUT_SECONDS)


def _cache_fail(k: UploadCacheKeys, err: Exception) -> None:
    cache.set(k.status, "FAILURE", timeout=CACHE_TIMEOUT_SECONDS)
    cache.set(k.error, str(err), timeout=CACHE_TIMEOUT_SECONDS)


def _cache_success(k: UploadCacheKeys, result_path: str) -> None:
    _cache_set_percent(k, 100)
    cache.set(k.status, "SUCCESS", timeout=CACHE_TIMEOUT_SECONDS)
    cache.set(k.result_path, result_path, timeout=CACHE_TIMEOUT_SECONDS)


# =============================================================================
# 2) Result dir helper
# =============================================================================

def _get_result_dir() -> Path:
    media_root = Path(getattr(settings, "MEDIA_ROOT", "media"))
    default_dir = media_root / "upload_results"
    result_dir = Path(getattr(settings, "UPLOAD_RESULT_DIR", default_dir))
    result_dir.mkdir(parents=True, exist_ok=True)
    return result_dir


# =============================================================================
# 3) 결과 리포트 엑셀 생성
# =============================================================================

def _make_result_wb(
    results: List[List[Any]],
    total: int,
    new_cnt: int,
    upd_cnt: int,
    skip_cnt: int,
    err_cnt: int,
    picked_sheet: str,
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = RESULT_SHEET_NAME

    ws.append(["Row", "사원번호", "성명", "부문", "부서", "지점", "권한(grade)", "상태", "Result"])

    fill_new = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_update = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    fill_skip = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fill_error = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row in results:
        ws.append(row)
        r = ws.max_row
        t = str(row[-1] or "")
        cell = ws[f"I{r}"]
        if "🟢" in t:
            cell.fill = fill_new
        elif "✅" in t:
            cell.fill = fill_update
        elif "⚠️" in t:
            cell.fill = fill_skip
        elif "❌" in t:
            cell.fill = fill_error

    ws.append([])
    ws.append(["선택된 시트", picked_sheet])
    ws.append(["총 데이터(행)", total])
    ws.append(["신규 추가", new_cnt])
    ws.append(["업데이트", upd_cnt])
    ws.append(["스킵", skip_cnt])
    ws.append(["오류", err_cnt])

    return wb


def _save_result_workbook(task_id: str, result_wb: Workbook) -> str:
    result_dir = _get_result_dir()
    filename = f"upload_result_{task_id}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    path = result_dir / filename
    result_wb.save(str(path))
    return str(path)


# =============================================================================
# 4) Celery Task: 영업가족직원조회 업로드/업데이트
# =============================================================================

@shared_task(bind=True)
def process_users_excel_task(self, task_id: str, file_path: str, batch_size: int = 500) -> dict:
    """
    ✅ '영업가족직원조회' 엑셀 업로드/업데이트 (SSOT)

    - 시트명 무관: REQUIRED_COLS 포함 시트 자동 탐색
    - division(총괄): 빈 문자열 저장
    - is_staff: 전체 False / is_superuser: 기본 False
    - is_active: grade != inactive
    - 관리자 보호(권장): 기존 superuser/head/leader는 grade/status/is_staff/is_superuser/is_active 덮어쓰기 금지
    - 보호필드(PROTECTED_FIELDS): 빈값으로 덮어쓰지 않음 (단, 재직→퇴사 전환 시 초기화 허용)
    - 진행률/상태/오류/결과경로: cache 기록
    - 배치 처리: batch_size 단위 transaction
    - 결과 리포트 엑셀 저장
    """
    k = _cache_init(task_id)
    logger.warning("[TASK START] tid=%s file=%s batch=%s", task_id, file_path, batch_size)

    wb = None
    try:
        # ---------------------------------------------------------------------
        # 1) Workbook open + 업로드 시트 자동 선택
        # ---------------------------------------------------------------------
        wb = load_workbook(file_path, read_only=True, data_only=True)
        sheet_name, ws, headers = pick_worksheet_by_required_cols(wb)

        if ws.sheet_state in ("hidden", "veryHidden"):
            raise ValueError("업로드 시트가 숨김 상태입니다. 숨김 해제 후 업로드하세요.")

        header_set = set(headers)
        missing = [c for c in REQUIRED_COLS if c not in header_set]
        if missing:
            raise ValueError(f"필수 컬럼 누락: {', '.join(missing)} (시트: {sheet_name})")

        total = max(int(ws.max_row) - 1, 0)  # 헤더 제외

        # ---------------------------------------------------------------------
        # 2) 사원번호 선 수집 → 기존 사용자 등급 조회(관리자 보호 판단)
        #    (read_only iterator 1회 소모 방지를 위해: 선 수집 후 workbook 재오픈)
        # ---------------------------------------------------------------------
        ids: List[str] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            emp_id, _name, _defaults = build_defaults_from_row(headers, row)
            if emp_id:
                ids.append(emp_id)

        existing_grade_map = dict(
            CustomUser.objects.filter(id__in=ids).values_list("id", "grade")
        )

        try:
            wb.close()
        except Exception:
            pass

        wb = load_workbook(file_path, read_only=True, data_only=True)
        sheet_name, ws, headers = pick_worksheet_by_required_cols(wb)

        # ---------------------------------------------------------------------
        # 3) 배치 처리 준비
        # ---------------------------------------------------------------------
        results: List[List[Any]] = []
        created = updated = skipped = err_cnt = 0
        processed = 0

        buffer_rows: List[Tuple[Any, ...]] = []
        current_excel_row_num = 2  # 엑셀 실제 행번호(헤더 다음)

        def set_percent_from_processed() -> None:
            if total <= 0:
                _cache_set_percent(k, 100)
                return
            p = int((processed / total) * 100)
            _cache_set_percent(k, p)

        @transaction.atomic
        def flush_chunk(rows_chunk: List[Tuple[Any, ...]], start_row_num: int) -> None:
            nonlocal created, updated, skipped, err_cnt, processed, results, existing_grade_map

            # 성능: chunk 내 기존 user 한 번에 미리 로드
            chunk_ids: List[str] = []
            built: List[Tuple[int, str, str, Dict[str, Any]]] = []  # (excel_row_num, emp_id, name, defaults)

            for offset, row in enumerate(rows_chunk):
                excel_row_num = start_row_num + offset
                emp_id, name, defaults = build_defaults_from_row(headers, row)

                if not emp_id:
                    skipped += 1
                    results.append([excel_row_num, "", name, "", "", "", "", "", "⚠️ 사원번호 누락(스킵)"])
                    processed += 1
                    continue

                built.append((excel_row_num, emp_id, name, defaults))
                chunk_ids.append(emp_id)

            users_by_id: Dict[str, CustomUser] = {
                u.id: u for u in CustomUser.objects.filter(id__in=chunk_ids)
            }

            for excel_row_num, emp_id, name, defaults in built:
                try:
                    user = users_by_id.get(emp_id)
                    channel = defaults.get("channel", "")
                    part = defaults.get("part", "")
                    branch = defaults.get("branch", "")
                    grade = defaults.get("grade", "")
                    status = defaults.get("status", "")
                    quit_ = defaults.get("quit")

                    # ---------------------------------------------------------
                    # Update path
                    # ---------------------------------------------------------
                    if user:
                        is_protected = user.grade in PROTECTED_GRADES
                        quit_newly_added = (user.quit is None and quit_ is not None)

                        # 1) 보호등급 + 퇴사일 신규 생성 아님 → 변경 차단
                        if is_protected and not quit_newly_added:
                            skipped += 1
                            results.append([
                                excel_row_num,
                                emp_id,
                                name,
                                channel,
                                part,
                                branch,
                                getattr(user, "grade", ""),
                                getattr(user, "status", ""),
                                "⚠️ 보호등급(superuser/head/leader) - 퇴사일 신규 없음(변경 차단)",
                            ])
                            processed += 1
                            continue

                        # 2) 보호등급 + 퇴사일 신규 생성 → resign/inactive 강제 전환
                        if is_protected and quit_newly_added:
                            forced_grade = "inactive" if ((not name) or ("*" in name)) else "resign"
                            forced_status = "재직" if forced_grade == "basic" else "퇴사"
                            defaults["grade"] = forced_grade
                            defaults["status"] = forced_status
                            defaults["is_active"] = (forced_grade != "inactive")
                            defaults["is_staff"] = False
                            defaults["is_superuser"] = False

                        update_fields: List[str] = []

                        for key, value in defaults.items():
                            # 보호 필드: 평소엔 빈값 덮어쓰기 금지, 단 quit 신규 생성이면 초기화 허용
                            if key in PROTECTED_FIELDS:
                                if value:
                                    setattr(user, key, value)
                                    update_fields.append(key)
                                else:
                                    if quit_newly_added:
                                        setattr(user, key, "")
                                        update_fields.append(key)
                                continue

                            # 일반 필드: None/빈문자열은 덮어쓰기 하지 않음(데이터 소실 방지)
                            if value is None:
                                continue
                            if isinstance(value, str) and value == "":
                                continue

                            setattr(user, key, value)
                            update_fields.append(key)

                        if update_fields:
                            # 중복 제거
                            user.save(update_fields=list(dict.fromkeys(update_fields)))

                        existing_grade_map[emp_id] = user.grade

                        updated += 1
                        results.append([
                            excel_row_num,
                            emp_id,
                            name,
                            channel,
                            part,
                            branch,
                            getattr(user, "grade", ""),
                            getattr(user, "status", ""),
                            "✅ 기존 업데이트",
                        ])

                    # ---------------------------------------------------------
                    # Create path
                    # ---------------------------------------------------------
                    else:
                        CustomUser.objects.create_user(
                            id=emp_id,
                            password=emp_id,  # 초기 비밀번호 = 사원번호
                            **defaults,
                        )
                        existing_grade_map[emp_id] = defaults.get("grade", "basic")

                        created += 1
                        results.append([
                            excel_row_num,
                            emp_id,
                            name,
                            channel,
                            part,
                            branch,
                            defaults.get("grade", ""),
                            defaults.get("status", ""),
                            "🟢 신규 등록",
                        ])

                except Exception as e:
                    err_cnt += 1
                    results.append([
                        excel_row_num,
                        emp_id,
                        name,
                        defaults.get("channel", ""),
                        defaults.get("part", ""),
                        defaults.get("branch", ""),
                        defaults.get("grade", ""),
                        defaults.get("status", ""),
                        f"❌ 오류: {e}",
                    ])

                processed += 1

            set_percent_from_processed()

        # ---------------------------------------------------------------------
        # 4) batch loop
        # ---------------------------------------------------------------------
        for row in ws.iter_rows(min_row=2, values_only=True):
            buffer_rows.append(row)
            if len(buffer_rows) >= batch_size:
                flush_chunk(buffer_rows, start_row_num=current_excel_row_num)
                current_excel_row_num += len(buffer_rows)
                buffer_rows = []

        if buffer_rows:
            flush_chunk(buffer_rows, start_row_num=current_excel_row_num)

        # ---------------------------------------------------------------------
        # 5) 결과 리포트 생성 + 저장
        # ---------------------------------------------------------------------
        result_wb = _make_result_wb(
            results=results,
            total=total,
            new_cnt=created,
            upd_cnt=updated,
            skip_cnt=skipped,
            err_cnt=err_cnt,
            picked_sheet=sheet_name,
        )
        result_path = _save_result_workbook(task_id, result_wb)

        # ---------------------------------------------------------------------
        # 6) cache finalize (SUCCESS)
        # ---------------------------------------------------------------------
        _cache_success(k, result_path)

        logger.warning(
            "[TASK DONE] tid=%s status=SUCCESS sheet=%s total=%s created=%s updated=%s skipped=%s errors=%s",
            task_id, sheet_name, total, created, updated, skipped, err_cnt
        )

        return {
            "status": "SUCCESS",
            "result_path": result_path,
            "sheet": sheet_name,
            "total": total,
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "errors": err_cnt,
        }

    except Exception as e:
        logger.exception("[TASK FAIL] tid=%s file=%s", task_id, file_path)
        _cache_fail(k, e)
        raise

    finally:
        try:
            if wb:
                wb.close()
        except Exception:
            pass
