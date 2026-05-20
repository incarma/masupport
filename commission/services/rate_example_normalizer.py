# commission/services/rate_example_normalizer.py
from __future__ import annotations

"""
RateExample 정규화 서비스.

역할:
- 업로드된 예시표 원본 파일을 보험사별 규칙에 따라 정규화 테이블로 적재한다.
- 원본 파일 저장/검증은 RateExampleService가 담당하고, 이 파일은 정규화만 담당한다.

보안/운영 원칙:
- 파일 URL 직접 접근 금지. FieldFile.path만 서버 내부에서 사용.
- 파싱 실패 시 예외를 삼키지 않고 호출부에서 rollback 가능하도록 raise.
- 동일 보험사·구분 정규화 데이터는 최신 업로드 기준으로 교체한다.
- 지원 대상 보험사에서 정규화 결과 0건이면 조용히 성공 처리하지 않는다.
"""

import logging
import os
import tempfile
import zipfile

from openpyxl import load_workbook

from commission.models import RateExample, RateExampleConversionRow
from commission.services.rate_example_normalizers.life_abl import build_life_abl_conversion_rows
from commission.services.rate_example_normalizers.life_db import build_life_db_conversion_rows
from commission.services.rate_example_normalizers.life_im import build_life_im_conversion_rows
from commission.services.rate_example_normalizers.life_kb import (
    build_life_kb_general_conversion_rows,
    build_life_kb_health_conversion_rows,
)
from commission.services.rate_example_normalizers.life_KDB import build_life_kdb_conversion_rows
from commission.services.rate_example_normalizers.life_kyobo import build_life_kyobo_conversion_rows
from commission.services.rate_example_normalizers.life_nh import build_life_nh_conversion_rows
from commission.services.rate_example_normalizers.life_dongyang import build_life_dongyang_conversion_rows
from commission.services.rate_example_normalizers.life_lina import (
    build_life_lina_conversion_rows,
    build_life_lina_pdf_conversion_rows,
)
from commission.services.rate_example_normalizers.life_met import build_life_met_conversion_rows
from commission.services.rate_example_normalizers.life_shinhan import build_life_shinhan_conversion_rows
from commission.services.rate_example_normalizers.life_chubb import build_life_chubb_pdf_conversion_rows
from commission.services.rate_example_normalizers.life_cardif import build_life_cardif_pdf_conversion_rows
from commission.services.rate_example_normalizers.life_mirae import build_life_mirae_conversion_rows
from commission.services.rate_example_normalizers.life_samsung import build_life_samsung_conversion_rows
from commission.services.rate_example_normalizers.life_fubon import build_life_fubon_pdf_conversion_rows
from commission.services.rate_example_normalizers.life_hana import build_life_hana_pdf_conversion_rows
from commission.services.rate_example_normalizers.life_heungkuk import build_life_heungkuk_pdf_conversion_rows
from commission.services.rate_example_normalizers.life_hanhwa import build_life_hanhwa_conversion_rows

from commission.services.rate_example_normalizers.fire_db import build_fire_db_conversion_rows
from commission.services.rate_example_normalizers.fire_kb import build_fire_kb_conversion_rows
from commission.services.rate_example_normalizers.fire_nh import build_fire_nh_conversion_rows
from commission.services.rate_example_normalizers.fire_samsung import build_fire_samsung_conversion_rows
from commission.services.rate_example_normalizers.fire_lotte import build_fire_lotte_conversion_rows
from commission.services.rate_example_normalizers.fire_hanhwa import build_fire_hanhwa_conversion_rows
from commission.services.rate_example_normalizers.fire_hyundai import build_fire_hyundai_conversion_rows
from commission.services.rate_example_normalizers.fire_aig import build_fire_aig_pdf_conversion_rows
from commission.services.rate_example_normalizers.fire_hana import build_fire_hana_pdf_conversion_rows
from commission.services.rate_example_normalizers.fire_meritz import build_fire_meritz_pdf_conversion_rows
from commission.services.rate_example_normalizers.fire_heungkuk import build_fire_heungkuk_conversion_rows

logger = logging.getLogger(__name__)

LEGACY_FIRE_TYPE = "nonlife"


def _effective_insurer_type(value: str) -> str:
    """레거시 nonlife를 현재 손보 SSOT인 fire로 보정한다."""
    raw = str(value or "").strip()
    return RateExample.TYPE_FIRE if raw == LEGACY_FIRE_TYPE else raw


def _delete_existing_conversion_rows(example: RateExample, *, effective_type: str) -> None:
    """replace 모드 삭제. 손보는 fire/nonlife 혼재 row를 함께 정리한다."""
    if effective_type == RateExample.TYPE_FIRE:
        RateExampleConversionRow.objects.filter(
            insurer_type__in=[RateExample.TYPE_FIRE, LEGACY_FIRE_TYPE],
            category=example.category,
            insurer=example.insurer,
        ).delete()
        return

    RateExampleConversionRow.objects.filter(
        insurer_type=effective_type,
        category=example.category,
        insurer=example.insurer,
    ).delete()


def _raise_if_supported_but_empty(
    *,
    example: RateExample,
    effective_type: str,
    normalized_count: int,
    original_name: str,
) -> None:
    """지원 대상인데 정규화 결과가 0건이면 실패 처리한다."""
    if normalized_count > 0:
        return

    label = "수정률" if effective_type == RateExample.TYPE_FIRE else "환산율"
    raise ValueError(
        f"{example.insurer} {label} 정규화 결과가 0건입니다. "
        f"원본 파일 형식 또는 parser 조건을 확인해 주세요. file={original_name}"
    )


def _load_workbook_safely(path: str):
    """openpyxl workbook 로드. custom property 오류 시 해당 XML만 제거한 임시본으로 재시도한다."""
    try:
        return load_workbook(path, data_only=True, read_only=False)
    except TypeError as exc:
        message = str(exc)
        if "openpyxl.packaging.custom" not in message and "CustomProperty" not in message:
            raise

        logger.warning(
            "openpyxl custom properties load failed. retry without docProps/custom.xml: %s",
            path,
            exc_info=True,
        )

        tmp_path = ""
        try:
            fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
            os.close(fd)

            with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(
                tmp_path, "w", zipfile.ZIP_DEFLATED
            ) as zout:
                for item in zin.infolist():
                    if item.filename in {"docProps/custom.xml", "docProps/_rels/custom.xml.rels"}:
                        continue
                    zout.writestr(item, zin.read(item.filename))

            return load_workbook(tmp_path, data_only=True, read_only=False)
        finally:
            if tmp_path:
                try:
                    os.remove(tmp_path)
                except OSError:
                    logger.warning("temporary sanitized xlsx cleanup failed: %s", tmp_path, exc_info=True)


def normalize_rate_example(
    example: RateExample,
    *,
    product_kind: str = "",
    normalize_mode: str = "replace",
) -> int:
    """
    RateExample 업로드 파일을 정규화한다.

    반환:
    - 생성된 RateExampleConversionRow 수
    """
    normalize_mode = (normalize_mode or "replace").strip()
    if normalize_mode not in {"replace", "append"}:
        raise ValueError(f"Invalid normalize_mode: {normalize_mode}")

    effective_type = _effective_insurer_type(example.insurer_type)

    # 지급률 정규화는 별도 SSOT로 위임한다.
    if example.category == RateExample.CAT_PAY:
        from commission.services.rate_example_pay_normalizer import normalize_pay_rate_example

        return normalize_pay_rate_example(example, normalize_mode=normalize_mode)

    is_life_conv_target = (
        effective_type == RateExample.TYPE_LIFE
        and example.category == RateExample.CAT_CONV
        and example.insurer
        in {
            "ABL", "DB", "IM", "KB", "KDB", "교보", "농협", "동양",
            "라이나", "메트", "미래", "삼성", "신한", "처브", "카디프",
            "푸본현대", "하나", "한화", "흥국",
        }
    )
    is_fire_conv_target = (
        effective_type == RateExample.TYPE_FIRE
        and example.category == RateExample.CAT_CONV
        and example.insurer
        in {"AIG", "DB", "KB", "농협", "롯데", "메리츠", "삼성", "하나", "한화", "현대", "흥국"}
    )

    if not (is_life_conv_target or is_fire_conv_target):
        return 0

    if not example.file:
        return 0

    original_name = str(example.original_name or "").lower()

    # PDF 정규화 분기
    if original_name.endswith(".pdf"):
        normalized_rows: list[RateExampleConversionRow] = []

        if effective_type == RateExample.TYPE_FIRE and example.insurer == "메리츠":
            normalized_rows = build_fire_meritz_pdf_conversion_rows(example)
        elif effective_type == RateExample.TYPE_FIRE and example.insurer == "하나":
            normalized_rows = build_fire_hana_pdf_conversion_rows(example)
        elif effective_type == RateExample.TYPE_FIRE and example.insurer == "AIG":
            normalized_rows = build_fire_aig_pdf_conversion_rows(example)
        elif effective_type == RateExample.TYPE_FIRE and example.insurer == "흥국":
            normalized_rows = build_fire_heungkuk_conversion_rows(example)
        elif effective_type == RateExample.TYPE_LIFE and example.insurer == "처브":
            normalized_rows = build_life_chubb_pdf_conversion_rows(example)
        elif effective_type == RateExample.TYPE_LIFE and example.insurer == "카디프":
            normalized_rows = build_life_cardif_pdf_conversion_rows(example)
        elif effective_type == RateExample.TYPE_LIFE and example.insurer == "푸본현대":
            normalized_rows = build_life_fubon_pdf_conversion_rows(example)
        elif effective_type == RateExample.TYPE_LIFE and example.insurer == "하나":
            normalized_rows = build_life_hana_pdf_conversion_rows(example)
        elif effective_type == RateExample.TYPE_LIFE and example.insurer == "흥국":
            normalized_rows = build_life_heungkuk_pdf_conversion_rows(example)
        elif effective_type == RateExample.TYPE_LIFE and example.insurer == "라이나":
            normalized_rows = build_life_lina_pdf_conversion_rows(example)
        else:
            return 0

        if normalize_mode == "replace":
            _delete_existing_conversion_rows(example, effective_type=effective_type)

        if normalized_rows:
            RateExampleConversionRow.objects.bulk_create(normalized_rows, batch_size=500)

        _raise_if_supported_but_empty(
            example=example,
            effective_type=effective_type,
            normalized_count=len(normalized_rows),
            original_name=original_name,
        )
        return len(normalized_rows)

    # XLSX 정규화 대상인데 확장자가 맞지 않으면 실패 처리한다.
    if not original_name.endswith(".xlsx"):
        raise ValueError(
            f"{example.insurer} 정규화는 현재 .xlsx 또는 지원 PDF만 가능합니다. "
            f"file={original_name}"
        )

    wb = _load_workbook_safely(example.file.path)
    normalized_rows: list[RateExampleConversionRow] = []

    # 손해보험 수정률
    if effective_type == RateExample.TYPE_FIRE and example.insurer == "DB":
        normalized_rows.extend(build_fire_db_conversion_rows(example, wb))
    elif effective_type == RateExample.TYPE_FIRE and example.insurer == "KB":
        normalized_rows.extend(build_fire_kb_conversion_rows(example, wb))
    elif effective_type == RateExample.TYPE_FIRE and example.insurer == "농협":
        normalized_rows.extend(build_fire_nh_conversion_rows(example, wb))
    elif effective_type == RateExample.TYPE_FIRE and example.insurer == "롯데":
        normalized_rows.extend(build_fire_lotte_conversion_rows(example, wb))
    elif effective_type == RateExample.TYPE_FIRE and example.insurer == "삼성":
        normalized_rows.extend(build_fire_samsung_conversion_rows(example, wb))
    elif effective_type == RateExample.TYPE_FIRE and example.insurer == "한화":
        normalized_rows.extend(build_fire_hanhwa_conversion_rows(example, wb))
    elif effective_type == RateExample.TYPE_FIRE and example.insurer == "현대":
        normalized_rows.extend(build_fire_hyundai_conversion_rows(example, wb))

    # 생명보험 환산율
    elif effective_type == RateExample.TYPE_LIFE and example.insurer == "ABL":
        normalized_rows.extend(build_life_abl_conversion_rows(example, wb))
    elif effective_type == RateExample.TYPE_LIFE and example.insurer == "DB":
        normalized_rows.extend(build_life_db_conversion_rows(example, wb))
    elif effective_type == RateExample.TYPE_LIFE and example.insurer == "IM":
        normalized_rows.extend(build_life_im_conversion_rows(example, wb))
    elif effective_type == RateExample.TYPE_LIFE and example.insurer == "KB":
        if product_kind == "general":
            normalized_rows.extend(build_life_kb_general_conversion_rows(example, wb))
        elif product_kind == "health":
            normalized_rows.extend(build_life_kb_health_conversion_rows(example, wb))
        else:
            raise ValueError("KB 생명보험 상품 구분 값이 올바르지 않습니다.")
    elif effective_type == RateExample.TYPE_LIFE and example.insurer == "KDB":
        normalized_rows.extend(build_life_kdb_conversion_rows(example, wb))
    elif effective_type == RateExample.TYPE_LIFE and example.insurer == "교보":
        normalized_rows.extend(build_life_kyobo_conversion_rows(example, wb))
    elif effective_type == RateExample.TYPE_LIFE and example.insurer == "농협":
        normalized_rows.extend(build_life_nh_conversion_rows(example, wb))
    elif effective_type == RateExample.TYPE_LIFE and example.insurer == "동양":
        normalized_rows.extend(build_life_dongyang_conversion_rows(example, wb))
    elif effective_type == RateExample.TYPE_LIFE and example.insurer == "라이나":
        normalized_rows.extend(build_life_lina_conversion_rows(example, wb))
    elif effective_type == RateExample.TYPE_LIFE and example.insurer == "메트":
        normalized_rows.extend(build_life_met_conversion_rows(example, wb))
    elif effective_type == RateExample.TYPE_LIFE and example.insurer == "미래":
        normalized_rows.extend(build_life_mirae_conversion_rows(example, wb))
    elif effective_type == RateExample.TYPE_LIFE and example.insurer == "삼성":
        normalized_rows.extend(build_life_samsung_conversion_rows(example, wb))
    elif effective_type == RateExample.TYPE_LIFE and example.insurer == "신한":
        normalized_rows.extend(build_life_shinhan_conversion_rows(example, wb))
    elif effective_type == RateExample.TYPE_LIFE and example.insurer == "한화":
        normalized_rows.extend(
            build_life_hanhwa_conversion_rows(
                example,
                wb,
                product_kind=product_kind,
            )
        )

    if normalize_mode == "replace":
        _delete_existing_conversion_rows(example, effective_type=effective_type)

    if normalized_rows:
        RateExampleConversionRow.objects.bulk_create(normalized_rows, batch_size=500)

    _raise_if_supported_but_empty(
        example=example,
        effective_type=effective_type,
        normalized_count=len(normalized_rows),
        original_name=original_name,
    )

    return len(normalized_rows)