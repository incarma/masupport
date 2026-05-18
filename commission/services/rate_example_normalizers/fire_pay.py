# django_ma/commission/services/rate_example_normalizers/fire_pay.py
from __future__ import annotations

"""
손해보험(fire) 지급률 정규화 parser.

역할:
- 손해보험 지급률 RAW 엑셀의 [① 5천만,3천만↑] 시트만 정규화한다.
- 해당 시트 내 (5천만원↑) 구간만 정규화하고 (3천만원↑) 구간은 제외한다.
- 정규화 결과는 RateExamplePayRow에 insurer_type="fire", category="pay"로 저장한다.

정규화 컬럼:
보험사 | 상품군 | 초회 | 2~6회 | 7~12회 | 13회 | 14회 | 15회

RateExamplePayRow 필드 매핑:
- col_first = 초회
- col_yr1   = 2~6회
- col_m13   = 7~12회
- col_yr2   = 13회
- col_yr3   = 14회
- col_m36   = 15회
- col_m37   = None
- col_yr4   = None
"""

import logging
import re
import unicodedata
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from commission.models import RateExample, RateExamplePayRow

logger = logging.getLogger(__name__)

TARGET_SHEET_NAME = "① 5천만,3천만↑"
TARGET_TIER = "5천만↑"

PRODUCT_GROUPS = {
    "보장",
    "보장(태아)",
    "연금",
    "저축",
    "단독실손(초회)",
    "단독실손(갱신)",
}

PAY_HEADER_MAP = {
    "초회": "col_first",
    "2~6회": "col_yr1",
    "7~12회": "col_m13",
    "13회": "col_yr2",
    "14회": "col_yr3",
    "15회": "col_m36",
}

INSURER_CANONICAL_MAP = {
    "현대해상": "현대",
    "현대": "현대",
    "DB손보": "DB",
    "DB손해보험": "DB",
    "DB": "DB",
    "KB손보": "KB",
    "KB손해보험": "KB",
    "KB": "KB",
    "메리츠화재": "메리츠",
    "메리츠": "메리츠",
    "한화손보": "한화",
    "한화손해보험": "한화",
    "한화": "한화",
    "롯데손보": "롯데",
    "롯데손해보험": "롯데",
    "롯데": "롯데",
    "흥국화재": "흥국",
    "흥국손보": "흥국",
    "흥국": "흥국",
    "삼성화재": "삼성",
    "삼성손보": "삼성",
    "삼성": "삼성",
    "AIG손보": "AIG",
    "AIG": "AIG",
    "MG손보": "MG",
    "MG": "MG",
    "농협손보": "농협",
    "농협손해보험": "농협",
    "농협": "농협",
    "하나손보": "하나",
    "하나손해보험": "하나",
    "하나": "하나",
    "라이나손보": "라이나",
    "라이나손해보험": "라이나",
    "라이나": "라이나",
}

INVALID_INSURER_NAMES = {
    "",
    "구분",
    "구 분",
    "구  분",
}


def _clean_text(value: Any) -> str:
    """셀 값을 비교 가능한 문자열로 정규화한다."""
    text = unicodedata.normalize("NFKC", str(value or ""))
    text = text.replace("\u00a0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _header_key(value: Any) -> str:
    """회차 헤더 비교용 key."""
    return _clean_text(value).replace(" ", "").replace("\n", "")


def _canonical_insurer(value: Any) -> str:
    """RAW 보험사명을 화면/계산 canonical 보험사명으로 정규화한다."""
    text = _clean_text(value)
    if text in INVALID_INSURER_NAMES:
        return ""
    return INSURER_CANONICAL_MAP.get(text, text)


def _to_decimal(value: Any) -> Decimal:
    """
    지급률 셀 값을 Decimal로 변환한다.

    정책:
    - 손보 지급률도 생보 지급률과 동일하게
      RAW 지급률 / 0.97 로 보정하여 저장한다.

    예:
    323.33 → 333.329896...
    """
    if value is None or value == "":
        return Decimal("0")

    if isinstance(value, Decimal):
        return value

    raw = str(value).replace(",", "").replace("%", "").strip()
    if not raw:
        return Decimal("0")

    try:
        value_decimal = Decimal(raw)

        # ── 지급률 0.97 보정 ───────────────────────────────────
        # 운영 계산식 기준:
        # 저장 지급률 = RAW 지급률 / 0.97
        #
        # 예:
        # 323.33 → 333.329896...
        #
        # Decimal 정밀도 유지 위해 float 사용 금지
        return value_decimal / Decimal("0.97")
    except (InvalidOperation, ValueError):
        logger.warning("fire pay normalizer: invalid decimal value=%r", value)
        return Decimal("0")


def _iter_table_blocks(ws: Worksheet) -> list[tuple[int, int, dict[str, int]]]:
    """
    보험사 테이블 블록을 탐지한다.

    RAW 구조:
    - 보험사 행: '구 분'이 있는 행과 같은 행에 보험사명이 배치
    - 회차 헤더 행: 보험사 행 + 2행
    - 지급률 컬럼: 초회, 2~6회, 7~12회, 13회, 14회, 15회

    예외:
    - 메리츠처럼 15회 컬럼이 없고 14회 다음에 '계', '유지보수' 등이
      오는 보험사도 존재한다.
    - 따라서 6개 헤더 완전 일치 방식이 아니라, 헤더명 기반 동적 매핑으로
      필수 컬럼 5개(초회~14회)가 있으면 블록으로 인정한다.
    - 15회가 없으면 col_m36은 None으로 저장한다.
    """
    blocks: list[tuple[int, int, dict[str, int]]] = []

    required_fields = {
        "col_first",  # 초회
        "col_yr1",    # 2~6회
        "col_m13",    # 7~12회
        "col_yr2",    # 13회
        "col_yr3",    # 14회
    }

    for row_no in range(1, ws.max_row + 1):
        first_value = _clean_text(ws.cell(row_no, 1).value)
        if "구" not in first_value or "분" not in first_value:
            continue

        header_row = row_no + 2

        for col_no in range(1, ws.max_column + 1):
            insurer = _canonical_insurer(ws.cell(row_no, col_no).value)
            if not insurer:
                continue

            # ── 지급률 헤더명 기반 동적 매핑 ───────────────────────────
            # 기존 6칸 완전 일치 방식은 메리츠처럼 15회가 없는 블록을
            # 정규화하지 못한다. 현재 보험사 블록 시작열 기준 우측 8칸만
            # 탐색하여 '계', '유지보수' 같은 비정규화 컬럼은 무시한다.
            header_map: dict[str, int] = {}
            for offset in range(0, 8):
                scan_col = col_no + offset
                header = _header_key(ws.cell(header_row, scan_col).value)
                field_name = PAY_HEADER_MAP.get(header)
                if not field_name:
                    continue
                header_map[field_name] = scan_col

            if not required_fields.issubset(set(header_map)):
                continue

            blocks.append((row_no, col_no, header_map))

    return blocks


def _target_data_rows(ws: Worksheet, start_row: int) -> list[int]:
    """
    (5천만원↑) 구간의 상품군 행만 반환한다.

    (3천만원↑) 구간이 시작되면 즉시 중단한다.
    """
    rows: list[int] = []
    in_target_tier = False

    for row_no in range(start_row + 1, ws.max_row + 1):
        tier_text = _clean_text(ws.cell(row_no, 1).value)
        product_group = _clean_text(ws.cell(row_no, 2).value)

        if "3천만원" in tier_text:
            break

        if "5천만원" in tier_text:
            in_target_tier = True

        if not in_target_tier:
            continue

        if product_group in PRODUCT_GROUPS:
            rows.append(row_no)

    return rows


def build_fire_pay_rows(example: RateExample) -> list[RateExamplePayRow]:
    """
    손해보험 지급률 정규화 row를 생성한다.

    주의:
    - DB 저장은 이 함수에서 하지 않는다.
    - 삭제/replace/append 정책은 rate_example_pay_normalizer.py가 담당한다.
    """
    if not example.file:
        return []

    wb = load_workbook(
        example.file.path,
        data_only=True,
        read_only=False,
    )

    try:
        if TARGET_SHEET_NAME not in wb.sheetnames:
            logger.warning(
                "fire pay normalizer: target sheet not found. file=%s sheets=%s",
                example.original_name,
                wb.sheetnames,
            )
            return []

        ws = wb[TARGET_SHEET_NAME]
        rows: list[RateExamplePayRow] = []

        for insurer_row, insurer_col, header_map in _iter_table_blocks(ws):
            insurer = _canonical_insurer(ws.cell(insurer_row, insurer_col).value)
            if not insurer:
                continue

            for row_no in _target_data_rows(ws, insurer_row):
                product_group = _clean_text(ws.cell(row_no, 2).value)
                if product_group not in PRODUCT_GROUPS:
                    continue

                values = {
                    "col_first": None,
                    "col_yr1": None,
                    "col_m13": None,
                    "col_yr2": None,
                    "col_yr3": None,
                    "col_m36": None,
                    "col_m37": None,
                    "col_yr4": None,
                }

                for field_name, col_no in header_map.items():
                    values[field_name] = _to_decimal(ws.cell(row_no, col_no).value)

                rows.append(
                    RateExamplePayRow(
                        source_file=example,
                        source_sheet=ws.title,
                        source_row_no=row_no,
                        insurer_type=RateExample.TYPE_FIRE,
                        category=RateExample.CAT_PAY,
                        insurer=insurer,
                        tier=TARGET_TIER,
                        coverage_type=product_group,
                        **values,
                    )
                )

        logger.info(
            "fire pay normalizer: created %s rows. file=%s",
            len(rows),
            example.original_name,
        )
        return rows

    finally:
        wb.close()