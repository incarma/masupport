# django_ma/commission/services/rate_example_normalizers/life_lina.py
from __future__ import annotations

"""
라이나생명 환산율/수정률 정규화 parser.

역할:
- 라이나생명 raw Excel/PDF 파일을 RateExampleConversionRow로 정규화한다.
- Excel:
  - 병합 셀 값을 row 단위로 전파한다.
  - raw "구분" 컬럼 묶음의 첫 번째 컬럼은 상품명, 마지막 컬럼은 납기로 사용한다.
- PDF:
  - PDF 텍스트 추출 결과를 기반으로 상품명/납기/환산율 패턴을 복원한다.
  - PDF에는 Excel 병합 셀 정보가 없으므로 "병합 셀 분리"가 아니라 텍스트 흐름 기반 정규화로 처리한다.
- 공통:
  - 납기 값에 "년납"이 포함된 행만 정규화한다.
  - raw 환산율 값을 1차년~4차년에 동일하게 저장한다.
"""

import logging
import re
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from commission.models import RateExample, RateExampleConversionRow
from commission.services.rate_example_normalizers._common.excel import (
    build_merged_value_map,
    cell_value_with_merged,
)

logger = logging.getLogger(__name__)

INSURER_NAME = "라이나"
YEAR_PAY_KEYWORD = "년납"

RATE_RE = re.compile(r"(?P<rate>\d+(?:\.\d+)?)\s*%")
PAY_PERIOD_RE = re.compile(r"(?P<pay>\d+\s*년납)")
PAGE_NOISE_KEYWORDS = (
    "LINA의 기밀",
    "복사나 배포",
    "허가 받은 자",
    "저작권",
    "라이나생명 상품별 환산율",
    "그룹",
    "상품군",
    "구분",
    "대상상품",
    "특약",
    "납기",
    "환산율",
)


# =============================================================================
# 공통 유틸
# =============================================================================

def _to_text(value: Any) -> str:
    """셀/PDF 값을 비교·저장 가능한 문자열로 정규화한다."""
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _clean_pdf_text(value: str) -> str:
    """
    PDF 추출 텍스트 정리.

    - 여러 공백을 1칸으로 축소
    - 한글 단어 사이에 들어간 불필요한 공백 일부 보정
    """
    text = _to_text(value)
    text = text.replace(" 년납", "년납")
    text = text.replace(" 년만기", "년만기")
    text = text.replace(" 회", "회")
    text = text.replace(" 대", "대")
    return text.strip()


def _to_decimal_rate(value: Any, *, number_format: str = "") -> Decimal | None:
    """
    raw 환산율을 백분율 수치 기준 Decimal로 변환한다.

    예:
    - "245%" → Decimal("245")
    - 245 → Decimal("245")
    - Excel percent cell 2.45 + number_format "%" → Decimal("245")
    """
    if value is None:
        return None

    if isinstance(value, Decimal):
        raw = value
        has_percent_format = "%" in (number_format or "")
        return raw * Decimal("100") if has_percent_format and abs(raw) <= 10 else raw

    if isinstance(value, (int, float)):
        raw = Decimal(str(value))
        has_percent_format = "%" in (number_format or "")
        return raw * Decimal("100") if has_percent_format and abs(raw) <= 10 else raw

    text = _to_text(value)
    if not text:
        return None

    has_percent_text = "%" in text
    text = text.replace("%", "").replace(",", "").strip()

    try:
        raw = Decimal(text)
    except InvalidOperation:
        return None

    if has_percent_text:
        return raw

    if "%" in (number_format or "") and abs(raw) <= 10:
        return raw * Decimal("100")

    return raw


def _coverage_type(product_name: str) -> str:
    """라이나 보종 판정 정책."""
    if "종신" in product_name:
        return "종신/CI"
    return "기타(보장성)"


def _build_row(
    *,
    example: RateExample,
    sheet_name: str,
    row_no: int,
    product_name: str,
    pay_period: str,
    rate: Decimal,
) -> RateExampleConversionRow:
    """RateExampleConversionRow 생성 규칙을 한 곳에 고정한다."""
    product_name = _to_text(product_name)
    pay_period = _to_text(pay_period)

    return RateExampleConversionRow(
        source_file=example,
        source_sheet=sheet_name,
        source_row_no=row_no,
        insurer_type=example.insurer_type,
        category=example.category,
        insurer=INSURER_NAME,
        coverage_type=_coverage_type(product_name),
        strategy_flag="",
        product_name=product_name,
        plan_type="",
        pay_period=pay_period,
        year1=rate,
        year2=rate,
        year3=rate,
        year4=rate,
    )


# =============================================================================
# Excel parser
# =============================================================================

def _merged_value_map(ws: Worksheet) -> dict[tuple[int, int], Any]:
    """
    병합 셀의 값을 모든 셀 좌표에 전파하기 위한 map을 만든다.

    원본 worksheet는 수정하지 않는다.
    """
    return build_merged_value_map(ws)


def _cell_value(ws: Worksheet, merged_map: dict[tuple[int, int], Any], row_no: int, col_no: int) -> Any:
    """병합 셀 전파값을 우선 적용해 셀 값을 반환한다."""
    return cell_value_with_merged(ws, merged_map, row_no, col_no)


def _find_header_row(ws: Worksheet, merged_map: dict[tuple[int, int], Any]) -> int | None:
    """
    라이나 Excel raw 헤더 행을 찾는다.

    기준:
    - 같은 행에 "구분"과 "환산율"이 모두 존재해야 한다.
    """
    for row_no in range(1, min(ws.max_row, 30) + 1):
        row_texts = [
            _to_text(_cell_value(ws, merged_map, row_no, col_no))
            for col_no in range(1, ws.max_column + 1)
        ]
        joined = " ".join(row_texts)

        if "구분" in joined and "환산율" in joined:
            return row_no

    return None


def _find_rate_col(ws: Worksheet, merged_map: dict[tuple[int, int], Any], header_row: int) -> int | None:
    """헤더 행에서 환산율 컬럼을 찾는다."""
    for col_no in range(1, ws.max_column + 1):
        text = _to_text(_cell_value(ws, merged_map, header_row, col_no))
        if "환산율" in text:
            return col_no
    return None


def _find_division_cols(
    ws: Worksheet,
    merged_map: dict[tuple[int, int], Any],
    header_row: int,
    rate_col: int,
) -> tuple[int, int] | None:
    """
    raw "구분" 컬럼 영역을 찾는다.

    기본 정책:
    - "구분" 헤더가 있는 병합 범위를 우선 사용한다.
    - 병합 범위를 찾지 못하면 구분 헤더 컬럼부터 환산율 직전 컬럼까지를 구분 영역으로 본다.
    """
    division_col = None

    for col_no in range(1, rate_col):
        text = _to_text(_cell_value(ws, merged_map, header_row, col_no))
        if "구분" in text:
            division_col = col_no
            break

    if not division_col:
        return None

    for merged_range in ws.merged_cells.ranges:
        if (
            merged_range.min_row <= header_row <= merged_range.max_row
            and merged_range.min_col <= division_col <= merged_range.max_col
        ):
            return merged_range.min_col, min(merged_range.max_col, rate_col - 1)

    return division_col, rate_col - 1


def _iter_excel_sheet_rows(example: RateExample, ws: Worksheet) -> list[RateExampleConversionRow]:
    """
    단일 Excel worksheet를 라이나 정규화 row 목록으로 변환한다.

    제외 정책:
    - 상품명 공란 제외
    - 납기 공란 제외
    - 납기에 "년납"이 없는 행 제외
    - 환산율 파싱 불가 행 제외
    """
    merged_map = _merged_value_map(ws)

    header_row = _find_header_row(ws, merged_map)
    if not header_row:
        return []

    rate_col = _find_rate_col(ws, merged_map, header_row)
    if not rate_col:
        return []

    division_cols = _find_division_cols(ws, merged_map, header_row, rate_col)
    if not division_cols:
        return []

    product_col, pay_period_col = division_cols

    rows: list[RateExampleConversionRow] = []

    for row_no in range(header_row + 1, ws.max_row + 1):
        product_name = _to_text(_cell_value(ws, merged_map, row_no, product_col))
        pay_period = _to_text(_cell_value(ws, merged_map, row_no, pay_period_col))

        if not product_name or not pay_period or YEAR_PAY_KEYWORD not in pay_period:
            continue

        rate_cell = ws.cell(row_no, rate_col)
        rate_value = _cell_value(ws, merged_map, row_no, rate_col)
        rate = _to_decimal_rate(rate_value, number_format=getattr(rate_cell, "number_format", ""))

        if rate is None:
            continue

        rows.append(
            _build_row(
                example=example,
                sheet_name=ws.title,
                row_no=row_no,
                product_name=product_name,
                pay_period=pay_period,
                rate=rate,
            )
        )

    return rows


def build_life_lina_conversion_rows(example: RateExample, workbook) -> list[RateExampleConversionRow]:
    """
    라이나생명 Excel 환산율/수정률 정규화 진입점.

    시트 정책:
    - workbook 전체 시트를 순회한다.
    - 단, 헤더("구분", "환산율")를 찾지 못한 시트는 자동 제외한다.
    """
    normalized_rows: list[RateExampleConversionRow] = []

    for ws in workbook.worksheets:
        normalized_rows.extend(_iter_excel_sheet_rows(example, ws))

    return normalized_rows


# =============================================================================
# PDF parser
# =============================================================================

def _extract_pdf_lines(example: RateExample) -> list[tuple[int, int, str]]:
    """
    PDF에서 텍스트 라인을 추출한다.

    반환:
    - (page_no, line_no, text)
    """
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise RuntimeError(
            "라이나 PDF 정규화를 위해 pypdf 패키지가 필요합니다. requirements.txt에 pypdf를 추가해 주세요."
        ) from exc

    lines: list[tuple[int, int, str]] = []

    with example.file.open("rb") as f:
        reader = PdfReader(f)

        for page_idx, page in enumerate(reader.pages, start=1):
            text = page.extract_text() or ""
            for line_idx, raw_line in enumerate(text.splitlines(), start=1):
                line = _clean_pdf_text(raw_line)
                if not line:
                    continue
                lines.append((page_idx, line_idx, line))

    return lines


def _is_pdf_noise_line(line: str) -> bool:
    """PDF 페이지 헤더/푸터/표 헤더 등 정규화 대상이 아닌 라인을 제외한다."""
    if not line:
        return True

    if any(keyword in line for keyword in PAGE_NOISE_KEYWORDS):
        return True

    if re.fullmatch(r"\d{4}\.\s*\d{1,2}\.\s*\d{1,2}", line):
        return True

    if re.fullmatch(r"\d{4}\.\s*\d{1,2}", line):
        return True

    if re.fullmatch(r"[-–—]+", line):
        return True

    return False


def _strip_rate_text(line: str) -> tuple[str, Decimal | None]:
    """
    라인 끝 또는 라인 내부의 환산율을 추출하고, 환산율 텍스트 제거 후 본문을 반환한다.
    """
    matches = list(RATE_RE.finditer(line))
    if not matches:
        return line, None

    match = matches[-1]
    rate = _to_decimal_rate(match.group("rate") + "%")
    body = (line[: match.start()] + " " + line[match.end():]).strip()
    body = _clean_pdf_text(body)
    return body, rate


def _extract_pay_period(line: str) -> tuple[str, str]:
    """
    라인에서 납기("20년납")를 추출한다.

    반환:
    - 납기 제거 후 본문
    - 납기
    """
    matches = list(PAY_PERIOD_RE.finditer(line))
    if not matches:
        return line, ""

    match = matches[-1]
    pay_period = _clean_pdf_text(match.group("pay"))
    body = (line[: match.start()] + " " + line[match.end():]).strip()
    body = _clean_pdf_text(body)
    return body, pay_period


def _looks_like_product_name(line: str) -> bool:
    """
    PDF 텍스트 라인이 상품명/특약명으로 볼 수 있는지 판단한다.

    라이나 PDF에서는 상품/특약명이 대체로 "무배당"으로 시작한다.
    """
    if not line:
        return False
    if _is_pdf_noise_line(line):
        return False
    if RATE_RE.search(line):
        return False
    if line in {"Dental", "Cancer", "Modular", "CI", "Whole Life", "GI", "Term", "Dementia"}:
        return False
    return "무배당" in line


def _looks_like_product_continuation(line: str) -> bool:
    """
    PDF 줄바꿈으로 분리된 상품명 continuation line 여부를 판단한다.

    예:
    - 1줄: 무배당새로담는간편건강보험(3.5.5)(해약환급금미지
    - 2줄: 급형Ⅱ)
    → 2줄은 단독 상품명은 아니지만 직전 상품명에 붙어야 한다.
    """
    if not line:
        return False
    if _is_pdf_noise_line(line):
        return False
    if RATE_RE.search(line):
        return False
    if PAY_PERIOD_RE.search(line):
        return False
    if line in {"주계약/기타특약", "주보험/기타특약", "기타특약", "주계약", "주보험"}:
        return False
    if line in {"Dental", "Cancer", "Modular", "CI", "Whole Life", "GI", "Term", "Dementia"}:
        return False
    
    # 신규 상품명은 continuation이 아니다.
    # 기존 버그:
    # - 다음 "무배당..." 상품명까지 직전 상품명에 이어붙어
    #   상품명이 두 번, 세 번 중복 표기됨.
    if "무배당" in line:
        return False

    # PDF 줄바꿈 조각만 허용한다.
    # 예: "급형Ⅱ)", "지급형_갱신형)", "형)"
    # 너무 긴 일반 라인은 다른 컬럼/다른 상품명일 가능성이 있어 제외한다.
    if len(line) > 40:
        return False

    return bool(re.search(r"[가-힣A-Za-z0-9_ⅡⅢⅣ]", line))


def _merge_pdf_product_lines(base: str, continuation: str) -> str:
    """
    PDF에서 분리된 상품명 줄을 자연스럽게 병합한다.

    핵심 보정:
    - '미지' + '급형Ⅱ)' → '미지급형Ⅱ)'
    - 일반 분리 줄은 공백 없이 결합해 원문 상품명 훼손을 줄인다.
    """
    base = _clean_pdf_text(base)
    continuation = _clean_pdf_text(continuation)

    if not base:
        return continuation
    if not continuation:
        return base

    return _clean_pdf_text(f"{base}{continuation}")


def _normalize_pdf_product_name(line: str) -> str:
    """PDF에서 추출한 상품명 텍스트를 정리한다."""
    text = _clean_pdf_text(line)
    text = RATE_RE.sub("", text).strip()
    text, _pay = _extract_pay_period(text)

    # 표의 중간 컬럼 값이 상품명에 붙은 경우 제거한다.
    for token in ("주계약/기타특약", "주보험/기타특약", "기타특약", "주계약", "주보험"):
        text = text.replace(token, " ")

    return _clean_pdf_text(text)


def _append_pdf_row_if_valid(
    *,
    rows: list[RateExampleConversionRow],
    example: RateExample,
    page_no: int,
    line_no: int,
    product_name: str,
    pay_period: str,
    rate: Decimal | None,
) -> None:
    """PDF에서 복원된 후보 행을 검증 후 정규화 row로 추가한다."""
    product_name = _normalize_pdf_product_name(product_name)
    pay_period = _clean_pdf_text(pay_period)

    if not product_name:
        return
    if not pay_period or YEAR_PAY_KEYWORD not in pay_period:
        return
    if rate is None:
        return

    rows.append(
        _build_row(
            example=example,
            sheet_name=f"PDF p.{page_no}",
            row_no=line_no,
            product_name=product_name,
            pay_period=pay_period,
            rate=rate,
        )
    )


def _parse_lina_pdf_lines(
    example: RateExample,
    lines: list[tuple[int, int, str]],
) -> list[RateExampleConversionRow]:
    """
    라이나 PDF 텍스트 라인을 정규화 row로 변환한다.

    처리 패턴:
    1. 한 줄에 상품명 + 납기 + 환산율이 모두 있는 경우
    2. 상품명 라인 이후, 다음 라인에 납기 + 환산율이 있는 경우
    3. 상품명 라인 이후, 납기 라인과 환산율 라인이 분리된 경우
    """
    rows: list[RateExampleConversionRow] = []

    current_product = ""
    current_product_page_no = 0
    current_product_line_no = 0
    pending_pay_period = ""
    pending_page_no = 0
    pending_line_no = 0

    for page_no, line_no, raw_line in lines:
        line = _clean_pdf_text(raw_line)

        if _is_pdf_noise_line(line):
            continue

        body_without_rate, rate = _strip_rate_text(line)
        body_without_rate = _clean_pdf_text(body_without_rate)
        body_without_pay, pay_period = _extract_pay_period(body_without_rate)

        # 패턴 1: 같은 라인에 상품명 + 납기 + 환산율이 모두 존재
        if pay_period and rate is not None and _looks_like_product_name(body_without_pay):
            _append_pdf_row_if_valid(
                rows=rows,
                example=example,
                page_no=page_no,
                line_no=line_no,
                product_name=body_without_pay,
                pay_period=pay_period,
                rate=rate,
            )
            current_product = _normalize_pdf_product_name(body_without_pay)
            pending_pay_period = ""
            continue

        # 패턴 2: 현재 상품명 + 같은 라인 납기 + 환산율
        if pay_period and rate is not None:
            _append_pdf_row_if_valid(
                rows=rows,
                example=example,
                page_no=page_no,
                line_no=line_no,
                product_name=current_product or body_without_pay,
                pay_period=pay_period,
                rate=rate,
            )
            pending_pay_period = ""
            continue

        # 패턴 3-1: 납기만 먼저 나온 경우
        if pay_period and rate is None:
            pending_pay_period = pay_period
            pending_page_no = page_no
            pending_line_no = line_no
            # 납기 앞쪽에 상품명이 같이 있으면 current_product 갱신
            if _looks_like_product_name(body_without_pay):
                current_product = _normalize_pdf_product_name(body_without_pay)
            continue

        # 패턴 3-2: 직전 납기 라인 다음에 환산율만 나온 경우
        if rate is not None and pending_pay_period:
            _append_pdf_row_if_valid(
                rows=rows,
                example=example,
                page_no=pending_page_no or page_no,
                line_no=pending_line_no or line_no,
                product_name=current_product or body_without_rate,
                pay_period=pending_pay_period,
                rate=rate,
            )
            pending_pay_period = ""
            continue

        # 상품명 후보 갱신
        if _looks_like_product_name(line):
            current_product = _normalize_pdf_product_name(line)
            current_product_page_no = page_no
            current_product_line_no = line_no
            pending_pay_period = ""

        # 상품명이 PDF 줄바꿈으로 분리된 경우 직전 상품명에 이어붙인다.
        # 예: "...해약환급금미지" + "급형Ⅱ)".
        if current_product and _looks_like_product_continuation(line):
            current_product = _normalize_pdf_product_name(
                _merge_pdf_product_lines(current_product, line)
            )
            # source_row_no는 상품명이 시작된 줄을 유지한다.
            current_product_page_no = current_product_page_no or page_no
            current_product_line_no = current_product_line_no or line_no
            pending_pay_period = ""
            continue

    return rows


def _dedupe_rows(rows: list[RateExampleConversionRow]) -> list[RateExampleConversionRow]:
    """
    PDF 텍스트 추출 중복 방어.

    기준:
    - 보험사 + 보종 + 상품명 + 납기 + 1차년 환산율
    """
    seen: set[tuple[str, str, str, str, Decimal | None]] = set()
    deduped: list[RateExampleConversionRow] = []

    for row in rows:
        key = (
            row.insurer,
            row.coverage_type,
            row.product_name,
            row.pay_period,
            row.year1,
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(row)

    return deduped


def build_life_lina_pdf_conversion_rows(example: RateExample) -> list[RateExampleConversionRow]:
    """
    라이나생명 PDF 환산율/수정률 정규화 진입점.

    전제:
    - PDF에서 텍스트 추출이 가능해야 한다.
    - 스캔 이미지형 PDF는 OCR 없이는 정규화할 수 없다.

    제외 정책:
    - 납기에 "년납"이 없는 항목 제외
    - "년만기"만 있는 항목 제외
    - 환산율 파싱 불가 항목 제외
    """
    lines = _extract_pdf_lines(example)
    rows = _parse_lina_pdf_lines(example, lines)

    deduped = _dedupe_rows(rows)

    logger.info(
        "lina pdf normalized: example_id=%s original=%s rows=%s",
        example.pk,
        example.original_name,
        len(deduped),
    )

    return deduped