# django_ma/commission/services/rate_example_normalizers/life_shinhan.py
from __future__ import annotations

"""
신한생명 환산율/수정률 정규화 parser.

역할:
- 신한 raw 예시표 xlsx 파일에서 일반상품/건강 시트를 표준 RateExampleConversionRow로 변환한다.
- 저장 단위는 기존 환산율 정책과 동일하게 백분율 수치 기준 Decimal이다.
  예: Excel 표시 100% → DB Decimal("100")

정규화 범위:
1) 시트명에 "일반상품" 포함
2) 시트명에 "건강" 포함
"""

import logging
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from commission.models import RateExample, RateExampleConversionRow

logger = logging.getLogger(__name__)

INSURER_NAME = "신한"


# ============================================================================
# 공통 helper
# ============================================================================

def _cell_text(value: Any) -> str:
    """
    raw 셀 값을 비교/저장 가능한 문자열로 정규화한다.

    - None은 공란
    - 줄바꿈/연속 공백은 한 칸으로 축약
    - 숫자형 20.0은 "20"으로 보정
    """
    if value is None:
        return ""

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    text = str(value).replace("\n", " ").replace("\r", " ").strip()
    return " ".join(text.split())


def _to_decimal_percent(value: Any, *, number_format: str = "") -> Decimal | None:
    """
    환산율 셀을 백분율 수치 기준 Decimal로 변환한다.

    기존 RateExample 환산율 저장 정책:
    - DB에는 1.0이 아니라 100.0을 저장한다.
    - Excel 셀이 percent format이고 openpyxl이 1.0으로 읽으면 100을 곱한다.
    """
    if value is None or value == "":
        return None

    if isinstance(value, str):
        raw = value.strip().replace(",", "")
        if not raw:
            return None

        has_percent = "%" in raw
        raw = raw.replace("%", "").strip()

        try:
            dec = Decimal(raw)
        except InvalidOperation:
            logger.debug("shinhan rate value skipped: value=%r", value)
            return None

        if has_percent and abs(dec) <= Decimal("10"):
            dec *= Decimal("100")

        return dec.quantize(Decimal("0.0001"))

    try:
        dec = Decimal(str(value))
    except (InvalidOperation, ValueError):
        logger.debug("shinhan rate value skipped: value=%r", value)
        return None

    if "%" in str(number_format or "") and abs(dec) <= Decimal("10"):
        dec *= Decimal("100")

    return dec.quantize(Decimal("0.0001"))


def _coverage_from_product_name(product_name: str) -> str:
    """
    신한 보종 매핑.

    우선순위:
    - 연금 + 변액: 변액연금
    - 경영: CEO정기
    - 종신: 종신/CI
    - 연금: 연금
    - 그 외: 기타(보장성)
    """
    name = product_name or ""

    if "연금" in name and "변액" in name:
        return "변액연금"
    if "경영" in name:
        return "CEO정기"
    if "종신" in name:
        return "종신/CI"
    if "연금" in name:
        return "연금"
    return "기타(보장성)"


def _join_plan_type(value_d: Any, value_e: Any) -> str:
    """
    일반상품 구분(D/E) 결합.

    - D/E 둘 다 있으면 "D, E"
    - 한쪽만 있으면 해당 값
    - 둘 다 없으면 공란
    """
    parts = [_cell_text(value_d), _cell_text(value_e)]
    return ", ".join([p for p in parts if p])


def _is_general_sheet(sheet_name: str) -> bool:
    return "일반상품" in str(sheet_name or "")


def _is_health_sheet(sheet_name: str) -> bool:
    return "건강" in str(sheet_name or "")


def _build_row(
    *,
    example: RateExample,
    sheet_name: str,
    row_no: int,
    coverage_type: str,
    product_name: str,
    plan_type: str,
    pay_period: str,
    year1: Decimal | None,
    year2: Decimal | None,
    year3: Decimal | None,
) -> RateExampleConversionRow:
    """RateExampleConversionRow 생성 공통 팩토리."""
    return RateExampleConversionRow(
        source_file=example,
        source_sheet=sheet_name,
        source_row_no=row_no,
        insurer_type=RateExample.TYPE_LIFE,
        category=RateExample.CAT_CONV,
        insurer=INSURER_NAME,
        coverage_type=coverage_type,
        strategy_flag="",
        product_name=product_name,
        plan_type=plan_type,
        pay_period=pay_period,
        year1=year1,
        year2=year2,
        year3=year3,
        year4=None,
    )


# ============================================================================
# 일반상품 parser
# ============================================================================

def _last_general_row(ws: Worksheet) -> int:
    """
    일반상품 대상 마지막 행.

    기준:
    - 7행부터 시작
    - 1Y(H열)에 마지막 데이터가 있는 행까지 정규화
    """
    last = 6
    for row_no in range(7, ws.max_row + 1):
        if ws.cell(row_no, 8).value not in (None, ""):
            last = row_no
    return last


def _parse_general_sheet(example: RateExample, ws: Worksheet) -> list[RateExampleConversionRow]:
    """
    신한 일반상품 시트 정규화.

    컬럼 매핑:
    - 상품명: C
    - 구분: D/E 결합
    - 납기: F
    - 1차년: H
    - 2차년: I
    - 3차년: J
    - 4차년: 없음
    """
    rows: list[RateExampleConversionRow] = []

    last_product_name = ""
    last_plan_by_product: dict[str, str] = {}

    last_row = _last_general_row(ws)
    if last_row < 7:
        return rows

    for row_no in range(7, last_row + 1):
        raw_product = _cell_text(ws.cell(row_no, 3).value)
        if raw_product:
            last_product_name = raw_product

        product_name = last_product_name
        if not product_name:
            continue

        plan_type = _join_plan_type(
            ws.cell(row_no, 4).value,
            ws.cell(row_no, 5).value,
        )

        # D/E가 모두 공란이면 동일 상품명 안에서 직전 구분값을 전파한다.
        if plan_type:
            last_plan_by_product[product_name] = plan_type
        else:
            plan_type = last_plan_by_product.get(product_name, "")

        pay_period = _cell_text(ws.cell(row_no, 6).value)

        year1_cell = ws.cell(row_no, 8)
        year2_cell = ws.cell(row_no, 9)
        year3_cell = ws.cell(row_no, 10)

        year1 = _to_decimal_percent(year1_cell.value, number_format=year1_cell.number_format)
        year2 = _to_decimal_percent(year2_cell.value, number_format=year2_cell.number_format)
        year3 = _to_decimal_percent(year3_cell.value, number_format=year3_cell.number_format)

        # 기준 열 H가 비어 있으면 대상 행이 아니다.
        if year1 is None:
            continue

        rows.append(
            _build_row(
                example=example,
                sheet_name=ws.title,
                row_no=row_no,
                coverage_type=_coverage_from_product_name(product_name),
                product_name=product_name,
                plan_type=plan_type,
                pay_period=pay_period,
                year1=year1,
                year2=year2,
                year3=year3,
            )
        )

    return rows


# ============================================================================
# 건강 parser
# ============================================================================

def _parse_health_sheet(example: RateExample, ws: Worksheet) -> list[RateExampleConversionRow]:
    """
    신한 건강 시트 정규화.

    조건:
    - 8행이 헤더
    - 9행부터 검사
    - A열 구분 값이 "주보험"인 행만 정규화

    컬럼 매핑:
    - 상품명: C
    - 보종: 기타(보장성)
    - 구분: 공란
    - 납기: G
    - 1차년: H
    - 2차년: I
    - 3차년: J
    - 4차년: 없음
    """
    rows: list[RateExampleConversionRow] = []

    for row_no in range(9, ws.max_row + 1):
        row_type = _cell_text(ws.cell(row_no, 1).value)
        if row_type != "주보험":
            continue

        product_name = _cell_text(ws.cell(row_no, 3).value)
        if not product_name:
            continue

        pay_period = _cell_text(ws.cell(row_no, 7).value)

        year1_cell = ws.cell(row_no, 8)
        year2_cell = ws.cell(row_no, 9)
        year3_cell = ws.cell(row_no, 10)

        year1 = _to_decimal_percent(year1_cell.value, number_format=year1_cell.number_format)
        year2 = _to_decimal_percent(year2_cell.value, number_format=year2_cell.number_format)
        year3 = _to_decimal_percent(year3_cell.value, number_format=year3_cell.number_format)

        if year1 is None:
            continue

        rows.append(
            _build_row(
                example=example,
                sheet_name=ws.title,
                row_no=row_no,
                coverage_type="기타(보장성)",
                product_name=product_name,
                plan_type="",
                pay_period=pay_period,
                year1=year1,
                year2=year2,
                year3=year3,
            )
        )

    return rows


# ============================================================================
# Public entrypoint
# ============================================================================

def build_life_shinhan_conversion_rows(
    example: RateExample,
    workbook,
) -> list[RateExampleConversionRow]:
    """
    신한생명 환산율/수정률 정규화 public entrypoint.

    dispatcher:
    - 시트명에 "일반상품" 포함 → 일반상품 parser
    - 시트명에 "건강" 포함 → 건강 parser
    """
    normalized_rows: list[RateExampleConversionRow] = []

    for ws in workbook.worksheets:
        sheet_name = str(ws.title or "")

        if _is_general_sheet(sheet_name):
            normalized_rows.extend(_parse_general_sheet(example, ws))
            continue

        if _is_health_sheet(sheet_name):
            normalized_rows.extend(_parse_health_sheet(example, ws))
            continue

    return normalized_rows