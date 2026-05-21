# django_ma/commission/services/rate_example_normalizers/life_hanhwa.py
from __future__ import annotations

"""
한화생명 환산율/수정률 정규화 모듈.

파일 위치:
    commission/services/rate_example_normalizers/life_hanhwa.py

역할:
- 한화생명 raw XLSX 파일을 RateExampleConversionRow 정규화 행으로 변환한다.
- 업로드 모달의 product_kind 값에 따라 세 가지 레이아웃을 분기한다.
  1) hanhwa_whole   : 종신보험
  2) hanhwa_annuity : 연금보험
  3) hanhwa_general : 일반보장

정규화 공통 원칙:
- 보험사 컬럼은 항상 "한화"
- 각 시트에서 "□ 주계약" 첫 번째 테이블만 정규화
- "독립특약" 또는 "종속특약" 문구가 포함된 행부터 아래는 제외
- 상품명은 시트명 사용
- 환산율은 백분율 수치로 저장
  예: raw 1.6 → Decimal("160.0000"), raw "160%" → Decimal("160.0000")
- 동일 납기 환산율은 year1~year4에 모두 동일하게 저장
"""

import logging
import re
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from typing import Iterable

from openpyxl.worksheet.worksheet import Worksheet

from commission.models import RateExampleConversionRow
from commission.services.rate_example_normalizers._common.text import clean_spaces

logger = logging.getLogger(__name__)

INSURER = "한화"
CATEGORY = "conv"

PRODUCT_KIND_WHOLE = "hanhwa_whole"
PRODUCT_KIND_ANNUITY = "hanhwa_annuity"
PRODUCT_KIND_GENERAL = "hanhwa_general"

VALID_PRODUCT_KINDS = {
    PRODUCT_KIND_WHOLE,
    PRODUCT_KIND_ANNUITY,
    PRODUCT_KIND_GENERAL,
}

STOP_KEYWORDS = ("독립특약", "종속특약")
MAIN_CONTRACT_KEYWORD = "주계약"

RATE_QUANT = Decimal("0.0001")


# =============================================================================
# 텍스트 / 숫자 유틸
# =============================================================================

def _clean_text(value) -> str:
    """셀 값을 비교·저장 가능한 단일 문자열로 정리한다."""
    return clean_spaces(str(value or "").replace("\u3000", " "))


def _row_text(ws: Worksheet, row_idx: int) -> str:
    """행 전체 텍스트를 합쳐 키워드 탐지에 사용한다."""
    return " ".join(
        _clean_text(ws.cell(row_idx, col_idx).value)
        for col_idx in range(1, ws.max_column + 1)
    ).strip()


def _contains_any(text: str, keywords: Iterable[str]) -> bool:
    return any(keyword in text for keyword in keywords)


def _to_decimal_percent(value) -> Decimal | None:
    """
    raw 환산율 셀 값을 백분율 Decimal로 변환한다.

    처리 예:
    - 1.6      → 160.0000
    - 0.85     → 85.0000
    - "160%"   → 160.0000
    - "160.0"  → 160.0000
    - "-" / "" → None
    """
    if value is None:
        return None

    text = _clean_text(value)
    if not text or text in {"-", "–", "—"}:
        return None

    has_percent = "%" in text
    text = text.replace("%", "").replace(",", "").strip()

    try:
        number = Decimal(text)
    except (InvalidOperation, TypeError, ValueError):
        return None

    # Excel percent format 또는 raw 소수 배율을 백분율 수치로 변환한다.
    if not has_percent and Decimal("0") < abs(number) <= Decimal("10"):
        number = number * Decimal("100")

    return number.quantize(RATE_QUANT, rounding=ROUND_HALF_UP)


def _make_row(
    *,
    example,
    source_sheet: str,
    source_row_no: int,
    coverage_type: str,
    product_name: str,
    plan_type: str,
    pay_period: str,
    rate: Decimal,
) -> RateExampleConversionRow:
    """RateExampleConversionRow 객체를 생성한다. DB 저장은 호출부 bulk_create가 담당한다."""
    return RateExampleConversionRow(
        source_file=example,
        source_sheet=source_sheet,
        source_row_no=source_row_no,
        insurer_type=example.insurer_type,
        category=CATEGORY,
        insurer=INSURER,
        coverage_type=coverage_type,
        strategy_flag="",
        product_name=product_name,
        plan_type=plan_type,
        pay_period=pay_period,
        year1=rate,
        year2=rate,
        year3=rate,
        year4=rate,
    )


# =============================================================================
# 시트 구조 탐지
# =============================================================================

def _find_main_contract_row(ws: Worksheet) -> int | None:
    """
    각 시트에서 첫 번째 "□ 주계약" 테이블 시작 행을 찾는다.

    raw 파일마다 "□ 주계약", "주계약" 등 표시 차이가 있을 수 있으므로
    행 전체 텍스트에서 "주계약" 포함 여부로 판단한다.
    """
    for row_idx in range(1, ws.max_row + 1):
        text = _row_text(ws, row_idx)
        if MAIN_CONTRACT_KEYWORD in text:
            return row_idx
    return None


def _find_stop_row(ws: Worksheet, start_row: int) -> int:
    """
    특약 시작 행을 찾는다.
    없으면 시트 마지막 행 다음 행을 반환해 range end로 사용한다.
    """
    for row_idx in range(start_row + 1, ws.max_row + 1):
        text = _row_text(ws, row_idx)
        if _contains_any(text, STOP_KEYWORDS):
            return row_idx
    return ws.max_row + 1


def _find_header_row_with_terms(
    ws: Worksheet,
    *,
    start_row: int,
    end_row: int,
    terms: tuple[str, ...],
) -> int | None:
    """start~end 범위에서 지정 단어가 포함된 헤더 행을 찾는다."""
    upper = min(end_row, start_row + 8, ws.max_row)
    for row_idx in range(start_row, upper + 1):
        values = [_clean_text(ws.cell(row_idx, col_idx).value) for col_idx in range(1, ws.max_column + 1)]
        if any(any(term in value for term in terms) for value in values):
            return row_idx
    return None


def _find_data_start_row(ws: Worksheet, header_row: int, stop_row: int) -> int:
    """
    헤더 다음 실제 데이터 시작 행을 찾는다.

    복합 헤더가 1~2행으로 구성되는 경우가 많으므로,
    header_row 이후 첫 번째 유효 데이터 행을 탐지한다.
    """
    for row_idx in range(header_row + 1, stop_row):
        text = _row_text(ws, row_idx)
        if not text:
            continue
        if "구분" in text and ("월납" in text or "비일시납" in text):
            continue
        if any(keyword in text for keyword in ("년납", "전기납", "월납")):
            # 연금보험은 "구분" 컬럼 자체가 납기인 경우가 있어 데이터일 수 있다.
            return row_idx
        if ws.cell(row_idx, 1).value is not None or ws.cell(row_idx, 2).value is not None:
            return row_idx
    return header_row + 1


def _is_cell_merged_across_cols(ws: Worksheet, row_idx: int, col_idx: int, target_cols: set[int]) -> bool:
    """
    특정 셀이 target_cols 범위를 가로로 병합하고 있는지 확인한다.

    한화 종신/일반보장 규칙:
    - A열과 B열이 병합된 형태이면 plan_type은 공란
    - A/B가 분리되어 있으면 B열 값을 plan_type으로 사용
    """
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row_idx <= merged.max_row and merged.min_col <= col_idx <= merged.max_col:
            covered_cols = set(range(merged.min_col, merged.max_col + 1))
            return target_cols.issubset(covered_cols)
    return False


# =============================================================================
# 종신보험 / 일반보장 공통 파서
# =============================================================================

def _find_non_single_pay_columns(
    ws: Worksheet,
    *,
    header_row: int,
    include_electric: bool,
) -> list[tuple[int, str]]:
    """
    메인 헤더 "비일시납" 하위의 세부 납기 컬럼을 찾는다.

    종신보험:
    - 6행 등 세부 헤더 중 "년납" 포함 컬럼만 사용

    일반보장:
    - "년납" 또는 "전기납" 포함 컬럼 사용
    """
    columns: list[tuple[int, str]] = []
    for col_idx in range(1, ws.max_column + 1):
        value = _clean_text(ws.cell(header_row, col_idx).value)
        if not value:
            continue

        has_year_pay = "년납" in value
        has_electric = include_electric and "전기납" in value

        if has_year_pay or has_electric:
            columns.append((col_idx, value))

    return columns


def _extract_plan_type_for_whole_or_general(ws: Worksheet, row_idx: int) -> str:
    """
    종신보험/일반보장 구분 컬럼 정규화.

    - A열과 B열이 병합되어 있으면 구분 공란
    - A/B가 분리되어 있으면 B열 데이터 사용
    """
    if _is_cell_merged_across_cols(ws, row_idx, 1, {1, 2}):
        return ""

    return _clean_text(ws.cell(row_idx, 2).value)


def _is_visible_sheet(ws: Worksheet) -> bool:
    """
    숨김 처리된 시트는 정규화에서 제외한다.

    openpyxl sheet_state:
    - visible
    - hidden
    - veryHidden
    """
    return getattr(ws, "sheet_state", "visible") == "visible"


def _should_use_plan_type(product_name: str) -> bool:
    """
    한화 일반보장 구분 컬럼 정책.

    - CEO정기 상품만 raw B열 구분을 사용한다.
    - 그 외 일반보장/종신/연금 시트는 구분을 공란으로 둔다.
    """
    return "정기" in product_name


def _build_whole_or_general_rows(
    *,
    example,
    workbook,
    product_kind: str,
) -> list[RateExampleConversionRow]:
    """
    종신보험 / 일반보장 raw 파일 정규화.

    product_kind:
    - hanhwa_whole   → 보종 전체 "종신/CI"
    - hanhwa_general → 보종 기본 "기타(보장성)", 상품명에 "경영" 포함 시 "CEO정기"
    """
    normalized: list[RateExampleConversionRow] = []
    include_electric = product_kind == PRODUCT_KIND_GENERAL

    for ws in workbook.worksheets:
        # 숨김/veryHidden 시트는 운영 정규화 대상에서 제외
        if not _is_visible_sheet(ws):
            continue

        product_name = _clean_text(ws.title)
        if not product_name:
            continue

        main_row = _find_main_contract_row(ws)
        if not main_row:
            logger.info("hanhwa normalizer: main contract not found. sheet=%s", ws.title)
            continue

        stop_row = _find_stop_row(ws, main_row)

        header_row = _find_header_row_with_terms(
            ws,
            start_row=main_row,
            end_row=stop_row,
            terms=("년납", "전기납"),
        )
        if not header_row:
            logger.info("hanhwa normalizer: pay-period header not found. sheet=%s", ws.title)
            continue

        pay_columns = _find_non_single_pay_columns(
            ws,
            header_row=header_row,
            include_electric=include_electric,
        )
        if not pay_columns:
            continue

        data_start = _find_data_start_row(ws, header_row, stop_row)

        coverage_type = "종신/CI"
        if product_kind == PRODUCT_KIND_GENERAL:
            coverage_type = "CEO정기" if "경영" in product_name else "기타(보장성)"

        for row_idx in range(data_start, stop_row):
            row_text = _row_text(ws, row_idx)
            if not row_text:
                continue
            if _contains_any(row_text, STOP_KEYWORDS):
                break

            # 구분 컬럼은 CEO정기 상품만 사용한다.
            # 나머지 시트는 수수료 계산 오류 방지를 위해 공란 처리한다.
            plan_type = (
                _extract_plan_type_for_whole_or_general(ws, row_idx)
                if _should_use_plan_type(product_name)
                else ""
            )

            for col_idx, pay_period in pay_columns:
                rate = _to_decimal_percent(ws.cell(row_idx, col_idx).value)
                if rate is None:
                    # 특정 납기 환산율 셀이 공란이면 해당 납기 row만 정규화 제외
                    continue

                normalized.append(_make_row(
                    example=example,
                    source_sheet=ws.title,
                    source_row_no=row_idx,
                    coverage_type=coverage_type,
                    product_name=product_name,
                    plan_type=plan_type,
                    pay_period=pay_period,
                    rate=rate,
                ))

    return normalized


# =============================================================================
# 연금보험 파서
# =============================================================================

def _coverage_for_annuity(product_name: str) -> str:
    """연금보험 보종을 상품명 키워드로 판별한다."""
    if "연금" in product_name and "변액" in product_name:
        return "변액연금"
    if "연금" in product_name and "저축" in product_name:
        return "연금저축"
    if "연금" in product_name:
        return "연금"
    return "연금"


def _find_column_by_header(ws: Worksheet, header_row: int, keyword: str) -> int | None:
    """헤더 행에서 특정 키워드가 포함된 첫 번째 컬럼을 찾는다."""
    for col_idx in range(1, ws.max_column + 1):
        if keyword in _clean_text(ws.cell(header_row, col_idx).value):
            return col_idx
    return None


def _find_annuity_header_row(ws: Worksheet, *, start_row: int, end_row: int) -> int | None:
    """
    연금보험 주계약 테이블의 실제 헤더 행을 찾는다.

    주의:
    - raw 4행에는 "(월납 P'대비율...)" 같은 안내 문구가 있어 "월납"이 포함된다.
    - 기존 공용 탐지 함수는 이 안내 행을 헤더로 오인할 수 있다.
    - 연금보험은 반드시 같은 행에 "구분" 헤더와 "월납" 헤더가 함께 있는 행만 인정한다.
    """
    upper = min(end_row, start_row + 8, ws.max_row)
    for row_idx in range(start_row, upper + 1):
        values = [
            _clean_text(ws.cell(row_idx, col_idx).value)
            for col_idx in range(1, ws.max_column + 1)
        ]
        has_period_header = any(value == "구분" for value in values)
        has_monthly_header = any("월납" in value for value in values)
        if has_period_header and has_monthly_header:
            return row_idx
    return None


def _build_annuity_rows(*, example, workbook) -> list[RateExampleConversionRow]:
    """
    연금보험 raw 파일 정규화.

    규칙:
    - 시트명에 "바로" 포함 시 제외
    - 상품명 = 시트명
    - 구분 = 공란
    - 납기 = 첫 번째 주계약 테이블의 "구분" 컬럼
    - 환산율 = "월납" 컬럼
    """
    normalized: list[RateExampleConversionRow] = []

    for ws in workbook.worksheets:
        # 숨김/veryHidden 시트는 운영 정규화 대상에서 제외
        if not _is_visible_sheet(ws):
            continue

        product_name = _clean_text(ws.title)
        if not product_name or "바로" in product_name:
            continue

        main_row = _find_main_contract_row(ws)
        if not main_row:
            continue

        stop_row = _find_stop_row(ws, main_row)

        header_row = _find_annuity_header_row(
            ws,
            start_row=main_row,
            end_row=stop_row,
        )
        if not header_row:
            logger.info("hanhwa annuity normalizer: header not found. sheet=%s", ws.title)
            continue

        pay_period_col = _find_column_by_header(ws, header_row, "구분") or 1
        monthly_rate_col = _find_column_by_header(ws, header_row, "월납")
        if not monthly_rate_col:
            logger.info("hanhwa annuity normalizer: monthly column not found. sheet=%s", ws.title)
            continue

        data_start = _find_data_start_row(ws, header_row, stop_row)
        coverage_type = _coverage_for_annuity(product_name)

        for row_idx in range(data_start, stop_row):
            row_text = _row_text(ws, row_idx)
            if not row_text:
                continue
            if _contains_any(row_text, STOP_KEYWORDS):
                break

            pay_period = _clean_text(ws.cell(row_idx, pay_period_col).value)
            rate = _to_decimal_percent(ws.cell(row_idx, monthly_rate_col).value)

            if not pay_period or rate is None:
                continue

            normalized.append(_make_row(
                example=example,
                source_sheet=ws.title,
                source_row_no=row_idx,
                coverage_type=coverage_type,
                product_name=product_name,
                plan_type="",
                pay_period=pay_period,
                rate=rate,
            ))

    return normalized


# =============================================================================
# 공개 진입점
# =============================================================================

def build_life_hanhwa_conversion_rows(
    example,
    workbook,
    *,
    product_kind: str,
) -> list[RateExampleConversionRow]:
    """
    한화생명 환산율/수정률 정규화 공개 함수.

    호출 위치:
        commission/services/rate_example_normalizer.py

    product_kind 필수:
        hanhwa_whole   : 종신보험
        hanhwa_annuity : 연금보험
        hanhwa_general : 일반보장
    """
    if product_kind not in VALID_PRODUCT_KINDS:
        raise ValueError("한화 상품 구분 값이 올바르지 않습니다.")

    if product_kind == PRODUCT_KIND_ANNUITY:
        rows = _build_annuity_rows(example=example, workbook=workbook)
    else:
        rows = _build_whole_or_general_rows(
            example=example,
            workbook=workbook,
            product_kind=product_kind,
        )

    logger.info(
        "hanhwa normalizer: created %d rows. pk=%s product_kind=%s",
        len(rows),
        getattr(example, "pk", None),
        product_kind,
    )
    return rows