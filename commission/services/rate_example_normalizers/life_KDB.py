# commission/services/rate_example_normalizers/life_kdb.py
from __future__ import annotations

"""
KDB 생명 환산율/수정률 정규화 parser.

역할:
- KDB(life) raw xlsx 파일에서 "GA 주계약" 시트만 정규화한다.
- 1~3행은 제외하고, 4행부터 데이터 행으로 처리한다.
- 정규화 결과는 RateExampleConversionRow master 테이블에 적재할 row 객체 목록으로 반환한다.

KDB raw 컬럼 매핑:
- C열: 상품명       -> product_name
- D열: 구분         -> 정규화 제외. plan_type은 공란 저장
- H열: 납기         -> pay_period
- I열: 연령/기준    -> H열 납기에 괄호로 병합
- K열: 변경후       -> year1~year4 동일 반영

보종 판정:
- 상품명에 "종신" 포함       -> 종신/CI
- 상품명에 "CEO" 포함        -> CEO정기
- 상품명에 "연금" 포함       -> 연금
- 상품명에 "연금" + "변액"   -> 변액연금
- 상품명에 "연금" + "저축"   -> 연금저축
- 그 외                     -> 기타(보장성)

저장 정책:
- 환산률은 DB에 백분율 수치 기준 Decimal로 저장한다.
  예: raw 100% -> Decimal("100"), raw 100.0 -> Decimal("100.0")
"""

import logging
from decimal import Decimal
from typing import Any

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from commission.models import RateExample, RateExampleConversionRow
from commission.services.rate_example_normalizers._common.excel import (
    build_merged_value_map,
    cell_value_with_merged,
)
from commission.services.rate_example_normalizers._common.decimal import (
    decimal_percent_value,
)

logger = logging.getLogger(__name__)

KDB_TARGET_SHEET = "GA 주계약"


def _cell_text(value: Any) -> str:
    """
    셀 값을 안전한 문자열로 변환한다.

    - None은 빈 문자열
    - 앞뒤 공백 제거
    - Excel 숫자형 20.0 같은 값은 "20"이 아니라 raw 의미 보존을 위해 str 기준 처리
    """
    if value is None:
        return ""
    return str(value).strip()


def _build_merged_cell_lookup(ws: Worksheet) -> dict[str, Any]:
    """
    병합 셀 값을 각 셀 좌표에 전파하기 위한 lookup 생성.

    KDB raw는 상품명(C열), 납기(H열)가 병합되어 여러 행을 점유한다.
    openpyxl은 병합 범위의 첫 셀에만 값이 있으므로,
    각 행을 독립 정규화하려면 병합 범위 내 모든 좌표에서
    좌상단 셀 값을 읽을 수 있어야 한다.
    """
    # legacy 함수명은 유지하되 실제 구현은 공통 helper를 사용한다.
    return build_merged_value_map(ws)


def _cell_value(ws: Worksheet, merged_values: dict[tuple[int, int], Any], row_no: int, col_no: int) -> Any:
    """
    일반 셀/병합 셀 값을 동일하게 읽는다.
    """
    return cell_value_with_merged(ws, merged_values, row_no, col_no)


def _build_pay_period(pay_period: str, age_basis: str) -> str:
    """
    납기(H열)에 연령/기준(I열)을 괄호로 결합한다.

    예:
    - H열: 3년만기
    - I열: 3년납
    - 결과: 3년만기(3년납)
    """
    pay_period = _cell_text(pay_period)
    age_basis = _cell_text(age_basis)
    if pay_period and age_basis:
        return f"{pay_period}({age_basis})"
    return pay_period or age_basis


def _to_decimal_percent(value: Any, *, number_format: str = "") -> Decimal | None:
    """
    KDB 변경후(K열) 값을 DB 저장용 Decimal로 변환한다.

    저장 기준:
    - DB에는 백분율 수치 기준으로 저장한다.
    - Excel 셀이 100% 표시이지만 실제값 1.0으로 읽히는 경우 number_format의 %를 보고 ×100 보정한다.
    - 문자열 "100%"는 "%" 제거 후 Decimal("100")으로 저장한다.
    """
    # KDB 정책: DB에는 백분율 표시 기준 Decimal을 저장한다.
    # 예: Excel value=1.0 + number_format='0%' → Decimal('100')
    return decimal_percent_value(value, number_format=number_format)


def _infer_kdb_coverage_type(product_name: str) -> str:
    """
    KDB 상품명 기반 보종 판정.

    주의:
    - "변액연금", "연금저축"은 일반 "연금"보다 먼저 판정해야 한다.
    - 사용자가 지정한 KDB 전용 분류 규칙을 이 함수에 집중시킨다.
    """
    name = product_name or ""

    if "연금" in name and "변액" in name:
        return "변액연금"
    if "연금" in name and "저축" in name:
        return "연금저축"
    if "종신" in name:
        return "종신/CI"
    if "CEO" in name.upper():
        return "CEO정기"
    if "연금" in name:
        return "연금"
    return "기타(보장성)"


def _get_sheet(wb: Workbook) -> Worksheet | None:
    """
    KDB 정규화 대상 시트를 조회한다.

    - 대상 시트명은 반드시 "GA 주계약"
    - 시트가 없으면 정규화 row 0건 반환
    """
    if KDB_TARGET_SHEET not in wb.sheetnames:
        logger.warning(
            "KDB rate example sheet not found: expected=%s sheets=%s",
            KDB_TARGET_SHEET,
            wb.sheetnames,
        )
        return None
    return wb[KDB_TARGET_SHEET]


def build_life_kdb_conversion_rows(
    example: RateExample,
    wb: Workbook,
) -> list[RateExampleConversionRow]:
    """
    KDB 생명 환산율/수정률 정규화 row를 생성한다.

    처리 규칙:
    - "GA 주계약" 시트만 정규화
    - 1~3행 제외, 4행부터 처리
    - 상품명(C열)이 비어 있으면 skip
    - 변경후(K열)이 비어 있거나 숫자 변환 불가하면 skip
    - K열 값을 year1~year4에 동일 반영
    """
    ws = _get_sheet(wb)
    if ws is None:
        return []

    rows: list[RateExampleConversionRow] = []
    seen_keys: set[tuple[str, str, str]] = set()
    merged_values = _build_merged_cell_lookup(ws)

    # openpyxl은 1-based index다.
    # C=3 상품명, H=8 납기, I=9 연령/기준, K=11 변경후
    for row_no in range(4, ws.max_row + 1):
        product_name = _cell_text(_cell_value(ws, merged_values, row_no, 3))
        if not product_name:
            continue

        rate_cell = ws.cell(row=row_no, column=11)
        rate_value = _to_decimal_percent(
            rate_cell.value,
            number_format=getattr(rate_cell, "number_format", ""),
        )
        if rate_value is None:
            continue

        coverage_type = _infer_kdb_coverage_type(product_name)
        # KDB는 정규화 테이블의 구분(plan_type)을 사용하지 않는다.
        plan_type = ""

        # 납기(H열)는 병합 셀 값을 행별로 전파하고,
        # 연령/기준(I열)이 있으면 납기 뒤에 괄호로 결합한다.
        pay_period = _build_pay_period(
            _cell_value(ws, merged_values, row_no, 8),
            _cell_value(ws, merged_values, row_no, 9),
        )

        # 최종 중복 제거 기준:
        # 상품명 + 구분 + 납기가 모두 같으면 같은 상품으로 판단한다.
        dedupe_key = (product_name, plan_type, pay_period)
        if dedupe_key in seen_keys:
            continue
        seen_keys.add(dedupe_key)

        rows.append(
            RateExampleConversionRow(
                source_file=example,
                source_sheet=KDB_TARGET_SHEET,
                source_row_no=row_no,
                insurer_type=example.insurer_type,
                category=example.category,
                insurer="KDB",
                coverage_type=coverage_type,
                strategy_flag="",
                product_name=product_name,
                plan_type=plan_type,
                pay_period=pay_period,
                year1=rate_value,
                year2=rate_value,
                year3=rate_value,
                year4=rate_value,
            )
        )

    return rows