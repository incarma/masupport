# django_ma/commission/services/rate_example_normalizers/life_samsung.py
from __future__ import annotations

"""
삼성생명 환산율/수정률 정규화 parser.

처리 대상:
- "보장성"
- "건강상해"
- "건강상해(" 포함 시트
- "연금저축"

정규화 정책:
- 보험사(insurer)는 "삼성"으로 저장한다.
- 원본 파일 저장/검증/DB 적재는 rate_example_normalizer.py가 담당한다.
- 본 파일은 workbook -> RateExampleConversionRow list 변환만 담당한다.
- 환산율은 프로젝트 정책대로 100% = Decimal("100") 기준으로 저장한다.
- Excel 셀이 % 서식이면 openpyxl 실제값 보정을 위해 x100 처리한다.
- 가로 병합은 좌측 셀만 값으로 인정한다.
- 세로 병합은 기존 점유 행 전체에 같은 값을 전파한다.
"""

from decimal import Decimal, InvalidOperation
import re
from typing import Any

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from commission.models import RateExample, RateExampleConversionRow


DEC4 = Decimal("0.0001")
NO_PLAN_TYPE = "사용안함"
SAMSUNG_INSURER = "삼성"


def _clean_text(value: Any) -> str:
    """셀 값을 DB 저장용 문자열로 정규화한다."""
    if value is None:
        return ""

    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    parts = [part.strip() for part in text.split("\n") if part.strip()]
    text = " ".join(parts) if parts else text.strip()
    return re.sub(r"\s+", " ", text).strip()


def _merged_value(ws: Worksheet, row: int, col: int) -> Any:
    """
    병합 셀 값을 반환한다.

    삼성 raw 규칙:
    - 가로 병합: 좌측 셀만 값 인정
    - 세로 병합: 점유 행 전체에 동일 값 전파
    """
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value

    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate not in merged_range:
            continue

        # 세로 병합 또는 혼합 병합에서는 좌측 열에 한해 행 방향 전파.
        if row >= merged_range.min_row and col == merged_range.min_col:
            return ws.cell(
                row=merged_range.min_row,
                column=merged_range.min_col,
            ).value

        # 가로 병합의 우측 셀은 공란으로 본다.
        return None

    return None


def _cell_text(ws: Worksheet, row: int, col: int) -> str:
    """병합 정책 적용 후 문자열을 반환한다."""
    return _clean_text(_merged_value(ws, row, col))


def _to_decimal(value: Any, *, number_format: str = "") -> Decimal | None:
    """
    raw 환산율 값을 Decimal("0.0001")로 변환한다.

    프로젝트 정책:
    - 120%는 Decimal("120")으로 저장한다.
    - Excel percent 서식의 0.8은 80으로 보정한다.
    """
    if value is None:
        return None

    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if not text or text in {"-", "－", "–"}:
            return None

        text = text.replace("%", "").strip()

        try:
            dec = Decimal(text)
        except InvalidOperation:
            return None

        return dec.quantize(DEC4)

    try:
        dec = Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return None

    if "%" in str(number_format or ""):
        dec *= Decimal("100")

    return dec.quantize(DEC4)


def _cell_decimal(ws: Worksheet, row: int, col: int) -> Decimal | None:
    """셀 값을 Decimal 환산율로 변환한다."""
    value = _merged_value(ws, row, col)
    number_format = ws.cell(row=row, column=col).number_format
    return _to_decimal(value, number_format=number_format)


def _pay_headers(
    ws: Worksheet,
    *,
    start_col: int,
    end_col: int,
    header_row: int,
) -> list[tuple[int, str]]:
    """납기 헤더 목록을 만든다."""
    headers: list[tuple[int, str]] = []

    for col in range(start_col, end_col + 1):
        header = _cell_text(ws, header_row, col)
        if not header:
            continue
        if header in {"납입기간", "-", "－", "–"}:
            continue
        headers.append((col, header))

    return headers


def _coverage_for_protection(product_name: str) -> str:
    """보장성 시트 보종 판정."""
    if "CEO" in product_name:
        return "CEO정기"
    if "종신" in product_name:
        return "종신/CI"
    return "기타(보장성)"


def _coverage_for_annuity(product_name: str) -> str:
    """연금저축 시트 보종 판정."""
    if "변액" in product_name:
        return "변액연금"
    if "연금" in product_name:
        return "연금"
    return "연금"


def _make_row(
    *,
    example: RateExample,
    ws: Worksheet,
    source_row_no: int,
    coverage_type: str,
    product_name: str,
    plan_type: str,
    pay_period: str,
    rate: Decimal,
) -> RateExampleConversionRow:
    """RateExampleConversionRow 생성을 중앙화한다."""
    return RateExampleConversionRow(
        source_file=example,
        source_sheet=ws.title,
        source_row_no=source_row_no,
        insurer_type=example.insurer_type,
        category=example.category,
        insurer=SAMSUNG_INSURER,
        coverage_type=coverage_type,
        strategy_flag="",
        product_name=product_name,
        plan_type=_clean_text(plan_type) or NO_PLAN_TYPE,
        pay_period=pay_period,
        year1=rate,
        year2=rate,
        year3=rate,
        year4=rate,
    )


def _append_unique(
    rows: list[RateExampleConversionRow],
    seen: set[tuple],
    row: RateExampleConversionRow,
) -> None:
    """파일 내부 완전 중복 row를 방지한다."""
    key = (
        row.source_sheet,
        row.coverage_type,
        row.product_name,
        row.plan_type,
        row.pay_period,
        row.year1,
        row.year2,
        row.year3,
        row.year4,
    )
    if key in seen:
        return

    seen.add(key)
    rows.append(row)


def _parse_standard_sheet(
    example: RateExample,
    ws: Worksheet,
    *,
    coverage_type: str | None,
    exclude_silson: bool = False,
) -> list[RateExampleConversionRow]:
    """
    "보장성", "건강상해" 단독 시트 정규화.

    공통 매핑:
    - 상품명: F열
    - 구분: G열
    - 납기 헤더: 9행 I~P열
    - 환산율 값: 각 납기 컬럼의 row 값
    """
    rows: list[RateExampleConversionRow] = []
    seen: set[tuple] = set()
    pay_headers = _pay_headers(ws, start_col=9, end_col=16, header_row=9)

    for row_no in range(10, ws.max_row + 1):
        product_name = _cell_text(ws, row_no, 6)
        if not product_name or product_name == "상품명":
            continue

        if exclude_silson and "실손" in product_name:
            continue

        plan_type = _cell_text(ws, row_no, 7)
        row_coverage = coverage_type or _coverage_for_protection(product_name)

        for col, pay_period in pay_headers:
            rate = _cell_decimal(ws, row_no, col)
            if rate is None:
                continue

            normalized = _make_row(
                example=example,
                ws=ws,
                source_row_no=row_no,
                coverage_type=row_coverage,
                product_name=product_name,
                plan_type=plan_type,
                pay_period=pay_period,
                rate=rate,
            )
            _append_unique(rows, seen, normalized)

    return rows


def _parse_named_health_sheet(
    example: RateExample,
    ws: Worksheet,
) -> list[RateExampleConversionRow]:
    """
    "건강상해(...)" 시트 정규화.

    매핑:
    - 상품명: F열
    - 구분: 사용안함
    - 보종: 기타(보장성)
    - 납기 헤더: 10행 I~P열
    - 환산율 값: 각 납기 컬럼의 row 값

    제외:
    - 상품명 공란
    - 상품명에 "특약" 포함
    - 상품명에 "판매중지" 포함
    - "판매중지" 행 바로 위 source row
    """
    candidate_rows: list[tuple[int, RateExampleConversionRow]] = []
    excluded_source_rows: set[int] = set()
    seen: set[tuple] = set()

    pay_headers = _pay_headers(ws, start_col=9, end_col=16, header_row=10)

    for row_no in range(11, ws.max_row + 1):
        product_name = _cell_text(ws, row_no, 6)
        if not product_name:
            continue

        if "판매중지" in product_name:
            excluded_source_rows.add(row_no)
            excluded_source_rows.add(row_no - 1)
            continue

        if "특약" in product_name:
            continue

        for col, pay_period in pay_headers:
            rate = _cell_decimal(ws, row_no, col)
            if rate is None:
                continue

            normalized = _make_row(
                example=example,
                ws=ws,
                source_row_no=row_no,
                coverage_type="기타(보장성)",
                product_name=product_name,
                plan_type=NO_PLAN_TYPE,
                pay_period=pay_period,
                rate=rate,
            )
            candidate_rows.append((row_no, normalized))

    rows: list[RateExampleConversionRow] = []
    for row_no, normalized in candidate_rows:
        if row_no in excluded_source_rows:
            continue
        _append_unique(rows, seen, normalized)

    return rows


def _parse_annuity_sheet(
    example: RateExample,
    ws: Worksheet,
) -> list[RateExampleConversionRow]:
    """
    "연금저축" 시트 정규화.

    매핑:
    - 상품명: F열
    - 구분: G열
    - 납기 헤더: 9행 H~M열
    - 환산율 값: 각 납기 컬럼의 row 값
    """
    rows: list[RateExampleConversionRow] = []
    seen: set[tuple] = set()
    pay_headers = _pay_headers(ws, start_col=8, end_col=13, header_row=9)

    for row_no in range(10, ws.max_row + 1):
        product_name = _cell_text(ws, row_no, 6)
        if not product_name:
            continue

        plan_type = _cell_text(ws, row_no, 7)

        for col, pay_period in pay_headers:
            rate = _cell_decimal(ws, row_no, col)
            if rate is None:
                continue

            normalized = _make_row(
                example=example,
                ws=ws,
                source_row_no=row_no,
                coverage_type=_coverage_for_annuity(product_name),
                product_name=product_name,
                plan_type=plan_type,
                pay_period=pay_period,
                rate=rate,
            )
            _append_unique(rows, seen, normalized)

    return rows


def build_life_samsung_conversion_rows(
    example: RateExample,
    workbook: Workbook,
) -> list[RateExampleConversionRow]:
    """
    삼성생명 환산율/수정률 정규화 진입점.

    반환:
    - 아직 DB에 저장하지 않은 RateExampleConversionRow 인스턴스 list
    """
    rows: list[RateExampleConversionRow] = []

    for ws in workbook.worksheets:
        sheet_name = str(ws.title or "").strip()

        if sheet_name == "보장성":
            rows.extend(
                _parse_standard_sheet(
                    example,
                    ws,
                    coverage_type=None,
                    exclude_silson=False,
                )
            )
            continue

        if sheet_name == "건강상해":
            rows.extend(
                _parse_standard_sheet(
                    example,
                    ws,
                    coverage_type="기타(보장성)",
                    exclude_silson=True,
                )
            )
            continue

        if sheet_name.startswith("건강상해("):
            rows.extend(_parse_named_health_sheet(example, ws))
            continue

        if sheet_name == "연금저축":
            rows.extend(_parse_annuity_sheet(example, ws))
            continue

    return rows