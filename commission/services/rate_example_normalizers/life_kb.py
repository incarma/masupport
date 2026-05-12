# commission/services/rate_example_normalizers/life_kb.py
from __future__ import annotations

"""
KB 생명보험 환산율/수정률 정규화 모듈.

지원 범위:
- 손생구분: 생명보험
- 구분: 환산율/수정률
- 보험사: KB
- 상품 구분: 일반상품

정규화 규칙:
1. 정규화 테이블의 보험사 컬럼에 "KB" 삽입
2. raw 파일 내 "(주계약)" 테이블만 정규화
   - "(특약)" 테이블은 사용하지 않음
3. 보종:
   - 상품(B열)에 "변액" 포함 → "변액연금"
   - 상품(B열)에 "연금" 포함 → "연금"
   - 상품(B열)에 "경영" 포함 → "CEO정기"
   - 상품(B열)에 "정기" 포함 → "종신/CI"
   - 그 외 → "종신/CI"
4. 상품명: 상품(B열)
5. 구분:
   - 기본값: 나이/보험료(K열)
   - 최소보험료(D열)에 데이터가 있으면 D열 사용
   - 가입금액(E열)에 데이터가 있으면 E열 사용
   - D/E 모두 "-"이면 공란
6. 납기: 납입기간(C열)
7. 환산율:
   - 1차년: F열
   - 2차년: G열
   - 3차년: H열
   - 4차년: I열
"""

from decimal import Decimal, InvalidOperation
import logging
import re
from typing import Any

from openpyxl.workbook.workbook import Workbook

from commission.models import RateExample, RateExampleConversionRow


logger = logging.getLogger(__name__)


RIDER_KEYWORD = "특약"
PAREN_RE = re.compile(r"\(([^()]*)\)")


def _clean_text(value: Any) -> str:
    """셀 값을 비교/저장 가능한 문자열로 정규화한다."""
    if value is None:
        return ""

    text = str(value).strip()
    if not text:
        return ""

    # Excel에서 숫자 코드/금액이 1000.0처럼 들어오는 경우 표시를 정리한다.
    if text.endswith(".0"):
        try:
            dec = Decimal(text)
            if dec == dec.to_integral_value():
                return str(dec.to_integral_value())
        except (InvalidOperation, ValueError):
            logger.warning(
                "KB text cleanup decimal conversion skipped: value=%r",
                value,
                exc_info=True,
            )

    return text


def _is_blank_or_dash(value: Any) -> bool:
    """빈 값 또는 '-' 계열 값을 비어있는 값으로 판정한다."""
    text = _clean_text(value)
    return text in {"", "-", "–", "—"}


def _to_decimal(value: Any) -> Decimal | None:
    """
    환산율 셀을 Decimal로 변환한다.

    처리 예:
    - 100.0 → Decimal("100.0")
    - "100.0%" → Decimal("100.0")
    - "-" / "" → None
    """
    if value is None:
        return None

    if isinstance(value, Decimal):
        return value

    text = _clean_text(value)
    if _is_blank_or_dash(text):
        return None

    text = text.replace("%", "").replace(",", "").strip()
    if not text:
        return None

    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def _rate_cell_to_decimal(cell) -> Decimal | None:
    """
    KB 환산율 셀을 raw 화면 표시 기준 백분율 숫자로 변환한다.

    Excel 백분율 서식 셀은 openpyxl에서 표시값 336.0%가 아니라
    실제값 3.36으로 읽힌다. number_format에 '%'가 있으면 ×100 해서
    DB에는 Decimal("336.0") 기준으로 저장한다.
    """
    value = cell.value
    dec = _to_decimal(value)
    if dec is None:
        return None

    number_format = str(getattr(cell, "number_format", "") or "")
    if "%" in number_format and "%" not in str(value):
        return dec * Decimal("100")

    return dec


def _has_any_rate_cell(*cells: Any) -> bool:
    """1~4차년 셀 중 하나라도 유효 환산율이 있는지 확인한다."""
    return any(_rate_cell_to_decimal(cell) is not None for cell in cells)


def _coverage_type(product_name: str) -> str:
    """
    KB 일반상품 보종 매핑.

    우선순위:
    - "변액"은 "연금"보다 먼저 판정해야 변액연금이 연금으로 오분류되지 않는다.
    - "경영"은 "정기"보다 먼저 판정해야 CEO정기로 분류된다.
    - 그 외는 기본값 "종신/CI"로 정규화한다.
    """
    name = product_name or ""
    if "변액" in name:
        return "변액연금"
    if "연금" in name:
        return "연금"
    if "경영" in name:
        return "CEO정기"
    if "정기" in name:
        return "종신/CI"
    return "종신/CI"


def _split_product_and_plan(raw_product: Any) -> tuple[str, str]:
    """
    KB 건강보험 상품(C열)을 상품명/구분으로 분리한다.

    - 상품명: 괄호 밖 텍스트
    - 구분: 괄호 안 텍스트 전체, 2개 이상이면 콤마로 결합
    """
    text = _clean_text(raw_product)
    if not text:
        return "", ""

    plans = [
        _clean_text(match.group(1))
        for match in PAREN_RE.finditer(text)
        if _clean_text(match.group(1))
    ]
    product_name = PAREN_RE.sub("", text).strip()
    product_name = re.sub(r"\s+", " ", product_name)

    return product_name, ", ".join(plans)


def _is_health_header_or_noise_row(
    *,
    raw_type: Any,
    raw_product: Any,
    pay_period: Any,
) -> bool:
    """
    KB 건강보험 raw의 헤더/노이즈 행을 제외한다.

    건강보험 raw는 일부 시트에서 4~5행에 헤더가 있고 6행부터 데이터가 시작된다.
    단순히 1~3행만 제외하면 C열 헤더 '상품'이 정규화 row로 들어갈 수 있으므로
    컬럼 헤더 텍스트를 명시적으로 차단한다.
    """
    type_text = _clean_text(raw_type)
    product_text = _clean_text(raw_product)
    pay_period_text = _clean_text(pay_period)

    if not type_text and not product_text and not pay_period_text:
        return True

    if product_text in {"상품", "상품명"}:
        return True

    if pay_period_text in {"납입기간", "납기"}:
        return True

    if type_text in {"구분"}:
        return True

    return False


def _plan_type(*, age_premium: Any, min_premium: Any, insured_amount: Any) -> str:
    """
    정규화 테이블의 구분 컬럼 산출.

    요구사항 문맥상 D/E 컬럼이 K열보다 우선한다.
    - D열 최소보험료에 데이터가 있으면 D열 사용
    - E열 가입금액에 데이터가 있으면 E열 사용
    - D/E 모두 '-'이면 공란
    - D/E 모두 빈 값이고 K열이 있으면 K열 사용
    """
    d_val = "" if _is_blank_or_dash(min_premium) else _clean_text(min_premium)
    e_val = "" if _is_blank_or_dash(insured_amount) else _clean_text(insured_amount)
    k_val = "" if _is_blank_or_dash(age_premium) else _clean_text(age_premium)

    if d_val:
        return d_val
    if e_val:
        return e_val

    # D/E가 명시적으로 '-'이면 공란 처리
    if _is_blank_or_dash(min_premium) and _is_blank_or_dash(insured_amount):
        return ""

    return k_val


def _is_header_or_noise_row(product_name: str) -> bool:
    """
    KB raw의 제목/헤더/구분행을 제외한다.

    정규화 대상은 B열 상품명 기준이다.
    """
    text = _clean_text(product_name)
    if not text:
        return True

    if text in {"상품", "상품명"}:
        return True

    # 표 제목/구분 행 제외
    if "주계약" in text:
        return True

    # 특약 상품 또는 특약 표시는 정규화 제외
    if RIDER_KEYWORD in text:
        return True

    return False


def build_life_kb_general_conversion_rows(
    example: RateExample,
    wb: Workbook,
) -> list[RateExampleConversionRow]:
    """
    KB 생명보험 일반상품 환산율/수정률 raw workbook을 정규화 행 목록으로 변환한다.

    DB 저장은 호출부(rate_example_normalizer.py)가 담당한다.
    """
    rows: list[RateExampleConversionRow] = []

    for ws in wb.worksheets:
        # ─────────────────────────────────────────────────────
        # KB 일반상품 단순 스캔 방식
        # - B열 상품명 기준으로 전체 행을 스캔한다.
        # - B열에 '특약' 문구가 포함된 행은 제외한다.
        # - 나머지 유효 상품 행만 정규화한다.
        # - 헤더 위치/주계약 marker 탐지에 의존하지 않는다.
        # ─────────────────────────────────────────────────────
        # raw 1~4행은 제목/안내/공백 영역으로 보고 정규화에서 제외한다.
        for row_no in range(5, ws.max_row + 1):
            product_name = _clean_text(ws.cell(row_no, 2).value)      # B: 상품
            pay_period = _clean_text(ws.cell(row_no, 3).value)        # C: 납입기간
            min_premium = ws.cell(row_no, 4).value                    # D: 최소보험료
            insured_amount = ws.cell(row_no, 5).value                 # E: 가입금액
            year1_cell = ws.cell(row_no, 6)                           # F: 1차년
            year2_cell = ws.cell(row_no, 7)                           # G: 2차년
            year3_cell = ws.cell(row_no, 8)                           # H: 3차년
            year4_cell = ws.cell(row_no, 9)                           # I: 4차년
            age_premium = ws.cell(row_no, 11).value                   # K: 나이/보험료

            # 빈 행/소계/헤더 잔여 행 방어
            if not product_name and not pay_period and not _has_any_rate_cell(
                year1_cell,
                year2_cell,
                year3_cell,
                year4_cell,
            ):
                continue

            if _is_header_or_noise_row(product_name):
                continue

            rows.append(
                RateExampleConversionRow(
                    source_file=example,
                    source_sheet=ws.title,
                    source_row_no=row_no,
                    insurer_type=example.insurer_type,
                    category=example.category,
                    insurer="KB",
                    coverage_type=_coverage_type(product_name),
                    strategy_flag="",
                    product_name=product_name,
                    plan_type=_plan_type(
                        age_premium=age_premium,
                        min_premium=min_premium,
                        insured_amount=insured_amount,
                    ),
                    pay_period=pay_period,
                    year1=_rate_cell_to_decimal(year1_cell),
                    year2=_rate_cell_to_decimal(year2_cell),
                    year3=_rate_cell_to_decimal(year3_cell),
                    year4=_rate_cell_to_decimal(year4_cell),
                )
            )

    return rows


def build_life_kb_health_conversion_rows(
    example: RateExample,
    wb: Workbook,
) -> list[RateExampleConversionRow]:
    """
    KB 생명보험 건강보험 환산율/수정률 raw workbook을 정규화 행 목록으로 변환한다.

    규칙:
    - 모든 시트 정규화
    - 각 시트 1~3행 제외
    - B열 구분 값에 '특약'이 포함된 행부터 해당 시트 하단 전체 제외
    - 보험사: KB
    - 보종: 기타(보장성)
    - 상품명: C열 괄호 밖 텍스트
    - 구분: C열 괄호 안 텍스트 전체, 2개 이상이면 콤마 결합
    - 납기: D열
    - 1~4차년: E~H열
    """
    rows: list[RateExampleConversionRow] = []

    for ws in wb.worksheets:
        for row_no in range(4, ws.max_row + 1):
            raw_type = _clean_text(ws.cell(row_no, 2).value)       # B: 구분
            raw_product = ws.cell(row_no, 3).value                 # C: 상품
            pay_period = _clean_text(ws.cell(row_no, 4).value)     # D: 납입기간
            year1_cell = ws.cell(row_no, 5)                        # E: 1차년
            year2_cell = ws.cell(row_no, 6)                        # F: 2차년
            year3_cell = ws.cell(row_no, 7)                        # G: 3차년
            year4_cell = ws.cell(row_no, 8)                        # H: 4차년

            if RIDER_KEYWORD in raw_type:
                break

            if _is_health_header_or_noise_row(
                raw_type=raw_type,
                raw_product=raw_product,
                pay_period=pay_period,
            ):
                continue

            product_name, plan_type = _split_product_and_plan(raw_product)

            if not product_name and not pay_period and not _has_any_rate_cell(
                year1_cell,
                year2_cell,
                year3_cell,
                year4_cell,
            ):
                continue

            if not product_name:
                continue

            rows.append(
                RateExampleConversionRow(
                    source_file=example,
                    source_sheet=ws.title,
                    source_row_no=row_no,
                    insurer_type=example.insurer_type,
                    category=example.category,
                    insurer="KB",
                    coverage_type="기타(보장성)",
                    strategy_flag="",
                    product_name=product_name,
                    plan_type=plan_type,
                    pay_period=pay_period,
                    year1=_rate_cell_to_decimal(year1_cell),
                    year2=_rate_cell_to_decimal(year2_cell),
                    year3=_rate_cell_to_decimal(year3_cell),
                    year4=_rate_cell_to_decimal(year4_cell),
                )
            )

    return rows