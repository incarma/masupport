# django_ma/commission/services/rate_example_normalizers/life_im.py
from __future__ import annotations

"""
IM생명 환산율/수정률 raw 정규화.

대상:
- 생명보험 / 환산율·수정률 / IM / xlsx

정규화 규칙:
- 첫 번째 시트 "(총괄)환산성적표"만 사용한다.
- E열 "구분" 값이 "주계약"인 행만 정규화한다.
- 보험사 컬럼은 "IM" 고정.
- 보종은 F열 "주계약" 상품명 기준으로 판정한다.
- L열 "기본형" 값을 1차년~4차년에 모두 동일하게 저장한다.
"""

from decimal import Decimal, InvalidOperation
from numbers import Number

from openpyxl.workbook.workbook import Workbook

from commission.models import RateExample, RateExampleConversionRow

SHEET_IM_SUMMARY = "(총괄)환산성적표"

# 원본 컬럼 위치
COL_KIND = 5          # E열: 구분
COL_PRODUCT = 6       # F열: 주계약
COL_GA = 8            # H열: GA
COL_PAY_PERIOD = 10   # J열: 납기
COL_PLAN = 11         # K열: 나이/보험료
COL_BASIC_RATE = 12   # L열: 기본형


def _clean_text(value) -> str:
    """엑셀 셀 값을 비교/저장 가능한 문자열로 정규화한다."""
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def _format_excel_display_text(cell) -> str:
    """
    엑셀 표시값에 가까운 문자열을 만든다.

    주요 목적:
    - IM raw의 납기(J열)가 실제 화면에서는 "20년 이상"처럼 보이지만
      openpyxl cell.value는 20으로 읽히는 케이스 보정
    """
    value = cell.value
    if value is None:
        return ""

    number_format = str(getattr(cell, "number_format", "") or "")

    # ── IM 납기 전용 보정: 20 -> 20년 이상 ─────────────────────
    if isinstance(value, Number) and "년" in number_format:
        if float(value).is_integer():
            base = str(int(value))
        else:
            base = str(value)

        if "이상" in number_format:
            return f"{base}년 이상"
        return f"{base}년"

    return _clean_text(value)


def _to_percent_decimal(cell):
    """
    환산율 값을 백분율 기준 Decimal로 변환한다.

    저장 기준:
    - raw 표시값이 126%이면 DB에는 Decimal("126") 저장
    - 추후 계산 시에는 year / 100 으로 비례 적용
    """
    value = cell.value
    if value is None:
        return None

    number_format = str(getattr(cell, "number_format", "") or "")
    text = _clean_text(value).replace(",", "")
    if not text or text in {"-", "–"}:
        return None

    try:
        if text.endswith("%"):
            return Decimal(text[:-1].strip())

        decimal_value = Decimal(text)

        # 엑셀 셀 표시 형식이 %이면 openpyxl 값은 보통 1.26,
        # 화면 표시값은 126%이므로 100을 곱해 저장한다.
        if "%" in number_format:
            return decimal_value * Decimal("100")

        return decimal_value
    except (InvalidOperation, ValueError):
        return None


def _coverage_type_for_im(product_name: str) -> str:
    """
    F열 주계약 상품명 기준 보종 판정.

    주의:
    - "변액"은 "연금"보다 우선한다.
      예: 변액연금 상품은 "연금"이 아니라 "변액연금"으로 저장한다.
    """
    name = _clean_text(product_name)

    if "변액" in name:
        return "변액연금"
    if "종신" in name:
        return "종신/CI"
    if "연금" in name:
        return "연금"
    return "기타(보장성)"


def _strategy_flag_for_im(ga_value: str) -> str:
    """
    H열 GA 상품분류 값을 정규화 테이블의 전략유무 값으로 변환한다.

    raw 파일에서 로마 숫자가 ASCII(I, II...) 또는 유니코드(Ⅰ, Ⅱ...)로
    들어올 수 있으므로 둘 다 허용한다.
    """
    text = _clean_text(ga_value).replace(" ", "")

    mapping = {
        "전략상품I": "전략상품1",
        "전략상품Ⅰ": "전략상품1",
        "전략상품II": "전략상품2",
        "전략상품Ⅱ": "전략상품2",
        "전략상품III": "전략상품3",
        "전략상품Ⅲ": "전략상품3",
        "전략상품IV": "전략상품4",
        "전략상품Ⅳ": "전략상품4",
    }
    return mapping.get(text, "")


def build_life_im_conversion_rows(
    example: RateExample,
    wb: Workbook,
) -> list[RateExampleConversionRow]:
    """
    IM생명 환산율/수정률 정규화 row 생성.

    반환:
    - DB 저장 전 RateExampleConversionRow 인스턴스 목록
    """
    if not wb.worksheets:
        raise ValueError("[rate_example][IM] workbook에 시트가 없습니다.")

    ws = wb.worksheets[0]

    if ws.title != SHEET_IM_SUMMARY:
        raise ValueError(
            "[rate_example][IM] 첫 번째 시트명이 올바르지 않습니다. "
            f"expected={SHEET_IM_SUMMARY!r}, actual={ws.title!r}"
        )

    rows: list[RateExampleConversionRow] = []

    for row_no in range(1, ws.max_row + 1):
        kind = _clean_text(ws.cell(row_no, COL_KIND).value)

        # E열 구분이 "주계약"인 행만 정규화한다.
        if kind != "주계약":
            continue

        product_name = _clean_text(ws.cell(row_no, COL_PRODUCT).value)
        plan_type = _clean_text(ws.cell(row_no, COL_PLAN).value)
        pay_period = _format_excel_display_text(ws.cell(row_no, COL_PAY_PERIOD))
        strategy_flag = _strategy_flag_for_im(ws.cell(row_no, COL_GA).value)
        basic_rate_cell = ws.cell(row_no, COL_BASIC_RATE)
        basic_rate_raw = _clean_text(basic_rate_cell.value)

        # ── 제외 조건 ─────────────────────────────────────────
        # IM raw의 L열 "기본형" 값이 "미판매"인 주계약 행은
        # 환산율 계산에 사용할 수 없으므로 정규화 대상에서 제외한다.
        if basic_rate_raw == "미판매":
            continue

        basic_rate = _to_percent_decimal(basic_rate_cell)

        # 상품명과 환산율이 모두 없으면 의미 없는 행으로 보고 제외한다.
        if not product_name and basic_rate is None:
            continue

        rows.append(
            RateExampleConversionRow(
                source_file=example,
                source_sheet=SHEET_IM_SUMMARY,
                source_row_no=row_no,
                insurer_type=example.insurer_type,
                category=example.category,
                insurer="IM",
                coverage_type=_coverage_type_for_im(product_name),
                strategy_flag=strategy_flag,
                product_name=product_name,
                plan_type=plan_type,
                pay_period=pay_period,
                year1=basic_rate,
                year2=basic_rate,
                year3=basic_rate,
                year4=basic_rate,
            )
        )

    return rows