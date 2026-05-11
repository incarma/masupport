# commission/services/rate_example_normalizers/life_kyobo.py
from __future__ import annotations

"""
교보 생명보험 환산율/수정률 정규화 parser.

역할:
- 교보 raw 예시표 중 "주계약(종속특약포함)" 시트만 정규화한다.
- 종신보험(B~F열), CI보험(H~L열), 연금보험(N~R열),
  정기보험(Z~AC열), 건강/어린이/기타보장(AE~AH열) 테이블을 정규화한다.
- 정규화 결과는 RateExampleConversionRow master 테이블에 적재할 row 객체 목록으로 반환한다.

정규화 정책 공통:
- 보험사: "교보"
- 시트: "주계약(종속특약포함)"
- 데이터 행: 6행부터 각 테이블 환산율 열의 마지막 데이터 행까지
- 5행은 헤더로 정규화 제외
- 상품명:
  - "판매중지" 또는 "특약" 포함 상품 제외 (하위 공란 행도 제외)
  - 공란 시 상단 마지막 상품명 전파
  - "(" 로 시작하고 ")" 로 끝나는 단독 키워드 → 직전 상품명에 합성
- 환산율: Excel % 셀은 number_format에 "%" 있으면 ×100 보정하여 백분율 수치로 저장

테이블별 정책:
- 종신보험(B~F열): 보종=종신/CI, 구분=E열, 납기=D열, 환산율=F열
- CI보험(H~L열): 보종=종신/CI, 구분=K열, 납기=J열, 환산율=L열
- 연금보험(N~R열): 보종=연금(변액→변액연금, 저축→연금저축), 구분=Q열, 납기=P열, 환산율=R열
- 정기보험(Z~AC열): 보종=종신/CI(경영→CEO정기), 구분=없음, 납기=AB열, 환산율=AC열
- 건강/어린이/기타보장(AE~AH열): 보종=기타(보장성), 구분=없음, 납기=AG열, 환산율=AH열
"""

import logging
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl.workbook.workbook import Workbook

from commission.models import RateExample, RateExampleConversionRow

logger = logging.getLogger(__name__)

TARGET_SHEET_NAME = "주계약(종속특약포함)"
DATA_START_ROW = 6


# =============================================================================
# 공통 헬퍼
# =============================================================================

def _text(value: Any) -> str:
    """Excel 셀 값을 문자열로 안전하게 정규화한다."""
    if value is None:
        return ""
    return str(value).strip()


def _should_exclude(product_name: str) -> bool:
    """
    상품명에 판매중지 또는 특약이 포함되어 있는지 판정한다.
    해당 상품 및 그 하위 공란 행은 정규화에서 제외한다.
    """
    name = _text(product_name)
    return "판매중지" in name or "특약" in name


def _is_subtype_keyword(value: str) -> bool:
    """
    셀 값이 직전 상품명의 서브타입 키워드인지 판정한다.

    판정 기준:
    - 문자열 전체가 "(" 로 시작하고 ")" 로 끝나고
    - 내부에 추가 괄호가 없는 경우 (단일 괄호 쌍)

    예: "(기본형)", "(체증형)" → True
    예: "(무)교보바로받는웰스연금(거치형)" → False (내부에 추가 괄호 있음)
    예: "기본형(플러스),보장강화형(플러스)" → False (괄호로 시작 안 함)
    """
    v = value.strip()
    if not (v.startswith("(") and v.endswith(")")):
        return False
    inner = v[1:-1]
    return "(" not in inner and ")" not in inner


def _to_decimal_percent(cell) -> Decimal | None:
    """
    환산율 셀 값을 DB 저장 정책에 맞는 백분율 수치 Decimal로 변환한다.

    Excel % 셀 처리:
    - value=0.75, number_format='0%' → Decimal("75.00")
    - value=150, number_format='General' → Decimal("150")
    """
    value = getattr(cell, "value", None)
    if value is None or value == "":
        return None

    try:
        dec = Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return None

    number_format = str(getattr(cell, "number_format", "") or "")
    if "%" in number_format:
        dec = dec * Decimal("100")

    return dec


def _last_data_row(ws, rate_col: int) -> int:
    """환산율 열(rate_col)의 마지막 데이터 행을 찾는다."""
    for row_no in range(ws.max_row, DATA_START_ROW - 1, -1):
        if _to_decimal_percent(ws.cell(row_no, rate_col)) is not None:
            return row_no
    return DATA_START_ROW - 1


def _resolve_product(
    product_raw: str,
    last_product_name: str,
    last_product_is_stopped: bool,
) -> tuple[str, bool, str, bool]:
    """
    상품명 공란/서브타입 전파 로직.

    반환: (current_product_name, current_is_stopped, new_last_name, new_last_stopped)
    """
    if product_raw:
        if _is_subtype_keyword(product_raw):
            # 서브타입 키워드: 직전 상품명 뒤에 합성
            current_name = last_product_name + product_raw
            current_stopped = _should_exclude(current_name)
        else:
            current_name = product_raw
            current_stopped = _should_exclude(product_raw)
        return current_name, current_stopped, current_name, current_stopped
    else:
        # 공란: 직전 상품명 전파
        return last_product_name, last_product_is_stopped, last_product_name, last_product_is_stopped


def _make_row(
    example: RateExample,
    source_sheet: str,
    source_row_no: int,
    coverage_type: str,
    product_name: str,
    plan_type: str,
    pay_period: str,
    rate_value: Decimal,
) -> RateExampleConversionRow:
    """RateExampleConversionRow 객체를 생성한다."""
    return RateExampleConversionRow(
        source_file=example,
        source_sheet=source_sheet,
        source_row_no=source_row_no,
        insurer_type=example.insurer_type,
        category=example.category,
        insurer="교보",
        coverage_type=coverage_type,
        strategy_flag="",
        product_name=product_name,
        plan_type=plan_type,
        pay_period=pay_period,
        year1=rate_value,
        year2=rate_value,
        year3=rate_value,
        year4=rate_value,
    )


# =============================================================================
# 테이블별 parser
# =============================================================================

def _parse_table(
    example: RateExample,
    ws,
    *,
    product_col: int,
    plan_type_col: int | None,
    pay_period_col: int,
    rate_col: int,
    coverage_type_fn,
) -> list[RateExampleConversionRow]:
    """
    단일 테이블 정규화 공통 로직.

    coverage_type_fn: (product_name: str) -> str
    """
    end_row = _last_data_row(ws, rate_col)
    if end_row < DATA_START_ROW:
        return []

    rows: list[RateExampleConversionRow] = []
    last_product_name = ""
    last_product_is_stopped = False

    for row_no in range(DATA_START_ROW, end_row + 1):
        product_raw = _text(ws.cell(row_no, product_col).value)
        pay_period = _text(ws.cell(row_no, pay_period_col).value)
        plan_type = _text(ws.cell(row_no, plan_type_col).value) if plan_type_col else ""
        rate_value = _to_decimal_percent(ws.cell(row_no, rate_col))

        # 환산율 없는 행은 정규화 대상 아님
        if rate_value is None:
            continue

        current_name, current_stopped, last_product_name, last_product_is_stopped = (
            _resolve_product(product_raw, last_product_name, last_product_is_stopped)
        )

        if not current_name:
            continue

        # 판매중지/특약 상품 및 하위 공란 행 제외
        if current_stopped:
            continue

        coverage_type = coverage_type_fn(current_name)

        rows.append(
            _make_row(
                example,
                source_sheet=TARGET_SHEET_NAME,
                source_row_no=row_no,
                coverage_type=coverage_type,
                product_name=current_name,
                plan_type=plan_type,
                pay_period=pay_period,
                rate_value=rate_value,
            )
        )

    return rows


def _coverage_shinsin(product_name: str) -> str:  # noqa: ARG001
    return "종신/CI"


def _coverage_annuity(product_name: str) -> str:
    if "변액" in product_name:
        return "변액연금"
    if "저축" in product_name:
        return "연금저축"
    return "연금"


def _coverage_term(product_name: str) -> str:
    if "경영" in product_name:
        return "CEO정기"
    return "종신/CI"


def _coverage_health(product_name: str) -> str:  # noqa: ARG001
    return "기타(보장성)"


# =============================================================================
# 진입점
# =============================================================================

def build_life_kyobo_conversion_rows(
    example: RateExample,
    wb: Workbook,
) -> list[RateExampleConversionRow]:
    """
    교보 생명보험 환산율/수정률 정규화 row 객체 목록을 생성한다.

    저장/replace/append 처리는 rate_example_normalizer.normalize_rate_example()이 담당한다.
    """
    if TARGET_SHEET_NAME not in wb.sheetnames:
        logger.warning("교보 정규화: 시트 '%s' 없음", TARGET_SHEET_NAME)
        return []

    ws = wb[TARGET_SHEET_NAME]
    rows: list[RateExampleConversionRow] = []

    # ── 종신보험 (B~F열) ────────────────────────────────────────────────
    # product=B(2), plan_type=E(5), pay_period=D(4), rate=F(6)
    rows.extend(_parse_table(
        example, ws,
        product_col=2,
        plan_type_col=5,
        pay_period_col=4,
        rate_col=6,
        coverage_type_fn=_coverage_shinsin,
    ))

    # ── CI보험 (H~L열) ──────────────────────────────────────────────────
    # product=H(8), plan_type=K(11), pay_period=J(10), rate=L(12)
    rows.extend(_parse_table(
        example, ws,
        product_col=8,
        plan_type_col=11,
        pay_period_col=10,
        rate_col=12,
        coverage_type_fn=_coverage_shinsin,
    ))

    # ── 연금보험 (N~R열) ────────────────────────────────────────────────
    # product=N(14), plan_type=Q(17), pay_period=P(16), rate=R(18)
    rows.extend(_parse_table(
        example, ws,
        product_col=14,
        plan_type_col=17,
        pay_period_col=16,
        rate_col=18,
        coverage_type_fn=_coverage_annuity,
    ))

    # ── 정기보험 (Z~AC열) ───────────────────────────────────────────────
    # product=Z(26), plan_type=없음, pay_period=AB(28), rate=AC(29)
    rows.extend(_parse_table(
        example, ws,
        product_col=26,
        plan_type_col=None,
        pay_period_col=28,
        rate_col=29,
        coverage_type_fn=_coverage_term,
    ))

    # ── 건강/어린이/기타보장 (AE~AH열) ─────────────────────────────────
    # product=AE(31), plan_type=없음, pay_period=AG(33), rate=AH(34)
    rows.extend(_parse_table(
        example, ws,
        product_col=31,
        plan_type_col=None,
        pay_period_col=33,
        rate_col=34,
        coverage_type_fn=_coverage_health,
    ))

    return rows