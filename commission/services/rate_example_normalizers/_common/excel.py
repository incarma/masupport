# django_ma/commission/services/rate_example_normalizers/_common/excel.py
from __future__ import annotations

"""
RateExample Excel parser 공통 helper.

주의:
- worksheet 자체를 unmerge 하거나 값을 직접 쓰지 않는다.
- 병합 셀 값을 parser 내부 lookup으로만 전파한다.
"""

from typing import Any

from openpyxl.worksheet.worksheet import Worksheet


def build_merged_value_map(ws: Worksheet) -> dict[tuple[int, int], Any]:
    """
    병합 셀의 좌상단 값을 병합 범위 전체 좌표에 매핑한다.
    """
    merged_map: dict[tuple[int, int], Any] = {}

    for merged_range in ws.merged_cells.ranges:
        top_left = ws.cell(
            row=merged_range.min_row,
            column=merged_range.min_col,
        ).value

        for row_no in range(merged_range.min_row, merged_range.max_row + 1):
            for col_no in range(merged_range.min_col, merged_range.max_col + 1):
                merged_map[(row_no, col_no)] = top_left

    return merged_map


def cell_value_with_merged(
    ws: Worksheet,
    merged_map: dict[tuple[int, int], Any],
    row_no: int,
    col_no: int,
) -> Any:
    """
    일반 셀/병합 셀 값을 동일하게 읽는다.

    우선순위:
    1. 실제 셀 값
    2. 병합 map 전파값
    """
    value = ws.cell(row=row_no, column=col_no).value
    if value not in (None, ""):
        return value
    return merged_map.get((row_no, col_no))


def build_worksheet_value_map(
    ws: Worksheet,
    *,
    include_empty: bool = True,
) -> dict[tuple[int, int], Any]:
    """
    worksheet 전체 값을 dict matrix로 만들고 병합 셀 좌상단 값을 범위 전체에 전파한다.

    주의:
    - worksheet 자체를 unmerge하거나 값을 쓰지 않는다.
    - parser 내부 lookup 전용이다.
    - include_empty=False이면 기존 fire_lotte.py처럼 값이 있는 셀만 먼저 담는다.
    """
    values: dict[tuple[int, int], Any] = {}

    for row in ws.iter_rows():
        for cell in row:
            if include_empty or cell.value is not None:
                values[(cell.row, cell.column)] = cell.value

    for merged_range in ws.merged_cells.ranges:
        top_left = ws.cell(
            row=merged_range.min_row,
            column=merged_range.min_col,
        ).value

        if not include_empty and top_left is None:
            continue

        for row_no in range(merged_range.min_row, merged_range.max_row + 1):
            for col_no in range(merged_range.min_col, merged_range.max_col + 1):
                values[(row_no, col_no)] = top_left

    return values


def filled_value_above(
    values: dict[tuple[int, int], Any],
    *,
    header_row: int,
    row_no: int,
    col_no: int,
    is_filled,
) -> Any:
    """
    현재 셀이 비어 있으면 header_row와 현재 행 사이 같은 컬럼의 최근 상단 값을 반환한다.

    용도:
    - 농협손보처럼 표 내부 공란 셀을 상단 텍스트로 보정하는 parser.

    is_filled:
    - parser별 텍스트 판정 정책을 유지하기 위한 callable.
    """
    current = values.get((row_no, col_no))
    if is_filled(current):
        return current

    for r in range(row_no - 1, header_row, -1):
        candidate = values.get((r, col_no))
        if is_filled(candidate):
            return candidate

    return current