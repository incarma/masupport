# django_ma/commission/services/rate_example_normalizers/fire_samsung.py
from __future__ import annotations

"""
삼성화재 손해보험 수정률 정규화 parser.

대상:
- insurer_type = fire
- category = conv
- insurer = 삼성

정규화 출력:
- 보험사     → insurer = "삼성"
- 상품군     → coverage_type
- 상품명     → product_name
- 구분       → plan_type
- 납기       → pay_period
- 수정률     → year1

중요:
- RateExampleConversionRow.year1은 DecimalField이므로 "%" 문자열을 저장하지 않는다.
- raw 수치 자체를 백분율 수치로 저장한다.
  예: raw 160 → Decimal("160") 저장 → 화면/API에서 "160%"로 표시
- raw가 Excel 내부 percent 값(예: 1.6, number_format="0%")이면 표시값 기준 160으로 보정한다.
"""

import re
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from commission.models import RateExample, RateExampleConversionRow
from commission.services.rate_example_normalizers._common.excel import (
    build_worksheet_value_map,
)
from commission.services.rate_example_normalizers._common.text import clean_spaces


INSURER = "삼성"

TITLE_SEARCH_MAX_ROWS = 12

HEADER_KEY_PLAN = "구분"
HEADER_KEY_PAY_PERIOD = "납기"
HEADER_KEY_RATE = "수정률"

PAY_PERIOD_HEADER_KEYWORDS = ("년", "만기", "최초", "갱신")
SKIP_TEXT_KEYWORDS = ("담보군", "비고", "주)", "※", "*")

TITLE_EXCLUDE_KEYWORDS = (
    "참고용",
    "한눈에 보기",
    "요약",
    "제작일자",
    "대외비",
    "현장관리자",
    "수수료지급",
    "자료로",
)


def build_fire_samsung_conversion_rows(
    example: RateExample,
    wb,
) -> list[RateExampleConversionRow]:
    """
    삼성화재 손보 수정률 raw workbook을 RateExampleConversionRow 목록으로 변환한다.

    처리 방식:
    1. 모든 시트 순회
    2. 병합셀 값을 parser 내부 matrix에서만 전파
    3. 테이블 제목행과 헤더행을 탐지
    4. 테이블별로 행형/매트릭스형 수정률을 공통 스키마로 펼침
    """
    rows: list[RateExampleConversionRow] = []

    for ws in wb.worksheets:
        values = _build_value_matrix(ws)

        # 삼성화재 자녀 전용 테이블 선처리
        # - 제목: "자녀 : NEW 마이 슈퍼스타 1-2종"
        # - 헤더: 갱신구분 / 납기 / 보장... 컬럼 구조
        # - 기존 일반 테이블 탐지 규칙과 충돌하지 않도록 별도 append 후 계속 일반 규칙도 수행
        rows.extend(_build_child_superstar_rows(example, ws, values))

        table_headers = _find_header_rows(ws, values)

        for idx, header in enumerate(table_headers):
            header_row, header_cols = header
            next_header_row = (
                table_headers[idx + 1][0]
                if idx + 1 < len(table_headers)
                else (ws.max_row or header_row) + 1
            )

            title_start_col, title_end_col = _header_title_search_bounds(header_cols)
            title = _find_table_title(
                values,
                header_row,
                start_col=title_start_col,
                end_col=title_end_col,
            )
            if not title:
                continue

            product_name = _normalize_product_name(title)
            if not product_name:
                continue

            plan_col = header_cols.get("plan")
            pay_period_col = header_cols.get("pay_period")
            rate_col = header_cols.get("rate")
            pay_period_rate_cols = header_cols.get("pay_period_rate_cols", [])

            for row_no in range(header_row + 1, next_header_row):
                row_values = [
                    _to_text(values.get((row_no, col_no)))
                    for col_no in range(1, (ws.max_column or 1) + 1)
                ]
                if _is_skip_row(row_values):
                    continue

                plan_type = _to_text(values.get((row_no, plan_col))) if plan_col else ""
                plan_type = _clean_text(plan_type)

                # A. 납기 컬럼 + 수정률 컬럼이 모두 있는 행형 테이블
                if pay_period_col and rate_col:
                    pay_period = _normalize_pay_period(values.get((row_no, pay_period_col)))
                    mod_rate = _to_decimal_percent(values.get((row_no, rate_col)))
                    if not pay_period or mod_rate is None:
                        continue

                    rows.append(_make_row(
                        example=example,
                        ws=ws,
                        row_no=row_no,
                        title=title,
                        product_name=product_name,
                        plan_type=plan_type,
                        pay_period=pay_period,
                        mod_rate=mod_rate,
                    ))
                    continue

                # B. 납기 컬럼 없이 헤더가 납기 역할을 하는 매트릭스형 테이블
                for col_no, header_text in pay_period_rate_cols:
                    pay_period = _normalize_pay_period(header_text)
                    mod_rate = _to_decimal_percent(values.get((row_no, col_no)))
                    if not pay_period or mod_rate is None:
                        continue

                    rows.append(_make_row(
                        example=example,
                        ws=ws,
                        row_no=row_no,
                        title=title,
                        product_name=product_name,
                        plan_type=plan_type,
                        pay_period=pay_period,
                        mod_rate=mod_rate,
                    ))

    # 최종 업로드 전 중복 방어:
    # 상품군 + 상품명 + 구분 + 납기가 모두 같으면 동일 상품으로 간주한다.
    return _dedupe_conversion_rows(rows)


def _dedupe_conversion_rows(
    rows: list[RateExampleConversionRow],
) -> list[RateExampleConversionRow]:
    """
    삼성화재 정규화 결과 중복 제거.

    기준:
    - coverage_type 상품군
    - product_name 상품명
    - plan_type 구분
    - pay_period 납기

    같은 키가 여러 번 생성되면 최초 row만 유지한다.
    """
    seen: set[tuple[str, str, str, str]] = set()
    deduped: list[RateExampleConversionRow] = []

    for row in rows:
        key = (
            _compact(row.coverage_type),
            _compact(row.product_name),
            _compact(row.plan_type),
            _compact(row.pay_period),
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(row)

    return deduped


def _build_child_superstar_rows(
    example: RateExample,
    ws: Worksheet,
    values: dict[tuple[int, int], Any],
) -> list[RateExampleConversionRow]:
    """
    '자녀 : NEW 마이 슈퍼스타 1-2종' 전용 정규화.

    규칙:
    1. 상품명 = ':' 오른쪽 텍스트 + '(' + 갱신구분 데이터 + ')'
    2. 구분 = '보장' 포함 헤더 텍스트
    3. 납기 = 납기 컬럼 데이터
    4. 상품군 = 납기에 '태아' 포함 시 보장(태아), 아니면 보장
    5. 수정률 = '보장' 포함 헤더 컬럼의 데이터
    """
    rows: list[RateExampleConversionRow] = []

    for title_row in range(1, (ws.max_row or 1) + 1):
        title_cells = [
            (col_no, _clean_text(values.get((title_row, col_no))))
            for col_no in range(1, (ws.max_column or 1) + 1)
        ]

        for title_col, title in title_cells:
            if "자녀" not in _title_left_of_colon(title):
                continue

            base_product_name = _normalize_product_name(title)
            if not base_product_name:
                continue

            header_row = title_row + 1
            header_map = {
                col_no: _clean_text(values.get((header_row, col_no)))
                for col_no in range(1, (ws.max_column or 1) + 1)
            }

            renew_col = _find_col_by_exact(header_map, "갱신구분")
            pay_col = _find_col_by_exact(header_map, HEADER_KEY_PAY_PERIOD)
            coverage_cols = [
                (col_no, header_text)
                for col_no, header_text in header_map.items()
                if "보장" in header_text
            ]

            if not renew_col or not pay_col or not coverage_cols:
                continue

            # 다음 제목/헤더가 나오기 전까지 현재 자녀 테이블 데이터로 본다.
            for row_no in range(header_row + 1, (ws.max_row or header_row) + 1):
                renew_type = _clean_text(values.get((row_no, renew_col)))
                pay_period = _normalize_pay_period(values.get((row_no, pay_col)))

                if not renew_type and not pay_period:
                    break

                if not pay_period:
                    continue

                product_name = f"{base_product_name} ({renew_type})" if renew_type else base_product_name
                coverage_type = "보장(태아)" if "태아" in pay_period else "보장"

                for rate_col, plan_type in coverage_cols:
                    mod_rate = _to_decimal_percent(values.get((row_no, rate_col)))
                    if mod_rate is None:
                        continue

                    rows.append(RateExampleConversionRow(
                        source_file=example,
                        source_sheet=ws.title,
                        source_row_no=row_no,
                        insurer_type=RateExample.TYPE_FIRE,
                        category=RateExample.CAT_CONV,
                        insurer=INSURER,
                        coverage_type=coverage_type,
                        strategy_flag="",
                        product_name=product_name,
                        plan_type=plan_type,
                        pay_period=pay_period,
                        year1=mod_rate,
                        year2=None,
                        year3=None,
                        year4=None,
                    ))

    return rows


def _make_row(
    *,
    example: RateExample,
    ws: Worksheet,
    row_no: int,
    title: str,
    product_name: str,
    plan_type: str,
    pay_period: str,
    mod_rate: Decimal,
) -> RateExampleConversionRow:
    """RateExampleConversionRow 생성 공통 helper."""
    return RateExampleConversionRow(
        source_file=example,
        source_sheet=ws.title,
        source_row_no=row_no,
        insurer_type=RateExample.TYPE_FIRE,
        category=RateExample.CAT_CONV,
        insurer=INSURER,
        coverage_type=_resolve_coverage_type(
            title=title,
            product_name=product_name,
            plan_type=plan_type,
            pay_period=pay_period,
        ),
        strategy_flag="",
        product_name=product_name,
        plan_type=plan_type,
        pay_period=pay_period,
        year1=mod_rate,
        year2=None,
        year3=None,
        year4=None,
    )


def _build_value_matrix(ws: Worksheet) -> dict[tuple[int, int], Any]:
    """
    병합셀 값을 포함한 value matrix를 만든다.

    실제 workbook은 변경하지 않고, parser 내부에서만 병합 범위 전체에
    좌상단 값을 전파한다.
    """
    return build_worksheet_value_map(ws)


def _find_header_rows(
    ws: Worksheet,
    values: dict[tuple[int, int], Any],
) -> list[tuple[int, dict[str, Any]]]:
    """
    테이블 헤더행을 찾는다.

    지원 형태:
    - 구분 / 납기 / 수정률
    - 구분 / 5년 / 10년 / 최초 / 갱신 등 납기형 헤더
    """
    headers: list[tuple[int, dict[str, Any]]] = []

    for row_no in range(1, (ws.max_row or 1) + 1):
        row_map: dict[int, str] = {
            col_no: _clean_text(values.get((row_no, col_no)))
            for col_no in range(1, (ws.max_column or 1) + 1)
        }

        non_empty = {col: text for col, text in row_map.items() if text}
        if len(non_empty) < 2:
            continue

        plan_col = _find_col_by_exact(non_empty, HEADER_KEY_PLAN)
        pay_period_col = _find_col_by_exact(non_empty, HEADER_KEY_PAY_PERIOD)
        rate_col = _find_col_by_exact(non_empty, HEADER_KEY_RATE)

        pay_period_rate_cols = [
            (col_no, text)
            for col_no, text in non_empty.items()
            if _looks_like_pay_period_header(text)
        ]

        # 행형: 구분 + 납기 + 수정률
        if plan_col and pay_period_col and rate_col:
            headers.append((
                row_no,
                {
                    "plan": plan_col,
                    "pay_period": pay_period_col,
                    "rate": rate_col,
                    "pay_period_rate_cols": [],
                },
            ))
            continue

        # 매트릭스형: 구분 + 납기형 헤더 1개 이상
        if plan_col and pay_period_rate_cols:
            headers.append((
                row_no,
                {
                    "plan": plan_col,
                    "pay_period": None,
                    "rate": None,
                    "pay_period_rate_cols": pay_period_rate_cols,
                },
            ))

    return headers


def _find_table_title(
    values: dict[tuple[int, int], Any],
    header_row: int,
    *,
    start_col: int,
    end_col: int,
) -> str:
    """
    헤더행 위쪽에서 가장 가까운 테이블 제목을 찾는다.

    제목은 보통 병합셀 또는 첫 번째 텍스트 셀에 있고,
    ':'가 있으면 상품명 파싱에도 사용된다.
    """
    start = max(1, header_row - TITLE_SEARCH_MAX_ROWS)

    for row_no in range(header_row - 1, start - 1, -1):
        texts = _unique_nonempty_texts(
            _clean_text(values.get((row_no, col_no)))
            for col_no in range(start_col, end_col + 1)
        )
        if not texts:
            continue

        joined = " ".join(texts)
        if _is_probable_title(joined):
            return joined

    return ""


def _header_title_search_bounds(header_cols: dict[str, Any]) -> tuple[int, int]:
    """
    현재 탐지된 테이블의 헤더 컬럼 범위만 제목 탐색 범위로 사용한다.

    이유:
    - 삼성 raw는 한 행에 여러 테이블이 가로로 배치되어 있다.
    - 전체 열을 제목 탐색 대상으로 삼으면 다른 테이블 제목/참고 문구가 섞인다.
    """
    cols: list[int] = []

    for key in ("plan", "pay_period", "rate"):
        col_no = header_cols.get(key)
        if isinstance(col_no, int):
            cols.append(col_no)

    for item in header_cols.get("pay_period_rate_cols", []) or []:
        if isinstance(item, tuple) and item and isinstance(item[0], int):
            cols.append(item[0])

    if not cols:
        return 1, 80

    return max(1, min(cols)), max(cols)


def _unique_nonempty_texts(values) -> list[str]:
    """
    병합셀 전파로 같은 제목이 여러 컬럼에 반복되는 현상을 제거한다.
    """
    result: list[str] = []
    seen: set[str] = set()

    for value in values:
        text = _clean_text(value)
        if not text:
            continue

        key = _compact(text)
        if key in seen:
            continue

        seen.add(key)
        result.append(text)

    return result


def _is_probable_title(text: str) -> bool:
    """테이블 제목 후보 여부를 판단한다."""
    if not text:
        return False
    if any(k in text for k in SKIP_TEXT_KEYWORDS):
        return False
    if any(k in text for k in TITLE_EXCLUDE_KEYWORDS):
        return False
    if HEADER_KEY_PLAN in text and (HEADER_KEY_RATE in text or HEADER_KEY_PAY_PERIOD in text):
        return False
    return True


def _normalize_product_name(title: str) -> str:
    """
    상품명 정규화.

    ':'가 있으면 오른쪽 텍스트만 상품명으로 사용하고,
    없으면 제목 전체를 상품명으로 사용한다.
    """
    text = _clean_text(title)
    if ":" in text:
        text = text.split(":", 1)[1]
    elif "：" in text:
        text = text.split("：", 1)[1]
    return _clean_text(text)


def _resolve_coverage_type(
    *,
    title: str,
    product_name: str,
    plan_type: str,
    pay_period: str,
) -> str:
    """
    상품군 정규화.

    - 테이블 제목의 ':' 왼쪽에 자녀 포함 → 보장(태아)
    - 제목에 실손 포함 + 납기 갱신 → 단독실손(갱신)
    - 그 외 → 보장
    """
    title_text = _clean_text(title)
    product_text = _clean_text(product_name)
    plan_text = _clean_text(plan_type)
    pay_text = _clean_text(pay_period)
    joined = f"{title_text} {product_text} {plan_text}"

    # 상품명 정규화(_normalize_product_name)로 ':' 오른쪽만 남기기 전에,
    # 원본 테이블 제목의 ':' 왼쪽 영역에서 자녀 여부를 먼저 판정한다.
    # 예: "자녀보험 : 삼성화재 XXX" → 상품명은 "삼성화재 XXX",
    #     상품군은 "보장(태아)" 유지
    title_left = _title_left_of_colon(title_text)
    if "자녀" in title_left:
        return "보장(태아)"

    if "실손" in joined:
        if "최초" in pay_text:
            return "단독실손(초회)"
        if "갱신" in pay_text:
            return "단독실손(갱신)"

    # 삼성화재 raw는 태아/자녀 정보가 테이블 제목뿐 아니라
    # 요약표의 세분/상품명/구분 영역에만 존재하는 경우가 있어
    # title + product_name + plan_type 전체를 기준으로 판정한다.
    if "자녀" in joined or "슈퍼스타" in joined:
        return "보장(태아)"

    return "보장"


def _title_left_of_colon(value: Any) -> str:
    """
    테이블 제목에서 ':' 또는 '：' 왼쪽 텍스트만 반환한다.

    상품명 정규화 단계에서는 ':' 오른쪽을 사용하지만,
    상품군 판정에서는 ':' 왼쪽의 분류성 문구를 먼저 확인해야 한다.
    """
    text = _clean_text(value)
    if ":" in text:
        return _clean_text(text.split(":", 1)[0])
    if "：" in text:
        return _clean_text(text.split("：", 1)[0])
    return text


def _normalize_pay_period(value: Any) -> str:
    """
    납기 정규화.

    - 숫자만 있으면 끝에 '년' 추가
    - Excel 숫자 10.0은 '10년' 처리
    - 텍스트는 줄바꿈 제거 후 공백 정리
    """
    text = _clean_text(value)
    if not text:
        return ""

    if re.fullmatch(r"\d+(\.0+)?", text):
        return f"{str(int(Decimal(text)))}년"

    return text


def _to_decimal_percent(value: Any) -> Decimal | None:
    """
    수정률을 Decimal로 변환한다.

    삼성화재 수정률은 raw 데이터의 숫자값을 그대로 저장한다.
    - raw 145   -> Decimal("145")
    - raw 240   -> Decimal("240")
    - raw 2000  -> Decimal("2000")
    - raw "145%" -> Decimal("145")

    주의:
    - number_format 기반 *100 보정 금지
    - 과대값 방어용 /100, /10 보정 금지
    - 화면 표시용 % 단위 처리는 프론트/API 표시 계층에서만 담당
    """
    if value is None:
        return None

    if isinstance(value, str):
        text = _clean_text(value)
        if not text or text in {"-", "–", "—"}:
            return None
        text = text.replace("%", "").replace(",", "")
    else:
        text = str(value).replace(",", "")

    try:
        dec = Decimal(text)
    except (InvalidOperation, ValueError):
        return None

    if dec == 0:
        return None
    
    # 삼성화재 raw 수정률은 실제 표시값의 100배 형태로 들어온다.
    #
    # 예:
    # raw 2000  -> 실제 20%
    # raw 14500 -> 실제 145%
    # raw 24000 -> 실제 240%
    #
    # 따라서 저장 직전 100으로 나눈 값을 저장한다.
    dec = dec / Decimal("100")

    return dec.quantize(Decimal("0.0001"))


def _find_col_by_exact(non_empty: dict[int, str], label: str) -> int | None:
    """공백 제거 후 정확히 일치하는 헤더 컬럼을 찾는다."""
    normalized_label = _compact(label)
    for col_no, text in non_empty.items():
        if _compact(text) == normalized_label:
            return col_no
    return None


def _looks_like_pay_period_header(text: str) -> bool:
    """납기 역할을 할 수 있는 헤더인지 판단한다."""
    value = _clean_text(text)
    if not value:
        return False
    if _compact(value) in {
        _compact(HEADER_KEY_PLAN),
        _compact(HEADER_KEY_PAY_PERIOD),
        _compact(HEADER_KEY_RATE),
    }:
        return False
    return any(keyword in value for keyword in PAY_PERIOD_HEADER_KEYWORDS)


def _is_skip_row(row_values: list[str]) -> bool:
    """주석/빈 행/합계성 행 제외."""
    texts = [v for v in row_values if v]
    if not texts:
        return True

    joined = " ".join(texts)
    if any(keyword in joined for keyword in SKIP_TEXT_KEYWORDS):
        return True

    # 데이터 없이 라벨만 있는 행 방어
    if len(texts) == 1 and _is_probable_title(texts[0]):
        return True

    return False


def _to_text(value: Any) -> str:
    """값을 문자열로 안전 변환한다."""
    if value is None:
        return ""
    return str(value)


def _clean_text(value: Any) -> str:
    """줄바꿈/중복 공백 정리."""
    return clean_spaces(_to_text(value).replace("\r\n", "\n").replace("\r", "\n"))


def _compact(value: Any) -> str:
    """비교용 문자열: 모든 공백 제거."""
    return re.sub(r"\s+", "", _clean_text(value))