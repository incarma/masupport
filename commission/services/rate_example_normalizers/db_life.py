from __future__ import annotations

"""
DB생명 환산률/수정률 정규화 모듈.

적용 대상:
- RateExample.insurer_type == life
- RateExample.category == conv
- RateExample.insurer == "DB"
- xlsx 원본 파일

정규화 규칙:
1. 시트명에 "특약", "방카교차"가 포함된 시트는 제외한다.
2. 각 시트의 첫 번째 테이블만 정규화한다.
   - 이후 등장하는 특약/의무부가 테이블은 제외한다.
3. 상품명은 각 시트 A1 셀 텍스트를 사용하되 "□" 문자는 제거한다.
4. 보종은 시트명 기준으로 판정한다.
5. 첫 번째 테이블 컬럼 매핑:
   - A열: 구분
   - B열: 납기
   - C열: 1차년
   - D열: 2차년
   - E열: 3차년
   - F열: 4차년
   - 단, F열 헤더가 "계"이면 4차년은 제외한다.
"""

import logging
import re
from decimal import Decimal, InvalidOperation
from typing import Iterable

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from commission.models import RateExample, RateExampleConversionRow

logger = logging.getLogger(__name__)

EXCLUDED_SHEET_KEYWORDS = ("특약", "방카교차")
STOP_TABLE_KEYWORDS = ("특약", "의무부가")


def _clean_text(value) -> str:
    """엑셀 셀 값을 문자열로 안전 정규화한다."""
    if value is None:
        return ""
    return re.sub(
        r"\s+",
        " ",
        str(value).replace("\r", "\n").replace("\n", " "),
    ).strip()


def _clean_product_name(value) -> str:
    """DB생명 상품명 원천인 A1 텍스트에서 특수문자와 불필요 공백을 제거한다."""
    return _clean_text(value).replace("□", "").strip()


def _to_decimal(value):
    """
    정수/실수/문자/퍼센트 값을 DecimalField 저장값으로 변환한다.

    변환 불가 값은 None으로 처리해 row 전체 실패를 방지한다.
    """
    if value is None or value == "":
        return None

    if isinstance(value, Decimal):
        return value.quantize(Decimal("0.0001"))

    if isinstance(value, (int, float)):
        return Decimal(str(value)).quantize(Decimal("0.0001"))

    text = _clean_text(value).replace(",", "").replace("%", "")
    if not text:
        return None

    try:
        return Decimal(text).quantize(Decimal("0.0001"))
    except InvalidOperation:
        logger.warning("[rate_example][DB] decimal parse skipped value=%r", value)
        return None


def _has_any_rate(*values) -> bool:
    """연차별 환산률 값이 하나라도 있으면 유효 데이터 행으로 본다."""
    return any(_to_decimal(v) is not None for v in values)


def _is_excluded_sheet(sheet_name: str) -> bool:
    """정규화 제외 시트 여부를 판정한다."""
    return any(keyword in sheet_name for keyword in EXCLUDED_SHEET_KEYWORDS)


def _coverage_type_from_product_name(product_name: str) -> str:
    """
    상품명(A1) 기반 보종 판정.

    우선순위:
    - 종신 → 종신/CI
    - 연금 → 연금
    - 경영 → CEO정기
    - 그 외 → 기타(보장성)
    """
    if "종신" in product_name:
        return "종신/CI"
    if "연금" in product_name:
        return "연금"
    if "경영" in product_name:
        return "CEO정기"
    return "기타(보장성)"


def _row_text(ws: Worksheet, row_no: int, start_col: int = 1, end_col: int = 8) -> str:
    """특약/의무부가 등 두 번째 테이블 시작 감지를 위해 행 텍스트를 합친다."""
    values = [
        _clean_text(ws.cell(row_no, col_no).value)
        for col_no in range(start_col, end_col + 1)
    ]
    return " ".join(v for v in values if v)


def _looks_like_header(ws: Worksheet, row_no: int) -> bool:
    """
    첫 번째 테이블 헤더 후보를 찾는다.

    DB생명 raw 파일은 헤더가 2줄 구조다.
    예:
    - 1행차: A=구분, B=납기, C=성적률
    - 2행차: C=1차년, D=2차년, E=3차년, F=계 또는 4차년

    원본 파일의 미세한 공백/개행 변형을 흡수하기 위해 포함 검색으로 판정한다.
    """
    a = _clean_text(ws.cell(row_no, 1).value)
    b = _clean_text(ws.cell(row_no, 2).value)
    c = _clean_text(ws.cell(row_no, 3).value)

    c_next = _clean_text(ws.cell(row_no + 1, 3).value)
    d_next = _clean_text(ws.cell(row_no + 1, 4).value)
    e_next = _clean_text(ws.cell(row_no + 1, 5).value)

    return (
        "구분" in a
        and "납기" in b
        and ("성적" in c or "환산" in c or "수정" in c)
        and "1" in c_next and "차" in c_next
        and "2" in d_next and "차" in d_next
        and "3" in e_next and "차" in e_next
    )


def _find_first_table_header_row(ws: Worksheet) -> int | None:
    """
    시트에서 첫 번째 정규화 대상 테이블의 헤더 행 번호를 찾는다.

    특약/의무부가 테이블 앞의 첫 번째 구분/납기/연차 테이블만 대상으로 한다.
    """
    for row_no in range(1, ws.max_row + 1):
        row_text = _row_text(ws, row_no)
        if any(keyword in row_text for keyword in STOP_TABLE_KEYWORDS):
            return None

        if _looks_like_header(ws, row_no):
            return row_no

    return None


def _has_year4_column(ws: Worksheet, header_row_no: int) -> bool:
    """
    F열 4차년 컬럼 사용 여부를 판정한다.

    - F열 헤더가 "계"이면 4차년 컬럼으로 보지 않는다.
    - F열 헤더가 비어 있으면 4차년 컬럼으로 보지 않는다.
    - F열 헤더에 "4차" 또는 "4차년" 의미가 있으면 사용한다.
    """
    # DB생명 raw는 연차 헤더가 header_row_no + 1 행에 존재한다.
    f_header = _clean_text(ws.cell(header_row_no + 1, 6).value)

    if not f_header:
        return False
    if f_header == "계" or "계" == f_header.replace(" ", ""):
        return False

    return "4" in f_header and "차" in f_header


def _normalize_db_sheet(
    example: RateExample,
    ws: Worksheet,
) -> Iterable[RateExampleConversionRow]:
    """
    DB생명 단일 시트의 첫 번째 테이블을 정규화한다.
    """
    header_row_no = _find_first_table_header_row(ws)
    if not header_row_no:
        logger.info("[rate_example][DB] table header not found sheet=%s", ws.title)
        return []

    rows: list[RateExampleConversionRow] = []

    product_name = _clean_product_name(ws.cell(1, 1).value)
    coverage_type = _coverage_type_from_product_name(product_name)
    has_year4 = _has_year4_column(ws, header_row_no)

    # 2줄 헤더 다음 행부터 첫 번째 테이블 데이터로 판단한다.
    # 데이터 시작 후 빈 행 또는 특약/의무부가 문구가 나오면 첫 번째 테이블 종료로 본다.
    started = False
    last_plan_type = ""

    for row_no in range(header_row_no + 2, ws.max_row + 1):
        row_text = _row_text(ws, row_no)
        if any(keyword in row_text for keyword in STOP_TABLE_KEYWORDS):
            break

        raw_plan_type = _clean_text(ws.cell(row_no, 1).value)
        plan_type = raw_plan_type or last_plan_type
        pay_period = _clean_text(ws.cell(row_no, 2).value)

        y1_raw = ws.cell(row_no, 3).value
        y2_raw = ws.cell(row_no, 4).value
        y3_raw = ws.cell(row_no, 5).value
        y4_raw = ws.cell(row_no, 6).value if has_year4 else None

        has_rate = _has_any_rate(y1_raw, y2_raw, y3_raw, y4_raw)

        if not plan_type and not pay_period and not has_rate:
            if started:
                break
            continue

        if not has_rate:
            continue

        started = True
        if raw_plan_type:
            last_plan_type = raw_plan_type

        rows.append(
            RateExampleConversionRow(
                source_file=example,
                source_sheet=ws.title,
                source_row_no=row_no,
                insurer_type=example.insurer_type,
                category=example.category,
                insurer="DB",
                coverage_type=coverage_type,
                strategy_flag="",
                product_name=product_name,
                plan_type=plan_type,
                pay_period=pay_period,
                year1=_to_decimal(y1_raw),
                year2=_to_decimal(y2_raw),
                year3=_to_decimal(y3_raw),
                year4=_to_decimal(y4_raw) if has_year4 else None,
            )
        )

    return rows


def build_db_life_conversion_rows(
    example: RateExample,
    wb: Workbook,
) -> list[RateExampleConversionRow]:
    """
    DB생명 xlsx 전체 workbook을 정규화 row 목록으로 변환한다.

    DB 저장/삭제는 상위 normalizer가 일괄 처리한다.
    """
    normalized_rows: list[RateExampleConversionRow] = []

    for sheet_name in wb.sheetnames:
        if _is_excluded_sheet(sheet_name):
            logger.info("[rate_example][DB] excluded sheet skipped sheet=%s", sheet_name)
            continue

        ws = wb[sheet_name]
        normalized_rows.extend(_normalize_db_sheet(example, ws))

    return normalized_rows