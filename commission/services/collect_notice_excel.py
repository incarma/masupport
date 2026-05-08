# django_ma/commission/services/collect_notice_excel.py
from __future__ import annotations

"""
Collect Notice Excel Service

역할:
- 환수내역 안내자료 원본 엑셀 파일들을 서버에서 파싱
- 기존 collect_notice.js의 전처리/마스킹 규칙을 서버로 이전
- openpyxl로 결과 xlsx 생성 및 서식 확정 적용

주의:
- 이 모듈은 HTTP 응답을 만들지 않는다.
- View는 이 서비스가 반환한 bytes/filename/row_count만 사용한다.
"""

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from contextlib import suppress
import shutil
import subprocess
import tempfile
import logging
from pathlib import Path
import re
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.page import PageMargins
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


logger = logging.getLogger(__name__)
PDF_MAGIC = b"%PDF"


# =============================================================================
# 1) 결과 엑셀 구조 SSOT
# =============================================================================

HEADERS: list[str] = [
    "월도",
    "항목구분",
    "지급/환수",
    "상품명",
    "증권번호",
    "계약자",
    "수납구분",
    "영수일",
    "회차",
    "영수보험료",
    "지급율(환수율)",
    "지급금액(환수금액)",
    "보험계약일",
    "모집자",
    "지급자",
]

SHEET_NAME = "환수내역"
FONT_NAME = "맑은 고딕"
HEADER_FILL = PatternFill(fill_type="solid", fgColor="F2F2F2")
SUMMARY_FILL = PatternFill(fill_type="solid", fgColor="E2EFDA")  # 합계 행 연한 초록
RIGHT_ALIGN_DATA_COLS = {10, 11, 12}  # J/K/L: 영수보험료, 지급율, 지급금액

# 사용자 요구사항 기준 열 너비
COLUMN_WIDTHS: dict[str, float] = {
    "A": 10,   # 월도
    "B": 10,   # 항목구분
    "C": 10,   # 지급/환수
    "D": 40,   # 상품명
    "E": 15,   # 증권번호
    "F": 10,   # 계약자
    "G": 10,   # 수납구분
    "H": 10,   # 영수일
    "I": 5,    # 회차
    "J": 10,   # 영수보험료
    "K": 15,   # 지급율(환수율)
    "L": 15,   # 지급금액(환수금액)
    "M": 10,   # 보험계약일
    "N": 10,   # 모집자
    "O": 10,   # 지급자
}


@dataclass(frozen=True)
class NoticeSourceFile:
    """
    업로드 원본 파일 1개와 해당 기준 월도.

    ym:
    - 화면의 행별 기준 연월
    - 형식: YYYY-MM
    """

    ym: str
    file: Any


@dataclass(frozen=True)
class NoticeWorkbookResult:
    """
    View에서 HttpResponse로 변환할 결과.
    """

    content: bytes
    filename: str
    row_count: int


@dataclass(frozen=True)
class NoticePdfResult:
    content: bytes
    filename: str
    row_count: int


# =============================================================================
# 2) 공개 API
# =============================================================================

def build_collect_notice_excel(
    *,
    target_name: str,
    target_branch: str,
    title_year: str,
    title_month: str,
    sources: list[NoticeSourceFile],
    manual_rows: list[dict[str, Any]] | None = None,
) -> NoticeWorkbookResult:
    """
    환수내역 안내자료 xlsx 생성.

    처리 흐름:
    1. 업로드 파일별 첫 번째 시트 파싱
    2. 기존 JS와 동일한 전처리/마스킹 수행
    3. 월도 기준 정렬
    4. openpyxl Workbook 생성
    5. 사용자 요구 서식 적용
    6. bytes 반환
    """
    _validate_meta(
        target_name=target_name,
        title_year=title_year,
        title_month=title_month,
        sources=sources,
        manual_rows=manual_rows or [],
    )

    rows: list[dict[str, str]] = []
    for src in sources:
        rows.extend(_clean_rows(_iter_first_sheet_rows(src.file), src.ym))
    rows.extend(_normalize_manual_rows(manual_rows or []))

    rows.sort(key=lambda r: r.get("_ym", ""))

    if not rows:
        raise ValueError("처리할 데이터가 없습니다. 지급금액이 모두 0이거나 유효 데이터가 없습니다.")

    wb = _build_workbook(
        rows=rows,
        target_name=target_name,
        target_branch=target_branch,
        title_year=title_year,
        title_month=title_month,
    )

    out = BytesIO()
    wb.save(out)

    filename = _build_filename(
        target_name=target_name,
        title_year=title_year,
        title_month=title_month,
    )

    return NoticeWorkbookResult(
        content=out.getvalue(),
        filename=filename,
        row_count=len(rows),
    )


def build_collect_notice_pdf(
    *,
    target_name: str,
    target_branch: str,
    title_year: str,
    title_month: str,
    sources: list[NoticeSourceFile],
    manual_rows: list[dict[str, Any]] | None = None,
) -> NoticePdfResult:
    """
    환수내역 안내자료 PDF 생성.

    처리 방식:
    1. 기존 openpyxl 결과 xlsx 생성
    2. xlsx 파일을 임시 디렉터리에 저장
    3. LibreOffice headless로 PDF 변환
    4. PDF bytes 반환

    주의:
    - openpyxl은 PDF 렌더링 기능이 없으므로 LibreOffice가 필요하다.
    - 운영 서버에 libreoffice 또는 soffice 실행파일이 설치되어 있어야 한다.
    """
    xlsx_result = build_collect_notice_excel(
        target_name=target_name,
        target_branch=target_branch,
        title_year=title_year,
        title_month=title_month,
        sources=sources,
        manual_rows=manual_rows or [],
    )

    pdf_content = _convert_xlsx_bytes_to_pdf(
        xlsx_content=xlsx_result.content,
        xlsx_filename=xlsx_result.filename,
    )

    if not pdf_content.startswith(PDF_MAGIC):
        raise RuntimeError(
            "PDF 변환 결과가 올바르지 않습니다. 생성 파일이 PDF 형식이 아닙니다."
        )

    pdf_filename = re.sub(r"\.xlsx$", ".pdf", xlsx_result.filename, flags=re.IGNORECASE)

    return NoticePdfResult(
        content=pdf_content,
        filename=pdf_filename,
        row_count=xlsx_result.row_count,
    )


# =============================================================================
# 3) 입력 검증
# =============================================================================

def _validate_meta(
    *,
    target_name: str,
    title_year: str,
    title_month: str,
    sources: list[NoticeSourceFile],
    manual_rows: list[dict[str, Any]],
) -> None:
    if not str(target_name or "").strip():
        raise ValueError("대상자 이름이 없습니다. 대상자를 다시 선택해주세요.")

    if not str(title_year or "").isdigit():
        raise ValueError("제목 기준 연도가 올바르지 않습니다.")

    if not str(title_month or "").isdigit():
        raise ValueError("제목 기준 월이 올바르지 않습니다.")

    month = int(title_month)
    if month < 1 or month > 12:
        raise ValueError("제목 기준 월은 1~12 범위여야 합니다.")

    if not sources and not manual_rows:
        raise ValueError("내역 파일 또는 수기 입력 행을 1개 이상 추가해주세요.")

    for src in sources:
        if not re.fullmatch(r"\d{4}-\d{2}", src.ym or ""):
            raise ValueError(f"파일 기준 연월 형식이 올바르지 않습니다: {src.ym}")


# =============================================================================
# 4) 원본 엑셀 파싱
# =============================================================================

def _iter_first_sheet_rows(uploaded_file: Any) -> Iterable[tuple[Any, ...]]:
    """
    업로드된 엑셀 파일의 첫 번째 시트를 values_only 행 iterator로 반환.

    - read_only=True: 대용량 파일 메모리 사용 최소화
    - data_only=True: 수식 셀은 계산된 값 기준
    """
    name = getattr(uploaded_file, "name", "unknown")
    ext = Path(str(name)).suffix.lower()
    if ext not in {".xlsx", ".xlsm"}:
        raise ValueError(f"지원하지 않는 엑셀 형식입니다: {name}")

    wb = None
    try:
        uploaded_file.seek(0)
        wb = load_workbook(uploaded_file, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(values_only=True):
            yield tuple(row)
    except Exception as exc:
        raise ValueError(f"파일 파싱 실패: {name}") from exc
    finally:
        with suppress(Exception):
            wb.close()  # type: ignore[name-defined]


def _cell(row: tuple[Any, ...], idx: int) -> Any:
    return row[idx] if len(row) > idx else ""


def _clean_rows(raw_rows: Iterable[tuple[Any, ...]], ym: str) -> list[dict[str, str]]:
    """
    기존 collect_notice.js 전처리 규칙을 서버로 이전.

    규칙:
    1. 첫 행은 헤더로 간주해 제거
    2. A열이 "전체 N건" 형식이면 합계행으로 제거
    3. 지급금액(index 12)이 0 또는 빈 값이면 제거
    4. 개인정보성 필드 마스킹
    """
    result: list[dict[str, str]] = []

    for idx, row in enumerate(raw_rows):
        if idx == 0:
            continue

        first_col = _to_str(_cell(row, 0))
        if re.fullmatch(r"전체\s*\d+\s*건", first_col):
            continue

        pay_raw = _cell(row, 12)
        pay_num = _to_number(pay_raw)
        if pay_num is None or pay_num == 0:
            continue

        result.append(
            {
                "_ym": ym,
                "항목구분": _to_str(_cell(row, 1)),
                "지급환수": _to_str(_cell(row, 2)),  # raw C열: 지급/환수
                "상품명": _to_str(_cell(row, 4)),
                "증권번호": _mask_policy(_to_str(_cell(row, 5))),
                "계약자": _mask_name(_to_str(_cell(row, 6))),
                "수납구분": _to_str(_cell(row, 7)),
                "영수일": _to_str(_cell(row, 8)),
                "회차": _to_str(_cell(row, 9)),
                "영수보험료": _money(_cell(row, 10)),
                "지급율": _rate(_cell(row, 11)),
                "지급금액": _money(_cell(row, 12)),
                "보험계약일": _to_str(_cell(row, 13)),
                "모집자": _mask_name(_strip_paren_id(_to_str(_cell(row, 14)))),
                "지급자": _mask_name(_strip_paren_id(_to_str(_cell(row, 15)))),
            }
        )

    return result


def _normalize_manual_rows(manual_rows: list[dict[str, Any]]) -> list[dict[str, str]]:
    """
    수기 입력 행을 결과 Workbook row 구조로 정규화한다.

    규칙:
    - 필수: 월도, 지급/환수, 상품명, 지급금액
    - 지급/환수: 지급 또는 환수만 허용
    - 증권번호/계약자/모집자/지급자는 원본 파일 처리와 동일하게 마스킹
    """
    result: list[dict[str, str]] = []

    for idx, row in enumerate(manual_rows, start=1):
        if not isinstance(row, dict):
            raise ValueError(f"수기 입력 {idx}행 형식이 올바르지 않습니다.")

        ym = _to_str(row.get("ym"))
        pay_refund = _to_str(row.get("pay_refund"))
        product_name = _to_str(row.get("product_name"))
        amount_raw = row.get("amount")

        if not ym or not pay_refund or not product_name or not _to_str(amount_raw):
            raise ValueError(f"수기 입력 {idx}행의 필수값(월도/지급·환수/상품명/지급금액)을 확인해주세요.")

        if not re.fullmatch(r"\d{4}-\d{2}", ym):
            raise ValueError(f"수기 입력 {idx}행의 월도 형식이 올바르지 않습니다.")

        if pay_refund not in {"지급", "환수"}:
            raise ValueError(f"수기 입력 {idx}행의 지급/환수 값은 '지급' 또는 '환수'만 가능합니다.")

        amount_num = _to_number(amount_raw)
        if amount_num is None:
            raise ValueError(f"수기 입력 {idx}행의 지급금액은 숫자여야 합니다.")

        rate_raw = _to_str(row.get("rate"))
        if rate_raw:
            rate_num = _to_number(rate_raw)
            if rate_num is None or rate_num < 0 or rate_num > 100:
                raise ValueError(f"수기 입력 {idx}행의 지급율은 0~100 범위여야 합니다.")

        result.append(
            {
                "_ym": ym,
                "항목구분": _to_str(row.get("item_type")),
                "지급환수": pay_refund,
                "상품명": product_name,
                "증권번호": _mask_policy(_to_str(row.get("policy_no"))),
                "계약자": _mask_name(_to_str(row.get("contractor"))),
                "수납구분": _to_str(row.get("payment_type")),
                "영수일": _to_str(row.get("receipt_date")),
                "회차": _money(row.get("round_no")),
                "영수보험료": _money(row.get("premium")),
                "지급율": _rate(row.get("rate")),
                "지급금액": _money(amount_raw),
                "보험계약일": _to_str(row.get("contract_date")),
                "모집자": _mask_name(_strip_paren_id(_to_str(row.get("recruiter")))),
                "지급자": _mask_name(_strip_paren_id(_to_str(row.get("payer")))),
            }
        )

    return result


# =============================================================================
# 5) 결과 Workbook 생성
# =============================================================================

def _build_workbook(
    *,
    rows: list[dict[str, str]],
    target_name: str,
    target_branch: str,
    title_year: str,
    title_month: str,
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    title_text = (
        f"{target_branch or '-'} {target_name} "
        f"{int(title_year)}년 {int(title_month):02d}월 기준 환수내역"
    )

    # -------------------------------------------------------------------------
    # 1행: 제목
    # -------------------------------------------------------------------------
    ws["A1"] = title_text
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))

    # -------------------------------------------------------------------------
    # 2행: 빈 행 / 3행: 헤더 / 4행~: 데이터
    # -------------------------------------------------------------------------
    for col_idx, header in enumerate(HEADERS, start=1):
        ws.cell(row=3, column=col_idx, value=header)

    for row_idx, item in enumerate(rows, start=4):
        values = [
            item["_ym"],
            item["항목구분"],
            item["지급환수"],
            item["상품명"],
            item["증권번호"],
            item["계약자"],
            item["수납구분"],
            item["영수일"],
            item["회차"],
            item["영수보험료"],
            item["지급율"],
            item["지급금액"],
            item["보험계약일"],
            item["모집자"],
            item["지급자"],
        ]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 합계 행: 지급금액(환수금액) 열 합산
    total_amount = sum((_to_number(item["지급금액"]) or Decimal(0)) for item in rows)
    summary_row_idx = 4 + len(rows)
    ws.cell(row=summary_row_idx, column=1, value="합계")
    ws.cell(row=summary_row_idx, column=12, value=f"{int(total_amount):,}")

    _apply_styles(ws, max_row=summary_row_idx, max_col=len(HEADERS))
    _apply_summary_row_styles(ws, row_idx=summary_row_idx, max_col=len(HEADERS))
    _apply_print_settings(ws)
    return wb


def _apply_print_settings(ws) -> None:
    """
    PDF 변환용 인쇄 설정.

    요구사항:
    - A4 가로
    - 모든 열을 한 페이지 너비에 맞춤
    """
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.page_margins = PageMargins(
        left=0.25,
        right=0.25,
        top=0.35,
        bottom=0.35,
        header=0.1,
        footer=0.1,
    )


def _apply_styles(ws, *, max_row: int, max_col: int) -> None:
    """
    사용자 요구 서식 적용.

    요구사항:
    - A1 높이 25
    - A1 글자 크기 14, bold, 상하좌우 가운데
    - A1 제외 전체 글자 크기 10
    - A1 제외 전체 행 높이 13.5
    - 지정 열 너비
    - 내용이 포함된 셀 테두리
    """
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 열 너비
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # 제목
    ws.row_dimensions[1].height = 25
    title_cell = ws["A1"]
    title_cell.font = Font(name=FONT_NAME, size=14, bold=True)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = border

    # 병합 제목 영역 전체 테두리 보강
    for col_idx in range(1, max_col + 1):
        ws.cell(row=1, column=col_idx).border = border

    # A1 제외 행 높이 + 글꼴
    for row_idx in range(2, max_row + 1):
        ws.row_dimensions[row_idx].height = 13.5
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(name=FONT_NAME, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # -------------------------------------------------------------
            # 데이터 정렬
            # - 헤더 행(3행)은 기존처럼 가운데 정렬
            # - 데이터 행(4행~)의 I/J/K열은 우측 정렬
            # -------------------------------------------------------------
            if row_idx >= 4 and col_idx in RIGHT_ALIGN_DATA_COLS:
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # 내용이 포함된 셀 + 표 영역에는 테두리 적용
            # 3행 헤더와 4행 이후 데이터는 표 영역으로 간주
            if row_idx >= 3:
                cell.border = border

    # 헤더 행: bold + 연한 회색 배경
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = Font(name=FONT_NAME, size=10, bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border


def _apply_summary_row_styles(ws, *, row_idx: int, max_col: int) -> None:
    """합계 행: bold + 연한 초록 배경. 지급금액 열은 우측 정렬."""
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = Font(name=FONT_NAME, size=10, bold=True)
        cell.fill = SUMMARY_FILL
        cell.border = border
        if col_idx == 12:
            cell.alignment = Alignment(horizontal="right", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center")


# =============================================================================
# 6) 데이터 정규화 / 마스킹
# =============================================================================

def _to_str(value: Any) -> str:
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")

    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip()


def _to_number(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None

    if isinstance(value, Decimal):
        return value

    if isinstance(value, int):
        return Decimal(value)

    if isinstance(value, float):
        return Decimal(str(value)) if value == value else None

    s = str(value).strip().replace(",", "").replace("%", "")
    if not s or s.lower() in {"nan", "none", "-"}:
        return None

    try:
        return Decimal(s)
    except Exception:
        return None


def _money(value: Any) -> str:
    n = _to_number(value)
    if n is None:
        return _to_str(value)

    # 기존 JS와 동일하게 정수 콤마 표기
    return f"{int(n):,}"


def _rate(value: Any) -> str:
    s = _to_str(value)
    if not s:
        return ""

    if "%" in s:
        return s

    return f"{s}%"


def _strip_paren_id(value: str) -> str:
    return re.sub(r"\([^)]*\)", "", value).strip()


def _mask_name(name: str) -> str:
    if not name:
        return ""

    if len(name) == 1:
        return "*"

    if len(name) == 2:
        return name[0] + "*"

    return name[0] + ("*" * (len(name) - 2)) + name[-1]


def _mask_policy(policy: str) -> str:
    if not policy:
        return ""

    if len(policy) <= 4:
        return "*" * len(policy)

    return ("*" * (len(policy) - 4)) + policy[-4:]


def _safe_filename_part(value: str) -> str:
    s = re.sub(r'[\\/:*?"<>|]+', "_", str(value or "").strip())
    return s or "대상자"


def _build_filename(*, target_name: str, title_year: str, title_month: str) -> str:
    name = _safe_filename_part(target_name)
    return f"환수내역_{name}_{int(title_year)}년{int(title_month):02d}월.xlsx"


# =============================================================================
# 7) XLSX → PDF 변환
# =============================================================================

def _find_libreoffice_binary() -> str:
    """
    LibreOffice 실행파일 탐색.

    Linux/Docker:
    - libreoffice
    - soffice

    Windows 개발환경:
    - soffice.exe가 PATH에 있으면 사용 가능
    """
    # 1) PATH 우선 탐색
    for cmd in ("libreoffice", "soffice", "soffice.exe"):
        found = shutil.which(cmd)
        if found:
            return found
        
    # 2) Windows 개발환경 기본 설치 경로 탐색
    windows_candidates = [
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    ]
    for p in windows_candidates:
        if Path(p).exists():
            return p

    raise RuntimeError(
        "PDF 변환을 위해 LibreOffice가 필요합니다. "
        "서버 또는 개발 PC에 LibreOffice를 설치하거나 soffice 실행파일을 PATH에 등록해주세요."
    )


def _convert_xlsx_bytes_to_pdf(*, xlsx_content: bytes, xlsx_filename: str) -> bytes:
    """
    xlsx bytes를 LibreOffice headless로 PDF 변환한다.

    주의:
    - LibreOffice는 한글/비ASCII 파일명을 subprocess 인자로 받을 때
      Windows 환경에서 --outdir을 무시하거나 PDF를 생성하지 않는 경우가 있다.
    - 이를 방지하기 위해 ASCII-only 임시 파일명을 사용한다.
    """
    lo_bin = _find_libreoffice_binary()

    with tempfile.TemporaryDirectory(prefix="collect_notice_pdf_") as tmpdir:
        tmp_path = Path(tmpdir)
        # 한글 파일명이 LibreOffice subprocess에 전달될 때 발생하는 인코딩 문제를 방지한다.
        xlsx_path = tmp_path / "collect_notice_temp.xlsx"
        xlsx_path.write_bytes(xlsx_content)

        cmd = [
            lo_bin,
            "--headless",
            "--norestore",
            "--nofirststartwizard",
            "--convert-to",
            "pdf",
            "--outdir",
            str(tmp_path),
            str(xlsx_path),
        ]

        try:
            completed = subprocess.run(
                cmd,
                check=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                timeout=120,
            )
        except subprocess.TimeoutExpired as exc:
            raise RuntimeError("PDF 변환 시간이 초과되었습니다. (120초)") from exc
        except subprocess.CalledProcessError as exc:
            stderr = (exc.stderr or b"").decode("utf-8", errors="replace")
            stdout = (exc.stdout or b"").decode("utf-8", errors="replace")
            logger.warning(
                "[collect_notice_pdf] libreoffice failed stdout=%s stderr=%s",
                stdout[:1000],
                stderr[:1000],
            )
            raise RuntimeError(f"PDF 변환에 실패했습니다.\nstdout={stdout}\nstderr={stderr}") from exc

        # LibreOffice는 입력 파일과 같은 이름의 .pdf를 --outdir에 생성한다.
        pdf_path = xlsx_path.with_suffix(".pdf")
        if not pdf_path.exists():
            candidates = list(tmp_path.glob("*.pdf"))
            if candidates:
                pdf_path = candidates[0]

        if not pdf_path.exists():
            stdout = (completed.stdout or b"").decode("utf-8", errors="replace")
            raise RuntimeError(
                f"PDF 파일이 생성되지 않았습니다. "
                f"LibreOffice 설치 상태를 확인하세요.\nstdout={stdout}"
            )

        pdf_bytes = pdf_path.read_bytes()
        if not pdf_bytes.startswith(PDF_MAGIC):
            raise RuntimeError(
                "PDF 파일은 생성되었지만 파일 형식이 올바르지 않습니다. "
                "LibreOffice PDF 변환 결과를 확인해주세요."
            )

        return pdf_bytes