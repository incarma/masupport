# commission/services/rate_example_pay_normalizer.py
from __future__ import annotations

"""
지급률 예시표 정규화 서비스.

파일 위치:
    commission/services/rate_example_pay_normalizer.py

역할:
- 업로드된 지급률 xlsx를 20개 생보사 컬럼 매핑에 따라 RateExamplePayRow로 적재한다.
- 대상 시트: '① 5천만, 3천만↑' (5천만원↑ 블록만 처리)
- IBK는 별도 레이아웃 처리 (coverage_type = "[IBK]{상품명}")
- 파서에서 예외 발생 시 삼키지 않고 호출부(rate_example_normalizer.py)에서 rollback

보안/운영 원칙:
- 파일 URL 직접 접근 금지. FieldFile.path만 서버 내부에서 사용.
- 파싱 실패 시 예외를 raise — 호출부 transaction.atomic()이 rollback 처리.

컬럼 설계 (8개, 영문 필드명 / verbose_name 한글):
- col_first : 초회
- col_yr1   : 1차년
- col_m13   : 13회
- col_yr2   : 2차년구간 (보험사별 13~24회, 14~24회 등 상이)
- col_yr3   : 3차년구간 (보험사별 25~36회, 25~35회 등 상이)
- col_m36   : 36회 — IM·한화·흥국만 별도 기재, 나머지 None
- col_m37   : 37회 — 삼성·신한·DB·KB·KDB·미래·카디프·메트만 별도 기재, 나머지 None
- col_yr4   : 4차년 이후 통합 구간 — ABL(37~48회)·처브(37~48회)·한화(37~42회)·교보(37-39회)만 존재, 나머지 None
"""

import logging
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP

from openpyxl import load_workbook
from commission.models import RateExample
from commission.services.rate_example.fire.pay.parser import build_fire_pay_rows

logger = logging.getLogger(__name__)


# ── 지급률 저장 정책 ─────────────────────────────────────────────────────────
#
# 생명보험 지급률 파일은 raw 지급률을 100% 기준 지급률로 역산 저장한다.
#
# 주의:
# - 환산율/수정률(conv) 저장 정책과 혼동 금지
# - 지급률(pay) 저장값 = raw 지급률 / 0.97
#
PAY_NORMALIZE_DIVISOR = Decimal("0.97")
PAY_QUANT = Decimal("0.0001")

# ── 대상 시트명 ───────────────────────────────────────────────────────────────
TARGET_SHEET = "① 5천만, 3천만↑"

# ── 상품군 정규화 매핑 ────────────────────────────────────────────────────────
# raw 파일 col4 값 → DB 저장 coverage_type
COVERAGE_MAP: dict[str, str] = {
    "종신,CI":     "종신/CI",
    "연금":        "연금",
    "변액연금":    "변액연금",
    "저축":        "저축",
    "VUL":        "VUL",
    "연금저축":    "연금저축",
    "기타(보장성)": "기타(보장성)",
    "CEO정기":    "CEO정기",
    "전략상품1":  "전략상품1",
    "전략상품2":  "전략상품2",
    "전략상품3":  "전략상품3",
    "전략상품4":  "전략상품4",
}

# ── 보험사별 컬럼 매핑 ─────────────────────────────────────────────────────────
# tuple 형식: (col_first, col_yr1, col_m13, col_yr2, col_yr3, col_m36, col_m37, col_yr4)
# 값: 1-indexed 열 번호. None = 해당 보험사에 없는 회차 (DB에 NULL 저장).
#
# ── 36회/37회/4차년 유무 근거 (파일 헤더 직접 분석) ──────────────────────────
# col_m36 있음: IM(col17=36회), 한화(col31=36회), 흥국(col23=36회)
# col_m37 있음: 삼성(col17=37회), 신한(col24=37회), DB(col10=37회),
#               KB(col24=37회), KDB(col10=37회), 미래(col17=37회),
#               카디프(col39=37회), 메트(col16=37회)
# col_yr4 있음: ABL(col10=37~48회), 처브(col24=37~48회),
#               한화(col32=37~42회), 교보(col36=37-39회)
# ─────────────────────────────────────────────────────────────────────────────

# 그룹1 (헤더행 3, 데이터 5천만↑ 행 5~16)
# ABL  : first=5, yr1=6, m13=7, yr2=8(14~24회/19~24회), yr3=9(25~36회), yr4=10(37~48회)
# 삼성 : first=12, yr1=13, m13=14, yr2=15(19~24회), yr3=16(25~36회), m37=17
# 신한 : first=19, yr1=20, m13=21, yr2=22(14~24회), yr3=23(25~36회), m37=24
# 하나 : first=26, yr1=27, m13=28, yr2=29(14~24회), yr3=30(25~36회)
_GROUP1_ROWS = (5, 16)
_GROUP1: dict[str, tuple] = {
    #        first  yr1   m13   yr2   yr3   m36   m37   yr4
    "ABL":  ( 5,    6,    7,    8,    9,   None, None,   10),
    "삼성":  (12,   13,   14,   15,   16,  None,   17,  None),
    "신한":  (19,   20,   21,   22,   23,  None,   24,  None),
    "하나":  (26,   27,   28,   29,   30,  None, None,  None),
    # IBK는 별도 처리 블록 사용
}

# 그룹2 (헤더행 30, 데이터 5천만↑ 행 32~43)
# DB   : first=5,  yr1=6,  m13=7,  yr2=8(13~24회), yr3=9(25~36회), m36=10, m37=11
# IM   : first=13, yr1=14, m13=15, yr2=16(14~24회/19~24회), yr3=17(25~35회), m36=18
# KB   : first=20, yr1=21, m13=22, yr2=23(14~24회), yr3=24(25~36회), m37=25
# 농협 : first=27, yr1=28, m13=29, yr2=30(19~24회), yr3=31(25~36회)
# 라이나: first=33, yr1=34, m13=35, yr2=36(14~24회/14~18회), yr3=37(25~36회)
_GROUP2_ROWS = (32, 43)
_GROUP2: dict[str, tuple] = {
    #          first  yr1   m13   yr2   yr3   m36   m37   yr4
    "DB":    ( 5,    6,    7,    8,    9,    10,   11,  None),
    "IM":    (13,   14,   15,   16,   17,    18,  None,  None),
    "KB":    (20,   21,   22,   23,   24,  None,   25,  None),
    "NH":    (27,   28,   29,   30,   31,  None, None,  None),
    "라이나":  (33,   34,   35,   36,   37,  None, None,  None),
}

# 그룹3 (헤더행 57, 데이터 5천만↑ 행 59~70)
# KDB  : first=5, yr1=6, m13=7, yr2=8(14~24회), yr3=9(25~36회), m37=10
# 미래 : first=12, yr1=13, m13=14, yr2=15(14~24회), yr3=16(25~36회), m37=17
# 처브 : first=19, yr1=20, m13=21, yr2=22(14~24회), yr3=23(25~36회), yr4=24(37~48회)
# 한화 : first=26, yr1=27, m13=28, yr2=29(14~24회), yr3=30(25~35회), m36=31, yr4=32(37~42회)
# 카디프: first=34, yr1=35(1차년/11~12회), m13=36, yr2=37(14~24회), yr3=38(25~36회), m37=39
_GROUP3_ROWS = (59, 70)
_GROUP3: dict[str, tuple] = {
    #           first  yr1   m13   yr2   yr3   m36   m37   yr4
    "KDB":    ( 5,    6,    7,    8,    9,   None,   10,  None),
    "미래":    (12,   13,   14,   15,   16,  None,   17,  None),
    "처브":    (19,   20,   21,   22,   23,  None, None,   24),
    "한화":    (26,   27,   28,   29,   30,    31, None,   32),
    "카디프":   (34,   35,   36,   37,   38,  None,   39,  None),
}

# 그룹4 (헤더행 84, 데이터 5천만↑ 행 86~97)
# 동양   : first=5, yr1=6, m13=7, yr2=8(19~24회), yr3=9(25~36회)
# 메트   : first=11, yr1=12, m13=13, yr2=14(14~24회), yr3=15(25~36회), m37=16
# 흥국   : first=18, yr1=19, m13=20, yr2=21(14~15회), yr3=22(25~36회), yr4=23(4차년)
#           ※ 흥국 특이: 2차년/3차년 구간과 4차년 컬럼명이 타 보험사와 상이
# 푸본현대: first=25, yr1=26, m13=27, yr2=28(14~24회), yr3=29(25~36회)
# 교보   : first=31, yr1=32, m13=33, yr2=34(14~24회), yr3=35(25~36회), yr4=36(37-39회)
_GROUP4_ROWS = (86, 97)
_GROUP4: dict[str, tuple] = {
    #             first  yr1   m13   yr2   yr3   m36   m37   yr4
    "동양":     ( 5,    6,    7,    8,    9,   None, None,  None),
    "메트":     (11,   12,   13,   14,   15,  None,   16,  None),
    "흥국":     (18,   19,   20,   21,   22,  None, None,   23),
    "푸본현대":  (25,   26,   27,   28,   29,  None, None,  None),
    "교보":     (31,   32,   33,   34,   35,  None, None,   36),
}

# IBK 전용 (그룹1 데이터 행 공유 5~16)
# IBK: first=35, yr1=36, m13=37, yr2=38(14~24회), yr3=39(25~36회)
#      m36/m37/yr4 없음
_IBK_ROWS        = (5, 16)
_IBK_PRODUCT_COL = 32       # 1-indexed: col32 = IBK 자체 상품명
#              first  yr1   m13   yr2   yr3   m36   m37   yr4
_IBK_COLS    = (35,   36,   37,   38,   39,  None, None,  None)

# 전체 그룹 목록 (순서 유지)
_GROUPS: list[tuple] = [
    (_GROUP1_ROWS, _GROUP1),
    (_GROUP2_ROWS, _GROUP2),
    (_GROUP3_ROWS, _GROUP3),
    (_GROUP4_ROWS, _GROUP4),
]


# ── 유틸 함수 ─────────────────────────────────────────────────────────────────

def _to_decimal(value) -> "Decimal | None":
    """
    raw 지급률 셀 값 → 100% 기준 지급률 Decimal(소수점 4자리).

    정책:
        최종 저장값 = raw 지급률 / 0.97
    """
    if value is None:
        return None
    
    try:
        raw = Decimal(str(value).replace(",", "").strip())
        return (raw / PAY_NORMALIZE_DIVISOR).quantize(      
            PAY_QUANT,
            rounding=ROUND_HALF_UP, 
        )

    except (InvalidOperation, TypeError, ValueError):
        return None


def _get_col(row: tuple, col_1indexed: "int | None") -> "Decimal | None":
    """1-indexed 열 번호 → row tuple에서 값 추출 후 Decimal 변환."""
    if col_1indexed is None:
        return None
    idx = col_1indexed - 1
    if idx < 0 or idx >= len(row):
        return None
    return _to_decimal(row[idx])


# ── 정규화 메인 함수 ───────────────────────────────────────────────────────────

def normalize_pay_rate_example(
    example,                       # commission.models.RateExample 인스턴스
    normalize_mode: str = "replace",
) -> int:
    """
    지급률 xlsx → RateExamplePayRow 정규화.

    반환: 생성된 RateExamplePayRow 수 (0이면 대상 없음)

    호출 조건:
        example.insurer_type == "life"
        example.category     == "pay"
        original_name 확장자  == ".xlsx"

    normalize_mode:
        "replace": 기존 insurer_type=life&category=pay 행 전체 삭제 후 새 적재 (기본값)
        "append" : 기존 행 유지 후 신규 행만 추가
    """
    # 늦은 import — 순환 참조 방지
    from commission.models import RateExamplePayRow

    if not example.file:
        logger.warning("pay normalizer: file field is empty. pk=%s", example.pk)
        return 0
    
    # ── 손해보험(fire) 지급률 정규화 ───────────────────────────────
    # 생보 지급률 고정 레이아웃과 손보 지급률 레이아웃이 다르므로
    # fire는 전용 parser로 위임한다.
    if example.insurer_type == RateExample.TYPE_FIRE:
        normalized = build_fire_pay_rows(example)

        if normalize_mode == "replace":
            deleted_count, _ = RateExamplePayRow.objects.filter(
                insurer_type=RateExample.TYPE_FIRE,
                category="pay",
            ).delete()
            logger.info(
                "fire pay normalizer: replace mode — deleted %d rows. pk=%s",
                deleted_count,
                example.pk,
            )

        if normalized:
            RateExamplePayRow.objects.bulk_create(normalized, batch_size=500)

        logger.info(
            "fire pay normalizer: created %d rows. pk=%s normalize_mode=%s",
            len(normalized),
            example.pk,
            normalize_mode,
        )
        return len(normalized)

    fname = str(example.original_name or "").lower()
    if not fname.endswith(".xlsx"):
        logger.info(
            "pay normalizer: skip non-xlsx file. pk=%s original_name=%s",
            example.pk, example.original_name,
        )
        return 0

    # ── Workbook 로드 (data_only=True로 수식 결과값 읽기) ──────────────────────
    wb = load_workbook(example.file.path, data_only=True, read_only=False)

    if TARGET_SHEET not in wb.sheetnames:
        logger.warning(
            "pay normalizer: target sheet '%s' not found. pk=%s sheets=%s",
            TARGET_SHEET, example.pk, wb.sheetnames,
        )
        wb.close()
        return 0

    ws = wb[TARGET_SHEET]
    # 최대 110행까지 읽음 (그룹4 5천만↑ 마지막 행 97 + 여유 13행)
    all_rows: list[tuple] = list(ws.iter_rows(min_row=1, max_row=110, values_only=True))
    wb.close()

    normalized: list = []

    # ── 일반 19개 보험사 처리 ──────────────────────────────────────────────────
    for (row_start, row_end), insurer_map in _GROUPS:
        for insurer_name, col_tuple in insurer_map.items():
            # col_tuple: (first, yr1, m13, yr2, yr3, m36, m37, yr4)
            c_first, c_yr1, c_m13, c_yr2, c_yr3, c_m36, c_m37, c_yr4 = col_tuple
            coverage_carry: "str | None" = None

            for row_1idx in range(row_start, row_end + 1):
                r = all_rows[row_1idx - 1]

                # col4(index 3) = 상품군 셀. 공란이면 직전 값 이어받기 (병합셀 대응)
                raw_cov = r[3]
                if raw_cov:
                    coverage_carry = str(raw_cov).strip()

                coverage_type = COVERAGE_MAP.get(coverage_carry or "", coverage_carry or "")

                normalized.append(RateExamplePayRow(
                    source_file   = example,
                    source_sheet  = TARGET_SHEET,
                    source_row_no = row_1idx,
                    insurer_type  = example.insurer_type,
                    category      = "pay",
                    insurer       = insurer_name,
                    tier          = "5천만↑",
                    coverage_type = coverage_type,
                    col_first = _get_col(r, c_first),
                    col_yr1   = _get_col(r, c_yr1),
                    col_m13   = _get_col(r, c_m13),
                    col_yr2   = _get_col(r, c_yr2),
                    col_yr3   = _get_col(r, c_yr3),
                    col_m36   = _get_col(r, c_m36),
                    col_m37   = _get_col(r, c_m37),
                    col_yr4   = _get_col(r, c_yr4),
                ))

    # ── IBK 전용 처리 ─────────────────────────────────────────────────────────
    # IBK는 상품군(col4) 대신 col32 자체 상품명을 coverage_type 키로 사용
    ibk_first, ibk_yr1, ibk_m13, ibk_yr2, ibk_yr3, ibk_m36, ibk_m37, ibk_yr4 = _IBK_COLS

    for row_1idx in range(_IBK_ROWS[0], _IBK_ROWS[1] + 1):
        r = all_rows[row_1idx - 1]
        raw_product = r[_IBK_PRODUCT_COL - 1]   # col32 → index 31
        if not raw_product:
            continue
        coverage_type = f"[IBK]{str(raw_product).strip()}"

        normalized.append(RateExamplePayRow(
            source_file   = example,
            source_sheet  = TARGET_SHEET,
            source_row_no = row_1idx,
            insurer_type  = example.insurer_type,
            category      = "pay",
            insurer       = "IBK",
            tier          = "5천만↑",
            coverage_type = coverage_type,
            col_first = _get_col(r, ibk_first),
            col_yr1   = _get_col(r, ibk_yr1),
            col_m13   = _get_col(r, ibk_m13),
            col_yr2   = _get_col(r, ibk_yr2),
            col_yr3   = _get_col(r, ibk_yr3),
            col_m36   = _get_col(r, ibk_m36),
            col_m37   = _get_col(r, ibk_m37),
            col_yr4   = _get_col(r, ibk_yr4),
        ))

    # ── replace / append 모드 처리 ─────────────────────────────────────────────
    if normalize_mode == "replace":
        deleted_count, _ = RateExamplePayRow.objects.filter(
            insurer_type=example.insurer_type,
            category="pay",
        ).delete()
        logger.info(
            "pay normalizer: replace mode — deleted %d existing rows. pk=%s",
            deleted_count, example.pk,
        )

    if normalized:
        RateExamplePayRow.objects.bulk_create(normalized, batch_size=500)

    logger.info(
        "pay normalizer: created %d rows. pk=%s insurer_type=%s normalize_mode=%s",
        len(normalized), example.pk, example.insurer_type, normalize_mode,
    )
    return len(normalized)