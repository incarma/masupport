# commission/services/rate_example/fire/lotte/parser.py
from __future__ import annotations

"""
롯데손해보험 수정률 정규화 parser.

역할:
- 롯데손해보험 RAW 수정률 xlsx를 RateExampleConversionRow 기준으로 정규화한다.
- 좌/우 병렬 테이블 구조를 상품 블록 pair 단위로 해석한다.

핵심 규칙:
- 우측 블록에 "좌동"이 있으면 좌측 블록 사용
- 우측 블록에 "판매중지"가 있으면 좌/우 블록 전체 제외
- 그 외에는 우측 블록 사용
- 상품/담보구분/수정률 병합셀은 parser 내부 value matrix에서만 전파
- 수정률은 raw 숫자 그대로 year1에 저장
  예: raw 130 → DB Decimal("130") → 화면 130%
"""

import logging
import re
from dataclasses import dataclass
from decimal import Decimal
from typing import Any

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from commission.models import RateExample, RateExampleConversionRow
from commission.services.rate_example.common import (
    build_worksheet_value_map,
    clean_spaces,
    decimal_from_text,
)

logger = logging.getLogger(__name__)


INSURER = "롯데"

STATUS_SAME = "same"
STATUS_STOPPED = "stopped"
STATUS_NORMAL = "normal"

LEFT_PRODUCT_COL = 2   # B: 상품
LEFT_PLAN_COL = 3      # C: 담보구분
LEFT_PAY_COL = 4       # D: 납입기간
LEFT_RATE_COL = 5      # E: 수정률/신계약

RIGHT_PRODUCT_COL = 9  # I: 상품
RIGHT_PLAN_COL = 10    # J: 담보구분 또는 좌동/판매중지
RIGHT_PAY_COL = 11     # K: 납입기간
RIGHT_RATE_COL = 12    # L: 수정률/신계약

START_DATA_ROW = 12

EXCLUDE_PRODUCT_START_KEYWORDS = (
    "공통사항",
    "변경전",
    "변경후",
    "상품",
    "■",
    "※",
)


@dataclass(frozen=True)
class TableCols:
    """좌/우 테이블의 상품/구분/납기/수정률 컬럼 위치."""

    product: int
    plan: int
    pay: int
    rate: int


@dataclass(frozen=True)
class ProductBlockPair:
    """동일 상품에 대한 좌/우 병렬 블록 pair."""

    start_row: int
    end_row: int
    status: str


LEFT_COLS = TableCols(
    product=LEFT_PRODUCT_COL,
    plan=LEFT_PLAN_COL,
    pay=LEFT_PAY_COL,
    rate=LEFT_RATE_COL,
)

RIGHT_COLS = TableCols(
    product=RIGHT_PRODUCT_COL,
    plan=RIGHT_PLAN_COL,
    pay=RIGHT_PAY_COL,
    rate=RIGHT_RATE_COL,
)


def _norm_text(value: Any) -> str:
    """
    셀 텍스트 정규화.

    - 여러 줄 텍스트는 한 줄로 연결
    - 중복 공백 제거
    - 앞뒤 공백 제거
    """
    if value is None:
        return ""

    return clean_spaces(str(value).replace("\r", "\n"))


def _to_decimal(value: Any) -> Decimal | None:
    """
    수정률 raw 값을 Decimal로 변환한다.

    주의:
    - 100으로 나누거나 곱하지 않는다.
    - "130"은 Decimal("130") 그대로 저장한다.
    - "없음", "-", 공란은 저장 제외한다.
    """
    text = _norm_text(value)
    if not text:
        return None

    if text in {"-", "없음"}:
        return None

    return decimal_from_text(text)


def _build_value_matrix(ws: Worksheet) -> dict[tuple[int, int], Any]:
    """
    병합셀 전개 matrix 생성.

    실제 workbook에는 unmerge_cells()를 실행하지 않는다.
    parser 내부 dict에서만 병합 범위 전체에 좌상단 값을 전파한다.
    """
    return build_worksheet_value_map(ws, include_empty=False)


def _looks_like_product_start(text: str) -> bool:
    """
    상품 블록 시작행 여부 판별.

    원본 셀 값 기준으로만 판단한다.
    병합셀 전파 matrix 기준으로 판단하면 모든 하위 행이 상품 시작행이 되는 문제가 있다.
    """
    value = _norm_text(text)
    if not value:
        return False

    if any(keyword in value for keyword in EXCLUDE_PRODUCT_START_KEYWORDS):
        return False

    return True


def _find_product_start_rows(ws: Worksheet) -> list[int]:
    """
    좌/우 병렬 상품 블록의 시작행을 찾는다.

    롯데 RAW는 좌측 B열 또는 우측 I열에 상품명이 새로 등장하는 행이
    하나의 상품 block pair 시작점이다.
    """
    starts: list[int] = []

    for row_no in range(START_DATA_ROW, ws.max_row + 1):
        left_product = _norm_text(ws.cell(row_no, LEFT_PRODUCT_COL).value)
        right_product = _norm_text(ws.cell(row_no, RIGHT_PRODUCT_COL).value)

        if _looks_like_product_start(left_product) or _looks_like_product_start(right_product):
            starts.append(row_no)

    return starts


def _detect_pair_status(
    values: dict[tuple[int, int], Any],
    *,
    start_row: int,
    end_row: int,
) -> str:
    """
    우측 블록 상태 판별.

    우측 block 범위에서 "판매중지"가 있으면 pair 전체 제외.
    "좌동"이 있으면 좌측 block을 기준으로 사용.
    그 외에는 우측 block 기준.
    """
    right_texts: list[str] = []

    for row_no in range(start_row, end_row + 1):
        for col_no in range(RIGHT_PRODUCT_COL, RIGHT_RATE_COL + 1):
            value = _norm_text(values.get((row_no, col_no)))
            if value:
                right_texts.append(value)

    joined = " ".join(right_texts)

    if "판매중지" in joined:
        return STATUS_STOPPED

    if "좌동" in joined:
        return STATUS_SAME

    return STATUS_NORMAL


def _build_block_pairs(
    ws: Worksheet,
    values: dict[tuple[int, int], Any],
) -> list[ProductBlockPair]:
    """상품 시작행 기준으로 좌/우 block pair 목록을 구성한다."""
    starts = _find_product_start_rows(ws)
    pairs: list[ProductBlockPair] = []

    for index, start_row in enumerate(starts):
        end_row = starts[index + 1] - 1 if index + 1 < len(starts) else ws.max_row
        status = _detect_pair_status(values, start_row=start_row, end_row=end_row)

        pairs.append(
            ProductBlockPair(
                start_row=start_row,
                end_row=end_row,
                status=status,
            )
        )

    return pairs


def _target_cols_for_pair(pair: ProductBlockPair) -> TableCols | None:
    """
    pair 상태에 따른 정규화 기준 테이블 결정.

    - 좌동: 좌측 테이블 기준
    - 판매중지: pair 전체 제외
    - 일반: 우측 테이블 기준
    """
    if pair.status == STATUS_STOPPED:
        return None

    if pair.status == STATUS_SAME:
        return LEFT_COLS

    return RIGHT_COLS


def _resolve_coverage_type(
    *,
    product_name: str,
    plan_type: str,
    pay_period: str,
) -> str:
    """
    롯데손해보험 상품군 정규화.

    우선순위:
    1. 상품명 또는 담보구분에 실손 포함
       - 납입기간에 최초 포함: 단독실손(초회)
       - 납입기간에 갱신 포함: 단독실손(갱신)
    2. 상품명에 연금 포함: 연금
    3. 상품명에 저축 포함: 저축
    4. 그 외: 보장
    """
    product = _norm_text(product_name)
    plan = _norm_text(plan_type)
    pay = _norm_text(pay_period)

    if "실손" in product or "실손" in plan:
        if "최초" in pay:
            return "단독실손(초회)"
        if "갱신" in pay:
            return "단독실손(갱신)"
        return "단독실손(갱신)"

    if "연금" in product:
        return "연금"

    if "저축" in product:
        return "저축"

    return "보장"


def _is_data_row(
    *,
    product_name: str,
    plan_type: str,
    pay_period: str,
    rate: Decimal | None,
) -> bool:
    """
    실제 저장 대상 행 여부.

    안내문/구분선/상품 제목만 있는 행은 rate가 없으므로 제외한다.
    """
    if not product_name or not plan_type or not pay_period:
        return False

    if rate is None:
        return False

    return True


def build_fire_lotte_conversion_rows(
    example: RateExample,
    wb: Workbook,
) -> list[RateExampleConversionRow]:
    """
    롯데손해보험 수정률 정규화 row 생성.

    반환 row는 normalize_rate_example()에서 replace/append 정책에 따라 bulk_create 된다.
    """
    rows: list[RateExampleConversionRow] = []

    for ws in wb.worksheets:
        values = _build_value_matrix(ws)
        pairs = _build_block_pairs(ws, values)

        for pair in pairs:
            target_cols = _target_cols_for_pair(pair)

            # 판매중지 pair는 좌/우 모두 정규화 제외
            if target_cols is None:
                continue

            for row_no in range(pair.start_row, pair.end_row + 1):
                product_name = _norm_text(values.get((row_no, target_cols.product)))
                plan_type = _norm_text(values.get((row_no, target_cols.plan)))
                pay_period = _norm_text(values.get((row_no, target_cols.pay)))
                rate = _to_decimal(values.get((row_no, target_cols.rate)))

                if not _is_data_row(
                    product_name=product_name,
                    plan_type=plan_type,
                    pay_period=pay_period,
                    rate=rate,
                ):
                    continue

                coverage_type = _resolve_coverage_type(
                    product_name=product_name,
                    plan_type=plan_type,
                    pay_period=pay_period,
                )

                rows.append(
                    RateExampleConversionRow(
                        source_file=example,
                        source_sheet=ws.title,
                        source_row_no=row_no,
                        insurer_type=RateExample.TYPE_FIRE,
                        category=RateExample.CAT_CONV,
                        insurer=INSURER,
                        coverage_type=coverage_type,
                        strategy_flag="",
                        product_name=product_name,
                        plan_type=plan_type,
                        pay_period=pay_period,
                        year1=rate,
                        year2=None,
                        year3=None,
                        year4=None,
                    )
                )

    logger.info(
        "lotte fire normalizer: created %s rows. pk=%s",
        len(rows),
        getattr(example, "pk", None),
    )

    return rows
