# commission/services/rate_example/fire/hanhwa/parser.py
from __future__ import annotations

"""
한화손해보험 수정률 정규화 parser.

역할:
- 한화손해보험 raw xlsx 파일을 RateExampleConversionRow 구조로 정규화한다.
- 손해보험 수정률은 단일 수정률 구조이므로 year1 필드에만 저장한다.

정규화 규칙:
- 보험사: "한화" 고정
- insurer_type: fire
- category: conv
- B열 값이 "구분"인 테이블을 상품 단위로 인식
- 최초 "구분" 셀 2행 상단의 "○" 포함 텍스트를 상품명으로 사용
- B/C열 구분이 다르면 "B (C)" 형태로 결합
- "신계약환산율" 세부 헤더 또는 "환산율" 포함 헤더를 납기로 사용
- 상품군이 "보장(태아)"인 경우 구분 컬럼에서 "주)" 텍스트 제거
- 수정률 데이터는 숫자만 추출하여 백분율 수치로 year1에 저장
"""

import re
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl.workbook.workbook import Workbook

from commission.models import RateExample, RateExampleConversionRow


PRODUCT_GROUPS = {
    "보장",
    "보장(태아)",
    "연금",
    "저축",
    "단독실손(초회)",
    "단독실손(갱신)",
}


def _to_text(value: Any) -> str:
    """셀 값을 비교/저장 가능한 단일 문자열로 정리한다."""
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n+", "\n", text)
    return text.strip()


def _flat_text(value: Any) -> str:
    """줄바꿈까지 공백 하나로 합친 비교용 문자열."""
    text = _to_text(value)
    return re.sub(r"\s+", " ", text).strip()


def _norm_key(value: Any) -> str:
    """
    헤더 비교용 정규화.

    한화 raw 파일은 B열 헤더가 '구분'이 아니라 '구  분'처럼
    중간 공백을 포함하므로, 헤더 비교 시 모든 공백을 제거한다.
    """
    return re.sub(r"\s+", "", _flat_text(value))


def _strip_product_marker(text: str) -> str:
    """
    상품명 앞 기호 제거.

    예:
    - "○ 한화 더건강한 한아름종합보험 무배당2604"
      -> "한화 더건강한 한아름종합보험 무배당2604"
    """
    text = _flat_text(text)
    text = re.sub(r"^[○●◯◎ㆍ\-\s]+", "", text)
    return text.strip()


def _clean_period(text: str) -> str:
    """
    납기 헤더 정리.

    - "신계약환산율" 자체는 납기로 쓰지 않는다.
    - "만기후재가입시 및 갱신시 환산율"은 "만기후재가입시 및 갱신시"로 저장한다.
    """
    text = _flat_text(text)
    text = text.replace("신계약환산율", "")
    text = text.replace("환산율", "")
    text = re.sub(r"\s+", " ", text)
    return text.strip(" :-")


def _build_plan_type(col_b: str, col_c: str) -> str:
    """
    B/C열 구분 조합 규칙.

    - B와 C가 같으면 B만 사용
    - B와 C가 다르면 "B (C)"로 저장
    """
    b = _flat_text(col_b)
    c = _flat_text(col_c)

    if not b:
        return ""
    if not c or b == c:
        return b
    return f"{b} ({c})"


def _cleanup_hanhwa_plan_type(plan_type: str, coverage_type: str) -> str:
    """
    한화손보 전용 구분 후처리.

    추가 요구사항:
    - 상품군이 "보장(태아)"인 경우 구분 컬럼 데이터에서 "주)" 텍스트 삭제
    - 다른 상품군은 영향 없도록 제한한다.
    """
    text = _flat_text(plan_type)

    if coverage_type == "보장(태아)":
        text = text.replace("주)", "")
        text = re.sub(r"\s+", " ", text).strip()

    return text


def _classify_coverage_type(product_name: str, plan_type: str, pay_period: str) -> str:
    """
    상품군 분류 규칙.

    우선순위:
    1. 실손: 납기 기준 초회/갱신 분류
    2. 연금
    3. 저축
    4. 태아관련
    5. 기본 보장
    """
    product = _flat_text(product_name)
    plan = _flat_text(plan_type)
    period = _flat_text(pay_period)

    if "실손" in product:
        if "최초" in period:
            return "단독실손(초회)"
        if "갱신" in period:
            return "단독실손(갱신)"
        return "단독실손(갱신)"

    if "연금" in product:
        return "연금"

    if "저축" in product:
        return "저축"

    if "태아관련" in plan:
        return "보장(태아)"

    return "보장"


def _to_percent_decimal(value: Any, number_format: str = "") -> Decimal | None:
    """
    수정률 값을 DB 저장용 Decimal로 변환한다.

    저장 정책:
    - 화면 표시 기준 백분율 수치 그대로 저장한다.
    - "160%" -> Decimal("160")
    - Excel 내부값 1.6 + percent format -> Decimal("160")
    - "160% 주)" 같은 부가 문구 -> Decimal("160")
    """
    if value is None:
        return None

    if isinstance(value, (int, float, Decimal)):
        try:
            dec = Decimal(str(value))
        except InvalidOperation:
            return None

        fmt = str(number_format or "")
        if "%" in fmt and dec <= Decimal("10"):
            dec = dec * Decimal("100")
        return dec.quantize(Decimal("0.0001"))

    text = _flat_text(value)
    if not text:
        return None

    # 숫자, 소수점, 음수부호 외 문자는 제거한다.
    match = re.search(r"-?\d+(?:\.\d+)?", text.replace(",", ""))
    if not match:
        return None

    try:
        dec = Decimal(match.group(0))
    except InvalidOperation:
        return None

    return dec.quantize(Decimal("0.0001"))


def _merged_matrix(ws) -> tuple[list[list[Any]], list[list[str]]]:
    """
    병합 셀 전개 matrix 생성.

    openpyxl 병합 셀은 좌상단 셀에만 값이 있으므로,
    정규화 전 병합 범위 전체에 좌상단 값을 전파한다.
    """
    max_row = ws.max_row
    max_col = ws.max_column

    values = [
        [None for _ in range(max_col + 1)]
        for _ in range(max_row + 1)
    ]
    formats = [
        ["" for _ in range(max_col + 1)]
        for _ in range(max_row + 1)
    ]

    for row in ws.iter_rows():
        for cell in row:
            values[cell.row][cell.column] = cell.value
            formats[cell.row][cell.column] = str(cell.number_format or "")

    for merged_range in ws.merged_cells.ranges:
        min_col, min_row, max_col_range, max_row_range = merged_range.bounds
        top_value = values[min_row][min_col]
        top_format = formats[min_row][min_col]

        for r in range(min_row, max_row_range + 1):
            for c in range(min_col, max_col_range + 1):
                values[r][c] = top_value
                formats[r][c] = top_format

    return values, formats


def _find_product_name(values: list[list[Any]], header_row: int, max_col: int) -> str:
    """
    B열 "구분" 헤더 기준 2행 상단에서 "○" 포함 상품명을 찾는다.

    기본 규칙은 header_row - 2 이지만,
    raw 병합/공백 변동에 대비하여 header_row - 4 ~ header_row - 1 범위도 보조 탐색한다.
    """
    candidate_rows = [header_row - 2, header_row - 3, header_row - 4, header_row - 1]

    for r in candidate_rows:
        if r < 1:
            continue

        row_texts = [_flat_text(values[r][c]) for c in range(1, max_col + 1)]
        marked = [text for text in row_texts if "○" in text]
        if marked:
            return _strip_product_marker(marked[0])

    return ""


def _find_rate_columns(
    values: list[list[Any]],
    header_row: int,
    max_col: int,
) -> list[tuple[int, str]]:
    """
    수정률 컬럼과 납기명을 찾는다.

    1. "신계약환산율" 헤더가 있으면 그 하위 세부 헤더를 납기로 사용한다.
    2. "신계약환산율"이 없으면 "구분" 컬럼 우측의 "환산율" 포함 헤더를 납기로 사용한다.
    """
    rate_cols: list[tuple[int, str]] = []

    # case 1: 신계약환산율 병합 헤더 + 하위 세부 헤더
    for c in range(4, max_col + 1):
        header = _flat_text(values[header_row][c])
        sub_header = _flat_text(values[header_row + 1][c]) if header_row + 1 < len(values) else ""

        if "신계약환산율" in header:
            pay_period = _clean_period(sub_header or header)
            if pay_period:
                rate_cols.append((c, pay_period))

    if rate_cols:
        return rate_cols

    # case 2: "구분" 우측의 "*환산율" 단일 헤더
    for c in range(4, max_col + 1):
        header = _flat_text(values[header_row][c])
        if "환산율" not in header:
            continue

        pay_period = _clean_period(header)
        if pay_period:
            rate_cols.append((c, pay_period))

    return rate_cols


def _is_next_table(values: list[list[Any]], row_no: int) -> bool:
    """다음 상품 테이블 시작 여부."""
    return _norm_key(values[row_no][2]) == "구분"


def _is_effective_data_row(values: list[list[Any]], row_no: int) -> bool:
    """구분 데이터가 있는 실제 데이터 행인지 판단한다."""
    col_b = _flat_text(values[row_no][2])
    if not col_b:
        return False
    if _norm_key(col_b) == "구분":
        return False
    if col_b.startswith("※") or col_b.startswith("주)"):
        return False
    return True


def _build_rows_from_table(
    *,
    example: RateExample,
    ws,
    values: list[list[Any]],
    formats: list[list[str]],
    header_row: int,
) -> list[RateExampleConversionRow]:
    """단일 상품 테이블을 RateExampleConversionRow 목록으로 변환한다."""
    max_col = ws.max_column
    max_row = ws.max_row

    product_name = _find_product_name(values, header_row, max_col)
    if not product_name:
        return []

    rate_columns = _find_rate_columns(values, header_row, max_col)
    if not rate_columns:
        return []

    rows: list[RateExampleConversionRow] = []

    start_row = header_row + 2
    if not any(_is_effective_data_row(values, r) for r in range(start_row, min(start_row + 3, max_row + 1))):
        start_row = header_row + 1

    for r in range(start_row, max_row + 1):
        if r != start_row and _is_next_table(values, r):
            break

        if not _is_effective_data_row(values, r):
            continue

        base_plan_type = _build_plan_type(
            _flat_text(values[r][2]),
            _flat_text(values[r][3]),
        )
        if not base_plan_type:
            continue

        for col_idx, pay_period in rate_columns:
            if _norm_key(pay_period) == "수금":
                continue

            rate_value = _to_percent_decimal(
                values[r][col_idx],
                number_format=formats[r][col_idx],
            )
            if rate_value is None:
                continue

            coverage_type = _classify_coverage_type(
                product_name=product_name,
                plan_type=base_plan_type,
                pay_period=pay_period,
            )

            if coverage_type not in PRODUCT_GROUPS:
                continue

            plan_type = _cleanup_hanhwa_plan_type(
                base_plan_type,
                coverage_type,
            )

            rows.append(
                RateExampleConversionRow(
                    source_file=example,
                    source_sheet=ws.title,
                    source_row_no=r,
                    insurer_type=example.insurer_type,
                    category=example.category,
                    insurer="한화",
                    coverage_type=coverage_type,
                    strategy_flag="",
                    product_name=product_name,
                    plan_type=plan_type,
                    pay_period=pay_period,
                    year1=rate_value,
                    year2=None,
                    year3=None,
                    year4=None,
                )
            )

    return rows


def _dedupe_rows(
    rows: list[RateExampleConversionRow],
) -> list[RateExampleConversionRow]:
    """
    동일 상품/구분/납기 중복 방지.

    병합 셀 전개 또는 헤더 중복 인식으로 동일 행이 반복 생성될 수 있으므로,
    저장 전 정규화 master 기준으로 중복을 제거한다.
    """
    seen: set[tuple[str, str, str, str, Decimal | None]] = set()
    deduped: list[RateExampleConversionRow] = []

    for row in rows:
        key = (
            row.coverage_type,
            row.product_name,
            row.plan_type,
            row.pay_period,
            row.year1,
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(row)

    return deduped


def build_fire_hanhwa_conversion_rows(
    example: RateExample,
    wb: Workbook,
) -> list[RateExampleConversionRow]:
    """
    한화손해보험 수정률 정규화 진입점.

    normalize_rate_example()에서 workbook을 안전 로드한 뒤 호출한다.
    """
    rows: list[RateExampleConversionRow] = []

    for ws in wb.worksheets:
        values, formats = _merged_matrix(ws)

        for r in range(1, ws.max_row + 1):
            if _norm_key(values[r][2]) != "구분":
                continue

            rows.extend(
                _build_rows_from_table(
                    example=example,
                    ws=ws,
                    values=values,
                    formats=formats,
                    header_row=r,
                )
            )

    return _dedupe_rows(rows)
