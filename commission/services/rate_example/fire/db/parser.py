# commission/services/rate_example/fire/db/parser.py
from __future__ import annotations

"""
DB 손해보험 수정률 정규화.

역할:
- DB손해보험 raw xlsx의 각 시트에서 "1. 수정률(GA)" 테이블만 읽는다.
- A열에서 "2. 수금수수료율"이 발견된 행부터 아래는 정규화하지 않는다.
- 정규화 결과는 RateExampleConversionRow에 저장한다.

저장 정책:
- insurer_type = "fire"
- category = "conv"
- insurer = "DB"
- product_name = 시트명
- year1 = 수정률 단일값
- year2~year4 = None
- 수정률은 raw 셀 값을 그대로 저장한다.
  예: Excel 내부값 2.4 → DB Decimal("2.4")
  즉, 2.4를 240으로 변환하지 않는다.
"""

from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from commission.models import RateExample, RateExampleConversionRow
from commission.services.rate_example.common.excel import (
    build_merged_value_map,
    cell_value_with_merged,
)
from commission.services.rate_example.common.text import clean_spaces


# =============================================================================
# 기본 텍스트/숫자 정규화
# =============================================================================


def _text(value: Any) -> str:
    """셀 값을 비교/저장 가능한 문자열로 정리한다."""
    return clean_spaces(str(value or "").replace("\r\n", "\n").replace("\r", "\n"))


def _compact(value: Any) -> str:
    """헤더 탐지용: 공백 제거 문자열."""
    return _text(value).replace(" ", "")


def _to_decimal_raw(value: Any, *, number_format: str = "") -> Decimal | None:
    """
    수정률 raw 값을 Decimal로 변환한다.

    중요:
    - DB 손보 수정률은 Excel 내부값(raw)을 그대로 저장한다.
    - 표시용 % 변환은 조회 API/UI에서만 처리한다.

    예:
    - Excel 내부값 2.4  + %서식 → DB 2.4 저장 → 화면 240%
    - Excel 내부값 0.55 + %서식 → DB 0.55 저장 → 화면 55%

    즉:
    - 저장 단계에서는 절대 ×100 보정하지 않는다.
    """
    if value is None:
        return None

    if isinstance(value, Decimal):
        return value

    if isinstance(value, int | float):
        # Excel 내부 raw 값을 그대로 저장한다.
        # 절대 ×100 보정하지 않는다.
        return Decimal(str(value))

    raw = _text(value)
    if not raw:
        return None

    # 텍스트 셀:
    # - "240%" → 240 저장
    # - "55%"  → 55 저장
    #
    # numeric 셀은 위에서 raw 그대로 처리된다.
    # 즉:
    # - numeric 0.55 → DB 0.55
    # - text "55%"   → DB 55
    #
    # 현재 DB손보 원본은 대부분 numeric percent format이므로
    # 실제 핵심 수정은 위 numeric branch의 ×100 제거이다.
    raw = raw.replace(",", "").replace("%", "").strip()

    try:
        return Decimal(raw)
    except (InvalidOperation, ValueError):
        return None


# =============================================================================
# 병합 셀 처리
# =============================================================================


def _merged_value_map(ws: Worksheet) -> dict[tuple[int, int], Any]:
    """
    병합 셀 범위 전체에 좌상단 값을 전파한 lookup map을 만든다.

    raw 파일의 병합을 실제로 해제하지 않고도,
    정규화 로직에서는 병합 해제 후 모든 셀에 같은 값이 들어간 것처럼 읽는다.
    """
    return build_merged_value_map(ws)


def _cell(ws: Worksheet, merged: dict[tuple[int, int], Any], row: int, col: int) -> Any:
    """병합 전파값을 우선하여 셀 값을 읽는다."""
    return cell_value_with_merged(ws, merged, row, col)


# =============================================================================
# 테이블 범위/헤더 탐지
# =============================================================================


def _find_table_start_row(ws: Worksheet, merged: dict[tuple[int, int], Any]) -> int | None:
    """A열에서 '1. 수정률(GA)'가 포함된 행을 찾는다."""
    for row in range(1, ws.max_row + 1):
        value = _compact(_cell(ws, merged, row, 1))
        if "1.수정률(GA)" in value or "1.수정률" in value:
            return row
    return None


def _find_table_stop_row(ws: Worksheet, merged: dict[tuple[int, int], Any], start_row: int) -> int:
    """
    A열에서 '2. 수금수수료율'이 발견되면 해당 행부터 제외한다.
    없으면 sheet 마지막 행 + 1을 반환한다.
    """
    for row in range(start_row + 1, ws.max_row + 1):
        value = _compact(_cell(ws, merged, row, 1))
        if "2.수금수수료율" in value:
            return row
    return ws.max_row + 1


def _header_texts(ws: Worksheet, merged: dict[tuple[int, int], Any], row: int) -> dict[int, str]:
    """한 행의 헤더 문자열 dict를 만든다."""
    headers: dict[int, str] = {}
    for col in range(1, ws.max_column + 1):
        text = _text(_cell(ws, merged, row, col))
        if text:
            headers[col] = text
    return headers


def _find_header_row(
    ws: Worksheet,
    merged: dict[tuple[int, int], Any],
    start_row: int,
    stop_row: int,
) -> int | None:
    """
    수정률 테이블의 헤더 행을 찾는다.

    필수 성격:
    - '수정률' 컬럼 존재
    - 일반 상품: 납기 컬럼 존재
    - 실손 상품: 납기 컬럼 없이 최초/갱신 컬럼 허용
    - 만기 또는 갱신주기 컬럼 존재
    """
    search_to = min(stop_row, start_row + 15)

    for row in range(start_row + 1, search_to):
        joined = " ".join(_header_texts(ws, merged, row).values())
        compact = joined.replace(" ", "")

        has_rate = "수정률" in compact
        has_pay = "납기" in compact
        has_first_renewal = "최초/갱신" in compact or "최초갱신" in compact
        has_maturity = "만기" in compact or "갱신주기" in compact

        if has_rate and has_maturity and (has_pay or has_first_renewal):
            return row

    return None


def _find_cols(headers: dict[int, str], keyword: str) -> list[int]:
    """헤더에 keyword가 포함된 컬럼 목록."""
    return [col for col, text in headers.items() if keyword in _compact(text)]


def _find_first_col(headers: dict[int, str], keywords: list[str]) -> int | None:
    """여러 keyword 중 하나라도 포함된 첫 컬럼."""
    for col, text in headers.items():
        compact = _compact(text)
        if any(keyword in compact for keyword in keywords):
            return col
    return None


def _find_plan_cols(headers: dict[int, str]) -> list[int]:
    """
    '구분' 컬럼 후보를 찾는다.

    규칙:
    - 구분 / 플랜 / 담보 / 최초/갱신 단어 포함 컬럼 사용
    - 단, 갱신주기는 납기 조합용 컬럼이므로 제외
    """
    cols: list[int] = []

    for col, text in headers.items():
        compact = _compact(text)

        if "갱신주기" in compact:
            continue

        if (
            "구분" in compact
            or "플랜" in compact
            or "담보" in compact
            or "최초/갱신" in compact
            or compact in {"최초", "갱신"}
        ):
            cols.append(col)

    return cols


def _find_rate_col(headers: dict[int, str]) -> int | None:
    """
    수정률 컬럼을 찾는다.

    수정률이 1차/2차로 나뉜 경우 '1차' 컬럼을 우선한다.
    """
    rate_cols = _find_cols(headers, "수정률")
    if not rate_cols:
        return None

    for col in rate_cols:
        compact = _compact(headers[col])
        if "1차" in compact:
            return col

    return rate_cols[0]


# =============================================================================
# 도메인 정규화 규칙
# =============================================================================


def _coverage_type(product_name: str, plan_type: str) -> str:
    """
    DB 손보 상품군/보종 정규화.

    요청 지침:
    - 태아 포함: 보장(태아)
    - 연금 포함: 연금
    - 저축 포함: 저축
    - 실손 포함 + 구분 최초: 단독실손(초회)
    - 실손 포함 + 구분 갱신: 단독실손(갱신)
    - 그 외: 보장
    """
    product = product_name or ""
    plan = plan_type or ""

    if "태아" in product:
        return "보장(태아)"

    if "연금" in product:
        return "연금"

    if "저축" in product:
        return "저축"

    if "실손" in product:
        if "최초" in plan:
            return "단독실손(초회)"
        if "갱신" in plan:
            return "단독실손(갱신)"

    return "보장"


def _join_unique(parts: list[str], sep: str = "/") -> str:
    """중복 없이 순서를 보존해 문자열을 결합한다."""
    result: list[str] = []
    seen: set[str] = set()

    for part in parts:
        value = _text(part)
        if not value or value in seen:
            continue
        seen.add(value)
        result.append(value)

    return sep.join(result)


def _build_plan_type(
    ws: Worksheet,
    merged: dict[tuple[int, int], Any],
    row: int,
    plan_cols: list[int],
    car_col: int | None,
) -> str:
    """
    구분 정규화.

    - 구분/플랜/담보/최초·갱신 컬럼값을 결합
    - 차종 컬럼이 있으면 끝에 '(차종값)' 추가
    """
    base = _join_unique([_text(_cell(ws, merged, row, col)) for col in plan_cols])
    car = _text(_cell(ws, merged, row, car_col)) if car_col else ""

    if car:
        if base:
            return f"{base}({car})"
        return f"({car})"

    return base


def _normalize_special_family_pay(value: Any) -> str:
    """
    참좋은훼밀리더블플러스 전용 납기 컬럼 정규화.
    - 갱신 포함: 갱신형
    - 그 외: raw 그대로
    """
    text = _text(value)
    if not text:
        return ""
    if "갱신" in text:
        return "갱신형"
    return text


def _normalize_special_family_maturity(value: Any) -> str:
    """
    참좋은훼밀리더블플러스 전용 만기 컬럼 정규화.
    - 숫자만 있으면 '년납' 추가
    - 그 외 raw 그대로
    """
    text = _text(value)
    if not text:
        return ""
    if text.isdigit():
        return f"{text}년납"
    return text


def _normalize_maturity(value: Any) -> str:
    """
    일반 상품 만기/갱신주기 정규화.
    - 10 → 10년만기
    - 10년 → 10년만기
    - 그 외 raw 그대로
    """
    text = _text(value)
    if not text:
        return ""
    if text.isdigit():
        return f"{text}년만기"
    if text.endswith("년") and text[:-1].isdigit():
        return f"{text}만기"
    return text


def _normalize_pay_period(value: Any) -> str:
    """
    일반 상품 납기 정규화.
    - 10 → 10년납
    - 10년 → 10년납
    - 그 외 raw 그대로
    """
    text = _text(value)
    if not text:
        return ""
    if text.isdigit():
        return f"{text}년납"
    if text.endswith("년") and text[:-1].isdigit():
        return f"{text}납"
    return text


def _build_pay_period(
    *,
    product_name: str,
    maturity_value: Any,
    pay_value: Any,
) -> str:
    """
    납기 최종 문자열 생성.

    특수 상품:
    - 참좋은훼밀리더블플러스:
      납기(만기) 형태
      예: 갱신형(10년갱신), 세만기(10년납)

    일반 상품:
    - 만기 또는 갱신주기 + (납기)
      예: 3년만기(3년납)
    """
    if "참좋은훼밀리더블플러스" in product_name:
        pay = _normalize_special_family_pay(pay_value)
        maturity = _normalize_special_family_maturity(maturity_value)

        if pay and maturity:
            return f"{pay}({maturity})"
        return pay or maturity

    maturity = _normalize_maturity(maturity_value)
    pay = _normalize_pay_period(pay_value)

    if maturity and pay:
        return f"{maturity}({pay})"
    return maturity or pay


def _is_blank_data_row(
    ws: Worksheet,
    merged: dict[tuple[int, int], Any],
    row: int,
    cols: list[int | None],
) -> bool:
    """주요 컬럼이 모두 공란이면 데이터 행이 아닌 것으로 본다."""
    for col in cols:
        if col and _text(_cell(ws, merged, row, col)):
            return False
    return True


# =============================================================================
# Public builder
# =============================================================================


def build_fire_db_conversion_rows(
    example: RateExample,
    workbook,
) -> list[RateExampleConversionRow]:
    """
    DB 손해보험 수정률 정규화 row 생성.

    반환된 row는 아직 DB에 저장되지 않은 unsaved model list다.
    저장/삭제 정책은 rate_example_normalizer.normalize_rate_example()이 담당한다.
    """
    rows: list[RateExampleConversionRow] = []

    for ws in workbook.worksheets:
        merged = _merged_value_map(ws)
        product_name = _text(ws.title)

        start_row = _find_table_start_row(ws, merged)
        if start_row is None:
            continue

        stop_row = _find_table_stop_row(ws, merged, start_row)
        header_row = _find_header_row(ws, merged, start_row, stop_row)
        if header_row is None:
            continue

        headers = _header_texts(ws, merged, header_row)

        plan_cols = _find_plan_cols(headers)
        car_col = _find_first_col(headers, ["차종"])
        pay_col = _find_first_col(headers, ["납기"])
        maturity_col = _find_first_col(headers, ["만기"])
        renewal_col = _find_first_col(headers, ["갱신주기"])
        rate_col = _find_rate_col(headers)

        # 필수 컬럼이 없으면 해당 시트는 안전하게 제외한다.
        # 단, DB 실손 상품은 '납기' 컬럼 없이 '만기 / 최초·갱신 / 수정률' 구조이므로
        # pay_col은 선택값으로 허용한다.
        if not rate_col:
            continue

        maturity_or_renewal_col = renewal_col or maturity_col
        if not maturity_or_renewal_col:
            continue

        for row_no in range(header_row + 1, stop_row):
            if _is_blank_data_row(
                ws,
                merged,
                row_no,
                [pay_col, maturity_or_renewal_col, rate_col, car_col, *plan_cols],
            ):
                continue

            plan_type = _build_plan_type(ws, merged, row_no, plan_cols, car_col)
            pay_period = _build_pay_period(
                product_name=product_name,
                maturity_value=_cell(ws, merged, row_no, maturity_or_renewal_col),
                pay_value=_cell(ws, merged, row_no, pay_col) if pay_col else "",
            )
            rate_cell = ws.cell(row=row_no, column=rate_col)
            mod_rate = _to_decimal_raw(
                _cell(ws, merged, row_no, rate_col),
                number_format=getattr(rate_cell, "number_format", ""),
            )

            if mod_rate is None:
                continue

            rows.append(
                RateExampleConversionRow(
                    source_file=example,
                    source_sheet=ws.title,
                    source_row_no=row_no,
                    insurer_type=RateExample.TYPE_FIRE,
                    category=RateExample.CAT_CONV,
                    insurer="DB",
                    coverage_type=_coverage_type(product_name, plan_type),
                    strategy_flag="",
                    product_name=product_name,
                    plan_type=plan_type,
                    pay_period=pay_period,
                    # 손해보험 수정률 단일 컬럼.
                    # raw 2.4는 2.4 그대로 저장한다.
                    year1=mod_rate,
                    year2=None,
                    year3=None,
                    year4=None,
                )
            )

    return rows
