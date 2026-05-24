# commission/services/rate_example/life/abl/parser.py
from __future__ import annotations

"""
ABL생명 환산율/수정률 정규화 모듈.

적용 대상:
- RateExample.insurer_type == life
- RateExample.category == conv
- RateExample.insurer == "ABL"
- xlsx 원본 파일

정규화 규칙:
1. 필수 시트 2개를 대상으로 정규화한다.
   - "주계약(저축성)"
   - "주계약(보장성)_12개월 선지급"
2. 저축성 시트:
   - 보종: 연금 고정
   - 상품명: A열
   - 구분: B열
   - 납기: C열
   - 1차년: D열
   - 2차년: E열
   - 3차년: F열
   - 4차년: 없음
3. 보장성 시트:
   - 보종: 상품명에 "종신" 포함 시 "종신/CI", 그 외 "기타(보장성)"
   - 상품명: A열
   - 구분: B열
   - 납기: E열
   - 1차년: F열
   - 2차년: G열
   - 3차년: H열
   - 4차년: I열
4. 병합셀/반복 생략 형태를 고려해 상품명·구분은 직전 값을 이어받는다.
"""

import logging
from decimal import Decimal, InvalidOperation
from typing import Iterable

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from commission.models import RateExample, RateExampleConversionRow
from commission.services.rate_example.common.text import clean_spaces

logger = logging.getLogger(__name__)

SHEET_ABL_SAVING = "주계약(저축성)"
SHEET_ABL_PROTECTION = "주계약(보장성)_12개월 선지급"


def _clean_text(value) -> str:
    """엑셀 셀 값을 화면 표시용 문자열로 정규화한다."""
    return clean_spaces(value)


def _to_decimal(value):
    """정수/실수/문자 퍼센트 값을 DecimalField 저장값으로 변환한다."""
    if value is None or value == "":
        return None

    if isinstance(value, Decimal):
        return value.quantize(Decimal("0.0001"))

    if isinstance(value, (int, float)):
        return Decimal(str(value)).quantize(Decimal("0.0001"))

    text = _clean_text(value).replace(",", "").replace("%", "")
    if not text:
        return None

    try:
        return Decimal(text).quantize(Decimal("0.0001"))
    except InvalidOperation:
        logger.warning("[rate_example][ABL] decimal parse skipped value=%r", value)
        return None


def _has_any_rate(*values) -> bool:
    """연차별 환산율 값이 하나라도 있으면 데이터 행으로 판단한다."""
    return any(_to_decimal(v) is not None for v in values)


def _coverage_type_for_protection(product_name: str) -> str:
    """보장성 시트의 상품명 기준 보종을 판정한다."""
    return "종신/CI" if "종신" in product_name else "기타(보장성)"


def _normalize_saving_sheet(
    example: RateExample,
    ws: Worksheet,
) -> Iterable[RateExampleConversionRow]:
    """
    ABL [주계약(저축성)] 시트 정규화.

    주요 기능:
    - A열 상품명, B열 구분이 비어 있으면 직전 값을 이어받는다.
    - 납기와 연차별 환산율이 모두 없는 행은 제외한다.
    """
    rows: list[RateExampleConversionRow] = []
    last_product = ""
    last_plan = ""

    for row_no in range(5, ws.max_row + 1):
        raw_product = _clean_text(ws.cell(row_no, 1).value)
        raw_plan = _clean_text(ws.cell(row_no, 2).value)

        product = raw_product or last_product
        plan = raw_plan or last_plan
        pay_period = _clean_text(ws.cell(row_no, 3).value)

        y1_raw = ws.cell(row_no, 4).value
        y2_raw = ws.cell(row_no, 5).value
        y3_raw = ws.cell(row_no, 6).value

        if raw_product:
            last_product = product
        if raw_plan:
            last_plan = plan

        if not product and not pay_period and not _has_any_rate(y1_raw, y2_raw, y3_raw):
            continue
        if not pay_period and not _has_any_rate(y1_raw, y2_raw, y3_raw):
            continue

        rows.append(
            RateExampleConversionRow(
                source_file=example,
                source_sheet=SHEET_ABL_SAVING,
                source_row_no=row_no,
                insurer_type=example.insurer_type,
                category=example.category,
                insurer="ABL",
                coverage_type="연금",
                strategy_flag="",
                product_name=product,
                plan_type=plan,
                pay_period=pay_period,
                year1=_to_decimal(y1_raw),
                year2=_to_decimal(y2_raw),
                year3=_to_decimal(y3_raw),
                year4=None,
            )
        )

    return rows


def _normalize_protection_sheet(
    example: RateExample,
    ws: Worksheet,
) -> Iterable[RateExampleConversionRow]:
    """
    ABL [주계약(보장성)_12개월 선지급] 시트 정규화.

    주요 기능:
    - A열 상품명, B열 구분이 비어 있으면 직전 값을 이어받는다.
    - 상품명에 "종신"이 포함되면 보종을 "종신/CI"로 저장한다.
    - 납기와 연차별 환산율이 모두 없는 행은 제외한다.
    """
    rows: list[RateExampleConversionRow] = []
    last_product = ""
    last_plan = ""

    for row_no in range(5, ws.max_row + 1):
        raw_product = _clean_text(ws.cell(row_no, 1).value)
        raw_plan = _clean_text(ws.cell(row_no, 2).value)

        product = raw_product or last_product
        plan = raw_plan or last_plan
        pay_period = _clean_text(ws.cell(row_no, 5).value)

        y1_raw = ws.cell(row_no, 6).value
        y2_raw = ws.cell(row_no, 7).value
        y3_raw = ws.cell(row_no, 8).value
        y4_raw = ws.cell(row_no, 9).value

        if raw_product:
            last_product = product
        if raw_plan:
            last_plan = plan

        if not product and not pay_period and not _has_any_rate(y1_raw, y2_raw, y3_raw, y4_raw):
            continue
        if not pay_period and not _has_any_rate(y1_raw, y2_raw, y3_raw, y4_raw):
            continue

        rows.append(
            RateExampleConversionRow(
                source_file=example,
                source_sheet=SHEET_ABL_PROTECTION,
                source_row_no=row_no,
                insurer_type=example.insurer_type,
                category=example.category,
                insurer="ABL",
                coverage_type=_coverage_type_for_protection(product),
                strategy_flag="",
                product_name=product,
                plan_type=plan,
                pay_period=pay_period,
                year1=_to_decimal(y1_raw),
                year2=_to_decimal(y2_raw),
                year3=_to_decimal(y3_raw),
                year4=_to_decimal(y4_raw),
            )
        )

    return rows


def build_life_abl_conversion_rows(
    example: RateExample,
    wb: Workbook,
) -> list[RateExampleConversionRow]:
    """
    ABL생명 xlsx 전체 workbook을 정규화 row 목록으로 변환한다.

    주요 기능:
    - 필수 시트 존재 여부를 먼저 검증한다.
    - 시트가 없으면 예외를 발생시켜 상위 transaction rollback을 유도한다.
    - DB 저장/삭제는 상위 normalizer가 일괄 처리한다.
    """
    missing = [
        sheet_name
        for sheet_name in (SHEET_ABL_SAVING, SHEET_ABL_PROTECTION)
        if sheet_name not in wb.sheetnames
    ]
    if missing:
        raise ValueError(f"ABL 환산율/수정률 필수 시트가 없습니다: {', '.join(missing)}")

    normalized_rows: list[RateExampleConversionRow] = []
    normalized_rows.extend(_normalize_saving_sheet(example, wb[SHEET_ABL_SAVING]))
    normalized_rows.extend(_normalize_protection_sheet(example, wb[SHEET_ABL_PROTECTION]))

    return normalized_rows
