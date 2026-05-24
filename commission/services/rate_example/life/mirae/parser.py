# commission/services/rate_example/life/mirae/parser.py
from __future__ import annotations

"""
미래에셋생명 환산율/수정률 정규화 parser.

역할:
- 미래에셋 raw xlsx 파일의 "보장성", "보장성_*", "저축성" 시트를
  RateExampleConversionRow 표준 구조로 변환한다.

정규화 정책:
- 원본 파일 저장/검증/DB 적재는 rate_example_normalizer.py가 담당한다.
- 본 파일은 workbook -> RateExampleConversionRow list 변환만 담당한다.
- 환산율/유지성적 값은 raw 숫자 기준으로 Decimal("0.0001") 형태로 저장한다.
- Excel 셀이 백분율 서식("%")인 경우 openpyxl 실제값 보정을 위해 x100 처리한다.
- 병합 셀은 병합 전 점유 범위의 값을 행 단위로 전파한다.
- 줄바꿈 상품명은 각 줄을 strip 후 붙여 한 줄로 정규화한다.
"""

from decimal import Decimal
import re
from typing import Any

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from commission.models import RateExample, RateExampleConversionRow
from commission.services.rate_example.common.decimal import (
    decimal_percent_value,
)
from commission.services.rate_example.common.rows import append_unique


DEC4 = Decimal("0.0001")
NO_PLAN_TYPE = "사용안함"


def _clean_text(value: Any) -> str:
    """
    셀 값을 화면/DB 저장용 문자열로 정규화한다.

    미래에셋 raw 파일은 상품명이 여러 줄로 들어오는 경우가 있으므로
    줄별 trim 후 첫 줄 끝에 그대로 붙여 한 줄로 만든다.
    """
    if value is None:
        return ""

    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    parts = [part.strip() for part in text.split("\n") if part.strip()]
    text = "".join(parts) if parts else text.strip()
    return re.sub(r"\s+", " ", text).strip()


def _merged_value(ws: Worksheet, row: int, col: int) -> Any:
    """
    병합 셀의 실제 값을 반환한다.

    openpyxl은 병합 범위의 좌상단 셀에만 값을 보관하므로,
    현재 좌표가 병합 범위에 포함되면 좌상단 값을 가져온다.
    """
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value

    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return ws.cell(
                row=merged_range.min_row,
                column=merged_range.min_col,
            ).value

    return None


def _cell_text(ws: Worksheet, row: int, col: int) -> str:
    """병합 셀 전파 + 줄바꿈 제거를 함께 적용한 문자열을 반환한다."""
    return _clean_text(_merged_value(ws, row, col))


def _to_decimal(value: Any, *, number_format: str = "") -> Decimal | None:
    """
    숫자 셀을 Decimal("0.0001")로 변환한다.

    주의:
    - 프로젝트 환산율 정책은 100%를 Decimal("100")처럼 백분율 수치로 저장한다.
    - 다만 미래에셋 raw의 환산성적이 1, 1.5처럼 일반 숫자로 제공되는 경우는
      raw 숫자 그대로 저장한다.
    - Excel 서식 자체가 %인 경우에만 openpyxl 실제값 보정을 위해 x100 한다.
    """
    dec = decimal_percent_value(value, number_format=number_format)
    if dec is None:
        return None

    return dec.quantize(DEC4)


def _cell_decimal(ws: Worksheet, row: int, col: int) -> Decimal | None:
    """병합 셀 전파 후 해당 셀의 숫자값을 Decimal로 변환한다."""
    value = _merged_value(ws, row, col)
    number_format = ws.cell(row=row, column=col).number_format
    return _to_decimal(value, number_format=number_format)


def _find_col(ws: Worksheet, labels: set[str], *, max_scan_row: int = 12) -> int | None:
    """
    상단 헤더 영역에서 특정 헤더명을 가진 컬럼을 찾는다.

    미래에셋 raw 파일은 시트별로 보종구분 컬럼 유무가 달라
    고정 열만 사용하면 일부 시트가 어긋난다.
    따라서 헤더명 탐지를 1순위로 사용하고, parser별 fallback 열을 2순위로 사용한다.
    """
    for row in range(1, min(ws.max_row, max_scan_row) + 1):
        for col in range(1, ws.max_column + 1):
            text = _cell_text(ws, row, col)
            if text in labels:
                return col
    return None


def _coverage_for_protection(product_name: str) -> str:
    """보장성 시트의 보종을 상품명 키워드 기준으로 판정한다."""
    if "경영" in product_name:
        return "CEO정기"
    if "종신" in product_name:
        return "종신/CI"
    return "기타(보장성)"


def _coverage_for_saving(product_name: str) -> str:
    """저축성 시트의 보종을 상품명 키워드 기준으로 판정한다."""
    if "적립" in product_name:
        return "VUL"
    if "변액" in product_name:
        return "변액연금"
    return "연금"


def _make_row(
    *,
    example: RateExample,
    ws: Worksheet,
    source_row_no: int,
    coverage_type: str,
    product_name: str,
    plan_type: str,
    pay_period: str,
    year1: Decimal | None,
    year2: Decimal | None,
    year3: Decimal | None,
    year4: Decimal | None,
) -> RateExampleConversionRow:
    """RateExampleConversionRow 생성을 중앙화해 필드 누락을 방지한다."""
    # 계산 화면 옵션 체인은 plan_type 선택지를 거쳐 pay_period를 조회한다.
    # plan_type=""로 저장하면 options API가 선택지를 반환하지 않아
    # "보험사, 상품명, 납기" validation 단계까지 진행하지 못한다.
    #
    # 따라서 미래에셋처럼 raw에 별도 구분이 없는 상품은
    # 카디프와 동일하게 "사용안함" sentinel 값으로 저장한다.
    plan_type = _clean_text(plan_type) or NO_PLAN_TYPE

    return RateExampleConversionRow(
        source_file=example,
        source_sheet=ws.title,
        source_row_no=source_row_no,
        insurer_type=example.insurer_type,
        category=example.category,
        insurer=example.insurer,
        coverage_type=coverage_type,
        strategy_flag="",
        product_name=product_name,
        plan_type=plan_type,
        pay_period=pay_period,
        year1=year1,
        year2=year2,
        year3=year3,
        year4=year4,
    )


def _append_unique(
    rows: list[RateExampleConversionRow],
    seen: set[tuple],
    row: RateExampleConversionRow,
) -> None:
    """
    동일 파일 내 완전 중복 row를 방지한다.

    append 업로드 정책은 DB 기존 row 보존 정책이므로 여기서는 파일 내부 중복만 방어한다.
    """
    key = (
        row.source_sheet,
        row.coverage_type,
        row.product_name,
        row.plan_type,
        row.pay_period,
        row.year1,
        row.year2,
        row.year3,
        row.year4,
    )
    append_unique(rows, seen, row, key)


def _parse_base_protection_sheet(
    example: RateExample,
    ws: Worksheet,
) -> list[RateExampleConversionRow]:
    """
    "보장성" 단독 시트 정규화.

    매핑:
    - 상품명: "상품명" 헤더 컬럼, fallback C열
    - 구분: "보종구분" 헤더 컬럼
    - 납기: "납입기간" 헤더 컬럼
    - 환산율: "환산성적" 헤더 컬럼
    - 1~4차년: 환산성적 동일 반영
    """
    rows: list[RateExampleConversionRow] = []
    seen: set[tuple] = set()

    product_col = _find_col(ws, {"상품명"}) or 3
    plan_col = _find_col(ws, {"보종구분"}) or 6
    pay_col = _find_col(ws, {"납입기간"}) or 7
    rate_col = _find_col(ws, {"환산성적"}) or 8

    current_product = ""

    for row_no in range(1, ws.max_row + 1):
        raw_product = _cell_text(ws, row_no, product_col)
        if raw_product and raw_product != "상품명":
            current_product = raw_product

        product_name = current_product
        if not product_name or product_name == "상품명":
            continue

        pay_period = _cell_text(ws, row_no, pay_col)
        if not pay_period or pay_period == "납입기간":
            continue

        rate = _cell_decimal(ws, row_no, rate_col)
        if rate is None:
            continue

        plan_type = _cell_text(ws, row_no, plan_col)
        if plan_type == "보종구분":
            plan_type = NO_PLAN_TYPE

        normalized = _make_row(
            example=example,
            ws=ws,
            source_row_no=row_no,
            coverage_type=_coverage_for_protection(product_name),
            product_name=product_name,
            plan_type=plan_type,
            pay_period=pay_period,
            year1=rate,
            year2=rate,
            year3=rate,
            year4=rate,
        )
        _append_unique(rows, seen, normalized)

    return rows


def _parse_named_protection_sheet(
    example: RateExample,
    ws: Worksheet,
) -> list[RateExampleConversionRow]:
    """
    "보장성_*" 시트 정규화.

    정책:
    - "보장성" 단독 시트는 이 함수 대상이 아니다.
    - raw의 "보장성_특약" 시트는 주계약 상품 시트가 아니므로 제외한다.
    - A열에서 "특약" 발견 시 해당 행 포함 이하를 제외한다.
    - 상품명: "상품명" 헤더 컬럼, fallback B열
    - 납기/환산성적: 헤더명 우선 탐지, 실패 시 요청 명세의 D/E fallback 사용
    - 보종/구분은 공란 유지
    """
    rows: list[RateExampleConversionRow] = []
    seen: set[tuple] = set()

    if "특약" in ws.title:
        return rows

    product_col = _find_col(ws, {"상품명"}) or 2
    pay_col = _find_col(ws, {"납입기간"}) or 4
    rate_col = _find_col(ws, {"환산성적"}) or 5

    current_product = ""

    for row_no in range(1, ws.max_row + 1):
        a_text = _cell_text(ws, row_no, 1)
        if "특약" in a_text:
            break

        raw_product = _cell_text(ws, row_no, product_col)
        if raw_product and raw_product != "상품명":
            current_product = raw_product

        product_name = current_product
        if not product_name or product_name == "상품명":
            continue

        pay_period = _cell_text(ws, row_no, pay_col)
        if not pay_period or pay_period == "납입기간":
            continue

        rate = _cell_decimal(ws, row_no, rate_col)
        if rate is None:
            continue

        normalized = _make_row(
            example=example,
            ws=ws,
            source_row_no=row_no,
            coverage_type="",
            product_name=product_name,
            plan_type=NO_PLAN_TYPE,
            pay_period=pay_period,
            year1=rate,
            year2=rate,
            year3=rate,
            year4=rate,
        )
        _append_unique(rows, seen, normalized)

    return rows


def _parse_saving_sheet(
    example: RateExample,
    ws: Worksheet,
) -> list[RateExampleConversionRow]:
    """
    "저축성" 시트 정규화.

    매핑:
    - 상품명: "상품명" 헤더 컬럼, fallback C열
    - 납기: "납입기간" 헤더 컬럼, fallback E열
    - 1차년/2차년: 환산성적
    - 3차년/4차년: 유지성적
    - 구분: "-" 고정
    """
    rows: list[RateExampleConversionRow] = []
    seen: set[tuple] = set()

    product_col = _find_col(ws, {"상품명"}) or 3
    pay_col = _find_col(ws, {"납입기간"}) or 5
    conversion_col = _find_col(ws, {"환산성적"}) or 6
    persist_col = _find_col(ws, {"유지성적"}) or 7

    current_product = ""

    for row_no in range(1, ws.max_row + 1):
        raw_product = _cell_text(ws, row_no, product_col)
        if raw_product and raw_product != "상품명":
            current_product = raw_product

        product_name = current_product
        if not product_name or product_name == "상품명":
            continue

        pay_period = _cell_text(ws, row_no, pay_col)
        if not pay_period or pay_period == "납입기간":
            continue

        conversion_rate = _cell_decimal(ws, row_no, conversion_col)
        persist_rate = _cell_decimal(ws, row_no, persist_col)

        if conversion_rate is None and persist_rate is None:
            continue

        normalized = _make_row(
            example=example,
            ws=ws,
            source_row_no=row_no,
            coverage_type=_coverage_for_saving(product_name),
            product_name=product_name,
            plan_type=NO_PLAN_TYPE,
            pay_period=pay_period,
            year1=conversion_rate,
            year2=conversion_rate,
            year3=persist_rate,
            year4=persist_rate,
        )
        _append_unique(rows, seen, normalized)

    return rows


def build_life_mirae_conversion_rows(
    example: RateExample,
    workbook: Workbook,
) -> list[RateExampleConversionRow]:
    """
    미래에셋생명 환산율/수정률 정규화 진입점.

    처리 대상:
    - "보장성"
    - "보장성_"으로 시작하는 시트
    - "저축성"

    반환:
    - 아직 DB에 저장하지 않은 RateExampleConversionRow 인스턴스 list
    """
    rows: list[RateExampleConversionRow] = []

    for ws in workbook.worksheets:
        sheet_name = str(ws.title or "").strip()

        if sheet_name == "보장성":
            rows.extend(_parse_base_protection_sheet(example, ws))
            continue

        if sheet_name.startswith("보장성_"):
            rows.extend(_parse_named_protection_sheet(example, ws))
            continue

        if sheet_name == "저축성":
            rows.extend(_parse_saving_sheet(example, ws))
            continue

    return rows
