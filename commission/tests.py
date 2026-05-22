# django_ma/commission/tests.py
from __future__ import annotations

"""
commission app 공통 회귀 방지 테스트.

목적:
- 기능 변화 0 리팩토링 시 핵심 helper 정책이 깨지지 않도록 보호한다.
- upload handler / rate_example normalizer 공통 helper 계약을 고정한다.
- upload_utils / fail token / deposit service / download response 회귀를 방지한다.
"""

from decimal import Decimal

from django.core.cache import cache
from django.test import RequestFactory, SimpleTestCase, TestCase
from openpyxl import Workbook

from accounts.models import CustomUser
from commission.models import DepositOther, DepositSummary, DepositSurety
from commission.services.deposit import calc_filtered_totals, calc_keep_totals_all
from commission.services.deposit_serializers import (
    apply_deposit_summary_totals,
    json_user_detail,
    other_to_payload,
    summary_to_payload,
    surety_to_payload,
    user_to_payload,
)
from commission.services.rate_example_normalizers._common.decimal import (
    decimal_percent_value,
)
from commission.services.rate_example_normalizers._common.excel import (
    build_merged_value_map,
    build_worksheet_value_map,
    cell_value_with_merged,
    filled_value_above,
)
from commission.services.rate_example_normalizers._common.rows import append_unique
from commission.services.rate_example_normalizers._common.text import (
    clean_spaces,
    clean_text,
    is_empty_like,
)
from commission.services.rate_example_normalizers._common.pdf import (
    clean_pdf_text,
    decimal_from_pdf_percent,
    dedupe_by_key,
    group_pdf_items_by_y,
    PdfTextItem,
)
from commission.upload_handlers._common import safe_cell_text, upload_result
from commission.upload_utils import (
    EMPTY_LIKE_VALUES,
    _extract_emp7_from_a,
    _is_empty_like,
    _norm_emp_id,
)
from commission.views._excel_export import XLSX_MIME, rows_to_excel_response
from commission.views.downloads import _can_download_fail_payload
from commission.views.utils_fail_excel import FAIL_TTL_SECONDS, store_fail_rows_as_excel


# =============================================================================
# RateExample Decimal helper
# =============================================================================
class RateExampleCommonDecimalTests(SimpleTestCase):
    def test_decimal_percent_value_scales_excel_percent_number_format(self):
        self.assertEqual(
            decimal_percent_value(Decimal("0.7"), number_format="0%"),
            Decimal("70.0"),
        )

    def test_decimal_percent_value_keeps_plain_number(self):
        self.assertEqual(decimal_percent_value("160"), Decimal("160"))

    def test_decimal_percent_value_keeps_percent_text_by_default(self):
        self.assertEqual(decimal_percent_value("0.8%"), Decimal("0.8"))

    def test_decimal_percent_value_can_scale_small_percent_text(self):
        self.assertEqual(
            decimal_percent_value("0.8%", scale_small_percent_text=True),
            Decimal("80.0"),
        )

    def test_decimal_percent_value_returns_none_for_empty(self):
        self.assertIsNone(decimal_percent_value(""))
        self.assertIsNone(decimal_percent_value(None))
        self.assertIsNone(decimal_percent_value("-"))

    def test_decimal_percent_value_keeps_already_scaled_percent(self):
        self.assertEqual(decimal_percent_value("80%"), Decimal("80"))


class RateExampleCommonTextTests(SimpleTestCase):
    def test_clean_text_preserves_line_policy(self):
        self.assertEqual(clean_text(" A\nB "), "A\nB")

    def test_clean_spaces_compacts_whitespace(self):
        self.assertEqual(clean_spaces(" A\n  B\tC "), "A B C")

    def test_is_empty_like(self):
        self.assertTrue(is_empty_like(None))
        self.assertTrue(is_empty_like("nan"))
        self.assertTrue(is_empty_like("-"))
        self.assertFalse(is_empty_like("0"))


class RateExampleCommonRowsTests(SimpleTestCase):
    def test_append_unique_dedupes_only_same_key(self):
        rows = []
        seen = set()

        append_unique(rows, seen, "A", ("p1", "10년"))
        append_unique(rows, seen, "A-duplicate", ("p1", "10년"))
        append_unique(rows, seen, "B", ("p2", "10년"))

        self.assertEqual(rows, ["A", "B"])


class RateExampleCommonPdfTests(SimpleTestCase):
    def test_clean_pdf_text_compacts_nbsp_and_whitespace(self):
        self.assertEqual(clean_pdf_text(" A\u00a0\n  B\tC "), "A B C")

    def test_decimal_from_pdf_percent_extracts_numeric_part(self):
        self.assertEqual(decimal_from_pdf_percent("160% 주)"), Decimal("160"))
        self.assertEqual(decimal_from_pdf_percent("1,234.5%"), Decimal("1234.5"))
        self.assertIsNone(decimal_from_pdf_percent("수정률"))

    def test_dedupe_by_key_preserves_order(self):
        items = [
            {"product": "A", "rate": "100"},
            {"product": "A", "rate": "100"},
            {"product": "B", "rate": "100"},
        ]

        result = dedupe_by_key(
            items,
            lambda item: (item["product"], item["rate"]),
        )

        self.assertEqual(
            result,
            [
                {"product": "A", "rate": "100"},
                {"product": "B", "rate": "100"},
            ],
        )

    def test_group_pdf_items_by_y_groups_rows_and_sorts_by_x(self):
        items = [
            PdfTextItem("B", x0=30, y0=10, x1=40, y1=20),
            PdfTextItem("A", x0=10, y0=11, x1=20, y1=20),
            PdfTextItem("C", x0=10, y0=30, x1=20, y1=40),
        ]

        rows = group_pdf_items_by_y(items, y_tolerance=3)

        self.assertEqual([[item.text for item in row] for row in rows], [["A", "B"], ["C"]])
    
    def test_group_pdf_items_by_y_respects_tolerance_boundary(self):
        items = [
            PdfTextItem("A", x0=10, y0=10, x1=20, y1=20),
            PdfTextItem("B", x0=20, y0=13, x1=30, y1=20),
            PdfTextItem("C", x0=10, y0=14.1, x1=20, y1=20),
        ]

        rows = group_pdf_items_by_y(items, y_tolerance=3)

        self.assertEqual([[item.text for item in row] for row in rows], [["A", "B"], ["C"]])

    def test_decimal_from_pdf_percent_handles_negative_and_plain_decimal(self):
        self.assertEqual(decimal_from_pdf_percent("-12.5%"), Decimal("-12.5"))
        self.assertEqual(decimal_from_pdf_percent("수정률 240"), Decimal("240"))

    def test_dedupe_by_key_supports_model_like_objects(self):
        class Row:
            def __init__(self, product_name, plan_type, pay_period):
                self.product_name = product_name
                self.plan_type = plan_type
                self.pay_period = pay_period

        rows = [
            Row("A", "보장", "10년납"),
            Row("A", "보장", "10년납"),
            Row("A", "적립", "10년납"),
        ]

        result = dedupe_by_key(
            rows,
            lambda row: (row.product_name, row.plan_type, row.pay_period),
        )

        self.assertEqual(len(result), 2)
        self.assertEqual(result[0].plan_type, "보장")
        self.assertEqual(result[1].plan_type, "적립")


# =============================================================================
# Upload handler common helper
# =============================================================================
class UploadHandlerCommonTests(SimpleTestCase):
    def test_safe_cell_text_normalizes_empty_like_values(self):
        self.assertEqual(safe_cell_text(None), "")
        self.assertEqual(safe_cell_text("nan"), "")
        self.assertEqual(safe_cell_text("none"), "")
        self.assertEqual(safe_cell_text("-"), "")

    def test_safe_cell_text_strips_normal_values(self):
        self.assertEqual(safe_cell_text("  지급  "), "지급")
        self.assertEqual(safe_cell_text(1234567), "1234567")

    def test_upload_result_contract(self):
        self.assertEqual(
            upload_result(
                inserted_or_updated=3,
                missing_users=2,
                missing_sample=["1000001", "1000002"],
            ),
            {
                "inserted_or_updated": 3,
                "missing_users": 2,
                "missing_sample": ["1000001", "1000002"],
            },
        )

    def test_upload_result_default_contract(self):
        self.assertEqual(
            upload_result(),
            {
                "inserted_or_updated": 0,
                "missing_users": 0,
                "missing_sample": [],
            },
        )

    def test_upload_result_always_returns_list_for_missing_sample(self):
        result = upload_result(inserted_or_updated=1, missing_users=0)
        self.assertIsInstance(result["missing_sample"], list)
        self.assertEqual(result["missing_sample"], [])

    def test_upload_result_preserves_zero_values(self):
        self.assertEqual(
            upload_result(
                inserted_or_updated=0,
                missing_users=0,
                missing_sample=[],
            ),
            {
                "inserted_or_updated": 0,
                "missing_users": 0,
                "missing_sample": [],
            },
        )


# =============================================================================
# Upload utils helper
# =============================================================================
class UploadUtilsConvertTests(SimpleTestCase):
    def test_empty_like_values_are_exported_as_frozenset(self):
        self.assertIsInstance(EMPTY_LIKE_VALUES, frozenset)
        self.assertIn("", EMPTY_LIKE_VALUES)
        self.assertIn("nan", EMPTY_LIKE_VALUES)
        self.assertIn("none", EMPTY_LIKE_VALUES)
        self.assertIn("-", EMPTY_LIKE_VALUES)

    def test_is_empty_like_normalizes_common_empty_values(self):
        self.assertTrue(_is_empty_like(None))
        self.assertTrue(_is_empty_like(""))
        self.assertTrue(_is_empty_like("  "))
        self.assertTrue(_is_empty_like("nan"))
        self.assertTrue(_is_empty_like("None"))
        self.assertTrue(_is_empty_like("-"))
        self.assertFalse(_is_empty_like("0"))
        self.assertFalse(_is_empty_like("1234567"))

    def test_norm_emp_id_keeps_string_pk_policy(self):
        self.assertEqual(_norm_emp_id("1234567.0"), "1234567")
        self.assertEqual(_norm_emp_id(" 1234567 "), "1234567")
        self.assertEqual(_norm_emp_id(1234567), "1234567")
        self.assertEqual(_norm_emp_id(None), "")
        self.assertEqual(_norm_emp_id("nan"), "")
        self.assertEqual(_norm_emp_id("-"), "")

    def test_extract_emp7_from_a_uses_existing_slice_policy(self):
        self.assertEqual(_extract_emp7_from_a("X1234567Z"), "1234567")
        self.assertEqual(_extract_emp7_from_a("ABC7654321X"), "7654321")
        self.assertEqual(_extract_emp7_from_a("1234567"), "")
        self.assertEqual(_extract_emp7_from_a("ABC12X456Z"), "")
        self.assertEqual(_extract_emp7_from_a(None), "")


# =============================================================================
# RateExample Excel helper
# =============================================================================
class RateExampleCommonExcelTests(SimpleTestCase):
    def test_build_merged_value_map_propagates_top_left_value(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "상품명"
        ws.merge_cells("A1:A3")

        merged_map = build_merged_value_map(ws)

        self.assertEqual(merged_map[(1, 1)], "상품명")
        self.assertEqual(merged_map[(2, 1)], "상품명")
        self.assertEqual(merged_map[(3, 1)], "상품명")

    def test_cell_value_with_merged_uses_real_value_first(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "병합값"
        ws.merge_cells("A1:A3")
        ws["B2"] = "실제값"

        merged_map = build_merged_value_map(ws)

        self.assertEqual(cell_value_with_merged(ws, merged_map, 2, 2), "실제값")

    def test_cell_value_with_merged_falls_back_to_merged_map(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "전파값"
        ws.merge_cells("A1:A3")

        merged_map = build_merged_value_map(ws)

        self.assertEqual(cell_value_with_merged(ws, merged_map, 3, 1), "전파값")

    def test_cell_value_with_merged_returns_none_for_plain_empty_cell(self):
        wb = Workbook()
        ws = wb.active

        merged_map = build_merged_value_map(ws)

        self.assertIsNone(cell_value_with_merged(ws, merged_map, 5, 5))

    def test_build_worksheet_value_map_keeps_plain_cells_and_expands_merged_cells(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "병합값"
        ws["B2"] = "일반값"
        ws.merge_cells("A1:A3")

        values = build_worksheet_value_map(ws)

        self.assertEqual(values[(1, 1)], "병합값")
        self.assertEqual(values[(2, 1)], "병합값")
        self.assertEqual(values[(3, 1)], "병합값")
        self.assertEqual(values[(2, 2)], "일반값")

    def test_build_worksheet_value_map_can_skip_plain_empty_cells(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "값"

        values = build_worksheet_value_map(ws, include_empty=False)

        self.assertEqual(values[(1, 1)], "값")
        self.assertNotIn((5, 5), values)

    def test_filled_value_above_uses_nearest_upper_value(self):
        values = {
            (1, 2): "헤더",
            (2, 2): "상단값",
            (3, 2): "",
            (4, 2): None,
        }

        result = filled_value_above(
            values,
            header_row=1,
            row_no=4,
            col_no=2,
            is_filled=lambda value: bool(str(value or "").strip()),
        )

        self.assertEqual(result, "상단값")


# =============================================================================
# Fail token / Excel response helper
# =============================================================================
class CommissionExcelHelperTests(TestCase):
    def setUp(self):
        cache.clear()

    def test_store_fail_rows_as_excel_returns_cache_token(self):
        token = store_fail_rows_as_excel(
            rows=[
                {"user_id": "1000001", "reason": "사용자 미존재"},
                {"user_id": "1000002", "reason": "스코프 제외"},
            ],
            filename="upload_fail_test.xlsx",
            owner_id="admin01",
        )

        self.assertTrue(token)

        payload = cache.get(f"commission:upload_fail:{token}")
        self.assertIsNotNone(payload)
        self.assertEqual(payload["filename"], "upload_fail_test.xlsx")
        self.assertEqual(payload["owner_id"], "admin01")
        self.assertIsInstance(payload["content"], bytes)
        self.assertGreater(len(payload["content"]), 0)
        self.assertEqual(FAIL_TTL_SECONDS, 60 * 60)

    def test_store_fail_rows_as_excel_returns_empty_for_no_rows(self):
        token = store_fail_rows_as_excel(
            rows=[],
            filename="empty.xlsx",
            owner_id="admin01",
        )
        self.assertEqual(token, "")

    def test_rows_to_excel_response_smoke(self):
        response = rows_to_excel_response(
            rows=[{"user_id": "1000001", "amount": 1234}],
            sheet_name="smoke",
            filename="테스트.xlsx",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], XLSX_MIME)
        self.assertIn("attachment;", response["Content-Disposition"])
        self.assertIn("filename*=", response["Content-Disposition"])
        self.assertGreater(len(response.content), 0)


# =============================================================================
# Deposit service
# =============================================================================
class DepositServiceAggregateTests(TestCase):
    def setUp(self):
        self.user = CustomUser.objects.create_user(
            id="1000001",
            name="테스트",
            password="pw",
            grade="basic",
        )

    def test_calc_filtered_totals(self):
        DepositSurety.objects.create(
            user=self.user,
            product_name="GA개인 보증",
            amount=1000,
            status="유지",
        )
        DepositSurety.objects.create(
            user=self.user,
            product_name="일반 보증",
            amount=2000,
            status="유지",
        )
        DepositOther.objects.create(
            user=self.user,
            product_name="기타",
            product_type="수수료 채권",
            amount=3000,
            status="유지인",
        )
        DepositOther.objects.create(
            user=self.user,
            product_name="기타",
            product_type="일반 채권",
            amount=4000,
            status="유지",
        )

        surety_total, other_total = calc_filtered_totals(self.user.pk)

        self.assertEqual(surety_total, 1000)
        self.assertEqual(other_total, 3000)

    def test_calc_keep_totals_all(self):
        DepositSurety.objects.create(
            user=self.user,
            product_name="GA개인 보증",
            amount=1000,
            status="유지",
        )
        DepositSurety.objects.create(
            user=self.user,
            product_name="GA개인 보증",
            amount=2000,
            status="해지",
        )
        DepositOther.objects.create(
            user=self.user,
            product_name="기타",
            product_type="수수료 채권",
            amount=3000,
            status="유지",
        )
        DepositOther.objects.create(
            user=self.user,
            product_name="기타",
            product_type="수수료 채권",
            amount=4000,
            status="유지인",
        )

        surety_keep_all, other_keep_all = calc_keep_totals_all(self.user.pk)

        self.assertEqual(surety_keep_all, 1000)
        self.assertEqual(other_keep_all, 3000)


# =============================================================================
# Deposit serializers
# =============================================================================
class DepositSerializerTests(TestCase):
    def setUp(self):
        self.user = CustomUser.objects.create_user(
            id="1000001",
            name="테스트",
            password="pw",
            grade="basic",
            part="MA",
            branch="서울",
        )

    def test_user_to_payload_keeps_deposit_home_contract(self):
        payload = user_to_payload(self.user)

        self.assertEqual(payload["id"], "1000001")
        self.assertEqual(payload["name"], "테스트")
        self.assertEqual(payload["part"], "MA")
        self.assertEqual(payload["branch"], "서울")
        self.assertIn("join_date_display", payload)
        self.assertIn("retire_date_display", payload)
        self.assertIn("enter", payload)
        self.assertIn("quit", payload)

    def test_json_user_detail_keeps_legacy_data_and_user_keys(self):
        payload = {"id": "1000001", "name": "테스트"}
        response = json_user_detail(payload)

        self.assertEqual(response.status_code, 200)
        self.assertJSONEqual(
            response.content.decode("utf-8"),
            {"ok": True, "data": payload, "user": payload},
        )

    def test_summary_to_payload_and_totals_contract(self):
        summary = DepositSummary.objects.create(
            user=self.user,
            final_payment=1000,
            sales_total=2000,
            surety_total=9999,
            other_total=8888,
        )
        DepositSurety.objects.create(
            user=self.user,
            product_name="GA개인 보증",
            amount=3000,
            status="유지",
        )
        DepositOther.objects.create(
            user=self.user,
            product_name="기타",
            product_type="수수료 채권",
            amount=4000,
            status="유지인",
        )

        payload = summary_to_payload(summary)
        payload = apply_deposit_summary_totals(payload, self.user.pk)

        self.assertEqual(payload["final_payment"], 1000)
        self.assertEqual(payload["sales_total"], 2000)
        self.assertEqual(payload["surety_total_all"], 9999)
        self.assertEqual(payload["other_total_all"], 8888)
        self.assertEqual(payload["surety_total"], 3000)
        self.assertEqual(payload["other_total"], 4000)
        self.assertIn("debt_keep_total", payload)

    def test_detail_serializers_keep_row_keys(self):
        surety = DepositSurety.objects.create(
            user=self.user,
            product_name="GA개인 보증",
            policy_no="P-1",
            amount=3000,
            status="유지",
        )
        other = DepositOther.objects.create(
            user=self.user,
            product_name="기타",
            product_type="수수료 채권",
            amount=4000,
            bond_no="B-1",
            status="유지",
        )

        surety_payload = surety_to_payload(surety)
        other_payload = other_to_payload(other)

        self.assertEqual(surety_payload["product_name"], "GA개인 보증")
        self.assertEqual(surety_payload["policy_no"], "P-1")
        self.assertEqual(surety_payload["amount"], 3000)
        self.assertEqual(other_payload["product_type"], "수수료 채권")
        self.assertEqual(other_payload["bond_no"], "B-1")
        self.assertEqual(other_payload["amount"], 4000)


# =============================================================================
# Fail download permission helper
# =============================================================================
class FailDownloadPermissionTests(TestCase):
    def setUp(self):
        self.factory = RequestFactory()
        self.superuser = CustomUser.objects.create_user(
            id="9000001",
            name="관리자",
            password="pw",
            grade="superuser",
            is_superuser=True,
        )
        self.other_superuser = CustomUser.objects.create_user(
            id="9000002",
            name="타관리자",
            password="pw",
            grade="superuser",
            is_superuser=True,
        )
        self.basic = CustomUser.objects.create_user(
            id="1000001",
            name="일반",
            password="pw",
            grade="basic",
        )

    def _request_for(self, user):
        request = self.factory.get("/commission/download/upload-fail/")
        request.user = user
        return request

    def test_owner_superuser_can_download_new_token(self):
        request = self._request_for(self.superuser)

        self.assertTrue(
            _can_download_fail_payload(
                request,
                {"owner_id": str(self.superuser.pk)},
            )
        )

    def test_other_superuser_cannot_download_owner_bound_token(self):
        request = self._request_for(self.other_superuser)

        self.assertFalse(
            _can_download_fail_payload(
                request,
                {"owner_id": str(self.superuser.pk)},
            )
        )

    def test_legacy_token_without_owner_allows_superuser_only(self):
        self.assertTrue(
            _can_download_fail_payload(
                self._request_for(self.superuser),
                {"owner_id": ""},
            )
        )
        self.assertFalse(
            _can_download_fail_payload(
                self._request_for(self.basic),
                {"owner_id": ""},
            )
        )