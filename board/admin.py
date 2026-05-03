# board/admin.py
#
# 변경 내용 (기존 코드 무변경):
#   - 기존 PostAdmin / IndustryArticleAdmin 등 전부 그대로 유지
#   - 하단에 WorkCategory / WorkTask / WorkTaskAttachment Admin 3개 추가
#
# WorkTask Admin 예외 정책 (worktask.md §17):
#   Django Admin 에서는 superuser(is_staff)가 모든 WorkTask를 열람·수정할 수 있다.
#   일반 뷰(/board/worktasks/)의 owner 격리 정책과는 완전히 별개이며,
#   운영 목적의 명시적 예외다.

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from django.contrib import admin
from django.http import HttpResponse
from django.urls import path
from django.utils.html import format_html
from django.utils.timezone import localtime

from accounts.custom_admin import custom_admin_site
from .industry_models import (
    IndustryArticle,
    IndustryCollectJobLog,
    IndustryRecommendation,
    IndustryUserPreference,
)
from .models import Post, KrHoliday, WorkCategory, WorkTask, WorkTaskAttachment


# =============================================================================
# KR Holiday Admin
# - API 수집 데이터 확인
# - 임시공휴일/대체공휴일 수동 보정
# =============================================================================
@admin.register(KrHoliday)
class KrHolidayAdmin(admin.ModelAdmin):
    list_display = (
        "date",
        "name",
        "is_holiday",
        "is_temporary",
        "source",
        "fetched_at",
        "updated_at",
    )
    list_filter = ("is_holiday", "is_temporary", "source")
    search_fields = ("name", "source_event_id")
    date_hierarchy = "date"
    ordering = ("-date",)
    readonly_fields = ("raw_payload", "fetched_at", "created_at", "updated_at")


# =========================================================
# Excel Export (Posts)
# =========================================================
def export_posts_as_excel(queryset, filename: str = "posts_export.xlsx") -> HttpResponse:
    """
    Post queryset을 엑셀로 내보내는 공용 함수
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Posts"

    headers = [
        "접수번호",
        "제목",
        "사번(요청자)",
        "성명(요청자)",
        "소속(요청자)",
        "성명(대상자)",
        "사번(대상자)",
        "담당자",
        "상태",
        "상태변경일",
        "최초등록일",
    ]
    ws.append(headers)

    # Header 스타일 + 고정
    header_font = Font(bold=True)
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=1, column=col_idx).font = header_font
    ws.freeze_panes = "A2"

    for post in queryset:
        ws.append([
            post.receipt_number,
            post.title,
            post.user_id,
            post.user_name,
            post.user_branch,
            post.fa,
            post.code,
            post.handler or "-",
            post.status,
            localtime(post.status_updated_at).strftime("%Y-%m-%d %H:%M") if post.status_updated_at else "-",
            localtime(post.created_at).strftime("%Y-%m-%d %H:%M"),
        ])

    # Auto filter
    ws.auto_filter.ref = ws.dimensions

    # 열 너비 자동 조정(상한 적용)
    MAX_W = 50
    PADDING = 2
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[col_letter].width = min(MAX_W, max(10, max_len + PADDING))

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename={filename}'
    wb.save(response)
    return response


# =========================================================
# Admin
# - custom_admin_site 접근 자체가 grade==superuser만 허용(기존 정책)
# =========================================================
@admin.register(Post, site=custom_admin_site)
class PostAdmin(admin.ModelAdmin):
    """
    업무요청(Post) 관리자 페이지
    """

    # -----------------------------
    # List / Filter / Search
    # -----------------------------
    list_display = (
        "get_receipt_number",
        "get_category",
        "get_title",
        "get_user_id",
        "get_user_name",
        "get_user_branch",
        "get_fa",
        "get_code",
        "get_handler",
        "colored_status",
        "get_status_updated_at",
        "get_created_at",
    )
    list_filter = ("status", "handler", "user_branch", "category", "created_at")
    search_fields = ("title", "content", "user_name", "user_id", "fa", "code", "handler")
    ordering = ("-created_at",)
    readonly_fields = ("created_at", "status_updated_at", "receipt_number")

    # -----------------------------
    # Fieldsets
    # -----------------------------
    fieldsets = (
        ("게시글 정보", {"fields": ("receipt_number", "category", "title", "content", "fa", "code")}),
        ("작성자 정보", {"fields": ("user_id", "user_name", "user_branch")}),
        ("담당자 / 상태 관리", {"fields": ("handler", "status", "status_updated_at")}),
        ("기타", {"fields": ("created_at",)}),
    )

    # -----------------------------
    # Actions
    # -----------------------------
    actions = ["export_selected_posts_to_excel"]

    def export_selected_posts_to_excel(self, request, queryset):
        return export_posts_as_excel(queryset, filename="selected_posts.xlsx")

    export_selected_posts_to_excel.short_description = "선택된 게시글을 Excel로 다운로드"

    # -----------------------------
    # Custom URLs (Export All)
    # -----------------------------
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "export-all/",
                self.admin_site.admin_view(self.export_all_posts_view),
                name="export_all_posts_excel",
            ),
        ]
        return custom_urls + urls

    def export_all_posts_view(self, request):
        return export_posts_as_excel(self.model.objects.all(), filename="all_posts.xlsx")

    # -----------------------------
    # UI Helpers
    # -----------------------------
    def colored_status(self, obj):
        color_map = {
            "확인중": "orange",
            "진행중": "green",
            "보완요청": "red",
            "완료": "black",
            "반려": "gray",
        }
        color = color_map.get(obj.status, "black")
        return format_html('<span style="color:{}; font-weight:600;">{}</span>', color, obj.status)

    colored_status.short_description = "상태"

    # -----------------------------
    # Display Labels (한글 컬럼명)
    # -----------------------------
    def get_receipt_number(self, obj): return obj.receipt_number
    get_receipt_number.short_description = "접수번호"

    def get_category(self, obj): return obj.category
    get_category.short_description = "구분"

    def get_title(self, obj): return obj.title
    get_title.short_description = "제목"

    def get_user_id(self, obj): return obj.user_id
    get_user_id.short_description = "사번(요청자)"

    def get_user_name(self, obj): return obj.user_name
    get_user_name.short_description = "성명(요청자)"

    def get_user_branch(self, obj): return obj.user_branch
    get_user_branch.short_description = "소속(요청자)"

    def get_fa(self, obj): return obj.fa
    get_fa.short_description = "성명(대상자)"

    def get_code(self, obj): return obj.code
    get_code.short_description = "사번(대상자)"

    def get_handler(self, obj): return obj.handler
    get_handler.short_description = "담당자"

    def get_status_updated_at(self, obj):
        return localtime(obj.status_updated_at).strftime("%Y-%m-%d %H:%M") if obj.status_updated_at else "-"
    get_status_updated_at.short_description = "상태변경일"

    def get_created_at(self, obj):
        return localtime(obj.created_at).strftime("%Y-%m-%d %H:%M")
    get_created_at.short_description = "최초등록일"


# =========================================================
# Industry Info Admin (Proxy Models)
# - 실제 DB는 support_* 테이블을 사용
# - board에서 운영/검수 화면을 통합 제공
# =========================================================
@admin.register(IndustryArticle, site=custom_admin_site)
class IndustryArticleAdmin(admin.ModelAdmin):
    list_display = (
        "id",
        "title",
        "source_name",
        "topic",
        "published_at",
        "is_active",
        "is_hidden",
    )
    list_filter = ("source_portal", "topic", "is_active", "is_hidden")
    search_fields = ("title", "summary", "source_name", "keyword_query")
    ordering = ("-published_at", "-id")


@admin.register(IndustryUserPreference, site=custom_admin_site)
class IndustryUserPreferenceAdmin(admin.ModelAdmin):
    list_display = (
        "id",
        "user",
        "article",
        "rating",
        "is_bookmarked",
        "is_hidden",
        "updated_at",
    )
    list_filter = ("is_bookmarked", "is_hidden", "is_read")
    search_fields = ("user__id", "article__title")


@admin.register(IndustryRecommendation, site=custom_admin_site)
class IndustryRecommendationAdmin(admin.ModelAdmin):
    list_display = (
        "id",
        "user",
        "article",
        "score",
        "reason_code",
        "model_version",
        "created_at",
    )
    list_filter = ("reason_code", "model_version", "clicked")
    search_fields = ("user__id", "article__title")


@admin.register(IndustryCollectJobLog, site=custom_admin_site)
class IndustryCollectJobLogAdmin(admin.ModelAdmin):
    list_display = (
        "id",
        "source",
        "query",
        "status",
        "fetched_count",
        "inserted_count",
        "error_count",
        "created_at",
    )
    list_filter = ("source", "status")
    search_fields = ("query", "message")


# =========================================================
# WorkTask 업무관리 Admin (Phase 1 신규 추가)
#
# 운영 예외 정책 (worktask.md §17):
#   - 이 Admin 에서는 모든 WorkTask 전체 열람·수정 가능
#   - 일반 뷰(/board/worktasks/)의 owner=request.user 격리와 별개
#   - custom_admin_site 접근 자체가 superuser 전용이므로 추가 제한 불필요
# =========================================================

# ---------------------------------------------------------
# WorkCategory — 업무 분류 마스터
# ---------------------------------------------------------
@admin.register(WorkCategory, site=custom_admin_site)
class WorkCategoryAdmin(admin.ModelAdmin):
    """
    관리자가 직접 업무 분류를 등록·관리.

    초기 등록 필요 데이터 (worktask.md §3.1):
        commission  수수료 업무
        bond        채권·환수
        risk        리스크관리
        biz_dev     제휴영업
        misc        기타
    """
    list_display  = ["code", "label", "sort_order", "is_active"]
    list_editable = ["label", "sort_order", "is_active"]
    ordering      = ["sort_order", "code"]


# ---------------------------------------------------------
# WorkTaskAttachment — WorkTask 하위 인라인
# ---------------------------------------------------------
class WorkTaskAttachmentInline(admin.TabularInline):
    """
    WorkTask 상세에서 첨부파일 인라인 관리.
    Admin 에서는 파일 직접 링크 허용 (운영 목적 예외).
    """
    model           = WorkTaskAttachment
    extra           = 0
    readonly_fields = ["original_name", "uploaded_by", "uploaded_at"]
    fields          = ["file", "original_name", "uploaded_by", "uploaded_at"]
    show_change_link = False


# ---------------------------------------------------------
# WorkTask — 업무 항목 (전체 열람 Admin)
# ---------------------------------------------------------
@admin.register(WorkTask, site=custom_admin_site)
class WorkTaskAdmin(admin.ModelAdmin):
    """
    운영용 전체 열람 Admin.
    ⚠️ 일반 뷰의 owner 격리(owner=request.user)가 적용되지 않는다.
    """

    # List
    list_display  = [
        "id", "owner", "category", "title",
        "status", "priority", "due_date",
        "recurrence_type", "target_ym",
        "is_notified", "created_at",
    ]
    list_filter   = ["status", "category", "recurrence_type", "is_notified"]
    search_fields = ["title", "owner__id", "owner__name"]
    ordering      = ["-created_at"]

    # Detail
    readonly_fields   = ["created_at", "updated_at"]
    raw_id_fields     = ["owner", "template_task"]
    filter_horizontal = ["related_users"]
    inlines           = [WorkTaskAttachmentInline]

    fieldsets = [
        (
            "소유자 / 분류",
            {"fields": ["owner", "category"]},
        ),
        (
            "내용",
            {"fields": ["title", "description", "related_users"]},
        ),
        (
            "일정 / 반복",
            {
                "fields": [
                    "due_date",
                    "recurrence_type", "recurrence_day",
                    "template_task", "target_ym",
                ],
            },
        ),
        (
            "상태 / 우선순위",
            {"fields": ["status", "priority"]},
        ),
        (
            "알림",
            {"fields": ["notify_days_before", "is_notified"]},
        ),
        (
            "감사 (읽기전용)",
            {
                "fields": ["created_at", "updated_at"],
                "classes": ["collapse"],
            },
        ),
    ]


# ---------------------------------------------------------
# WorkTaskAttachment — 단독 조회 Admin
# ---------------------------------------------------------
@admin.register(WorkTaskAttachment, site=custom_admin_site)
class WorkTaskAttachmentAdmin(admin.ModelAdmin):
    """첨부파일 단독 운영 조회."""
    list_display    = ["id", "task", "original_name", "uploaded_by", "uploaded_at"]
    search_fields   = ["original_name", "task__title"]
    raw_id_fields   = ["task", "uploaded_by"]
    readonly_fields = ["uploaded_at"]
    ordering        = ["-uploaded_at"]